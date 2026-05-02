#!/usr/bin/env python3
"""
Twedar Project Progress Tracker - Excel Generator
Generates a comprehensive multi-sheet workbook comparing the project document
requirements against the current implementation status.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# === Color Definitions ===
HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
SUBHEADER_FONT = Font(name="Segoe UI", size=10, bold=True, color="1B2A4A")

DONE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
DONE_FONT = Font(name="Segoe UI", size=10, color="006100")
PROGRESS_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PROGRESS_FONT = Font(name="Segoe UI", size=10, color="9C5700")
NOT_STARTED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NOT_STARTED_FONT = Font(name="Segoe UI", size=10, color="9C0006")

ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

TITLE_FONT = Font(name="Segoe UI", size=16, bold=True, color="1B2A4A")
SUBTITLE_FONT = Font(name="Segoe UI", size=12, bold=False, color="4472C4")
BODY_FONT = Font(name="Segoe UI", size=10, color="333333")
BOLD_BODY = Font(name="Segoe UI", size=10, bold=True, color="333333")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

STATUS_MAP = {
    "Done": (DONE_FILL, DONE_FONT),
    "In Progress": (PROGRESS_FILL, PROGRESS_FONT),
    "Not Started": (NOT_STARTED_FILL, NOT_STARTED_FONT),
    "Not Tested": (NOT_STARTED_FILL, NOT_STARTED_FONT),
    "Passed": (DONE_FILL, DONE_FONT),
    "Failed": (NOT_STARTED_FILL, NOT_STARTED_FONT),
}

COLUMNS = ["ID", "Feature Name", "Description", "Status", "Backend", "Frontend",
           "Implementation Notes", "Remaining Work", "Priority", "Doc Reference"]
COL_WIDTHS = [8, 28, 50, 14, 14, 14, 45, 45, 14, 16]


def style_header_row(ws, row=1):
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def set_col_widths(ws):
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_headers(ws):
    for col_idx, name in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=name)
    style_header_row(ws)
    set_col_widths(ws)
    ws.freeze_panes = "A2"


def add_status_validation(ws, start_row, end_row):
    dv = DataValidation(type="list", formula1='"Done,In Progress,Not Started"', allow_blank=True)
    dv.error = "Please select a valid status"
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)
    for r in range(start_row, end_row + 1):
        dv.add(ws.cell(row=r, column=4))


def write_data_row(ws, row_idx, data):
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    if row_idx % 2 == 0:
        for col_idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = ALT_ROW_FILL

    status = data[3] if len(data) > 3 else ""
    if status in STATUS_MAP:
        fill, font = STATUS_MAP[status]
        status_cell = ws.cell(row=row_idx, column=4)
        status_cell.fill = fill
        status_cell.font = font
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

        be_status = data[4] if len(data) > 4 else ""
        if be_status in STATUS_MAP:
            be_cell = ws.cell(row=row_idx, column=5)
            be_fill, be_font = STATUS_MAP[be_status]
            be_cell.fill = be_fill
            be_cell.font = be_font
            be_cell.alignment = Alignment(horizontal="center", vertical="center")

        fe_status = data[5] if len(data) > 5 else ""
        if fe_status in STATUS_MAP:
            fe_cell = ws.cell(row=row_idx, column=6)
            fe_fill, fe_font = STATUS_MAP[fe_status]
            fe_cell.fill = fe_fill
            fe_cell.font = fe_font
            fe_cell.alignment = Alignment(horizontal="center", vertical="center")


def write_section_header(ws, row_idx, title):
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(COLUMNS))
    cell = ws.cell(row=row_idx, column=1, value=title)
    cell.fill = SUBHEADER_FILL
    cell.font = SUBHEADER_FONT
    cell.alignment = Alignment(vertical="center")
    for col in range(2, len(COLUMNS) + 1):
        ws.cell(row=row_idx, column=col).fill = SUBHEADER_FILL
    return row_idx + 1


def populate_feature_sheet(ws, sections):
    write_headers(ws)
    row = 2
    for section_title, items in sections:
        row = write_section_header(ws, row, section_title)
        for item in items:
            write_data_row(ws, row, item)
            row += 1
        row += 1  # blank separator
    add_status_validation(ws, 2, row)


# ========== DATA DEFINITIONS ==========

AUTH_DATA = [
    ("Registration & Login", [
        ["A-01", "Email/Password Registration", "Users register with email and password. Password hashed with bcrypt. Role selected at registration.", "Done", "Done", "Done", "Better Auth handles registration with accountType field mapped to role. Frontend register page with role selector.", "", "Must Have", "3.2.1"],
        ["A-02", "Email Verification", "System sends verification email after registration. User clicks link to verify.", "Done", "Done", "Done", "Better Auth sendVerificationEmail configured with Nodemailer. Frontend /verify-email and /check-email pages.", "", "Must Have", "3.2.1"],
        ["A-03", "Email/Password Login", "Authenticated login with session cookie creation.", "Done", "Done", "Done", "Better Auth signIn.email. Frontend login page with redirect to role dashboard.", "", "Must Have", "3.4.2 UC2"],
        ["A-04", "Google OAuth (Couples)", "Social login via Google for quick couple signup.", "Done", "Done", "Done", "Better Auth socialProviders.google configured. Frontend redirects to /api/auth/social/google. Callback page handles edge cases.", "", "Must Have", "3.2.1"],
        ["A-05", "Apple OAuth (Couples)", "Social login via Apple for quick couple signup.", "Done", "Done", "Done", "Better Auth socialProviders.apple configured. Frontend callback handles Apple flow.", "", "Should Have", "3.2.1"],
        ["A-06", "Password Reset / Forgot Password", "Email-based password reset flow.", "Done", "Done", "Done", "Better Auth forgetPassword + resetPassword. Frontend forget-password and reset-password pages.", "", "Must Have", "3.2.1"],
    ]),
    ("Session & Security", [
        ["A-07", "HTTP-Only Session Cookies", "Secure session management without client-side token exposure.", "Done", "Done", "Done", "Better Auth configured with httpOnly cookies, 30-day expiry, SameSite lax/none, secure in production.", "", "Must Have", "4.11.1"],
        ["A-08", "RBAC Middleware", "Role-based access control enforced at API level.", "Done", "Done", "Done", "requireAuth + requireRole middleware in auth feature. Frontend AuthGuard component checks session + role.", "", "Must Have", "4.11.4"],
        ["A-09", "Session Validation on App Load", "Check session status on initial page load for routing.", "Done", "Done", "Done", "Frontend useSession hook from better-auth/react. AuthGuard validates on mount.", "", "Must Have", "4.11.1"],
        ["A-10", "401 Global Handling", "Redirect to login on unauthorized responses.", "Done", "Done", "Done", "apiFetch in auth.service.ts checks 401 and redirects to /login.", "", "Must Have", "4.11.1"],
    ]),
    ("Role Management", [
        ["A-11", "Couple Role", "Standard user with access to wedding planning features.", "Done", "Done", "Done", "Role 'couple' created at registration. Couple layout with dashboard shell.", "", "Must Have", "4.11.4"],
        ["A-12", "Vendor Owner Role", "Organization owner with full vendor dashboard access.", "Done", "Done", "In Progress", "Backend auto-creates Organization on vendor registration. Frontend has vendor layout but no org management UI.", "Add organization member list UI, staff management.", "Must Have", "4.11.4"],
        ["A-13", "Vendor Staff Role", "Organization member with limited access (chat, schedule).", "In Progress", "Done", "Not Started", "Better Auth Organization plugin configured with owner/member roles and permissions. No frontend staff view.", "Build vendor staff invitation UI, restricted staff dashboard.", "Should Have", "4.11.4"],
        ["A-14", "Super Admin Role", "Full system access for platform management.", "Done", "Done", "In Progress", "Admin plugin configured. Admin layout exists with vendor moderation. Missing users/reports/settings pages.", "Build admin users list, reports, and settings pages.", "Must Have", "4.11.4"],
        ["A-15", "Content Moderator Role", "Limited admin privileges assigned by Super Admin.", "In Progress", "In Progress", "Not Started", "Admin plugin supports role assignment. No dedicated moderator UI or role differentiation in frontend.", "Add moderator role assignment UI, restricted admin views.", "Could Have", "4.11.4"],
    ]),
    ("Profile Management", [
        ["A-16", "Couple Profile Update", "Couples update personal info, wedding preferences.", "In Progress", "In Progress", "Not Started", "Better Auth user fields exist. No dedicated couple profile/preferences page.", "Build couple profile page with wedding date, theme, location fields.", "Must Have", "3.4.2 UC3"],
        ["A-17", "Vendor Profile Management", "Vendors manage business info, services, pricing.", "Done", "Done", "Done", "Full vendor profile CRUD with setup wizard page.", "", "Must Have", "3.4.2 UC3"],
        ["A-18", "Admin Profile/Settings", "Admin can update system settings.", "Not Started", "Not Started", "Not Started", "No admin settings implementation.", "Build admin settings page and backend config API.", "Could Have", "3.4.2 UC3"],
    ]),
]

VENDOR_DATA = [
    ("Vendor Profile & Listing", [
        ["V-01", "Vendor Registration", "Vendor registers and auto-creates organization.", "Done", "Done", "Done", "Registration with accountType=vendor triggers org creation via databaseHooks.", "", "Must Have", "3.2.2, US-V-001"],
        ["V-02", "Business Profile Creation", "Vendor creates profile with business name, category, description, location.", "Done", "Done", "Done", "POST /api/v1/vendor/profile. Frontend setup wizard with all fields.", "", "Must Have", "3.2.2"],
        ["V-03", "Service & Pricing Packages", "Vendor defines service types and pricing ranges.", "Done", "Done", "Done", "vendor_profiles table has price_range_min/max, category. Setup form includes pricing section.", "", "Must Have", "3.2.2"],
        ["V-04", "Location with Map", "Vendor sets location with latitude/longitude on map.", "Done", "Done", "Done", "Backend stores lat/long with btree_gist index. Frontend Leaflet map picker component.", "", "Must Have", "3.2.2"],
        ["V-05", "Portfolio Upload", "Vendor uploads images/videos showcasing previous work.", "In Progress", "Done", "In Progress", "Backend has portfolio JSONB field in vendor_profiles. Frontend vendor nav has Portfolio link but no dedicated page.", "Build dedicated portfolio page with media grid, upload/delete functionality.", "Must Have", "3.4.2 UC11"],
        ["V-06", "Social Links", "Vendor adds social media links to profile.", "Done", "Done", "Done", "social_links JSONB field in vendor_profiles. Setup form includes social links section.", "", "Could Have", "3.2.2"],
        ["V-07", "Vendor Availability/Schedule", "Vendor sets available dates and working hours.", "Not Started", "Not Started", "Not Started", "No availability/calendar implementation.", "Build availability schema, API endpoints, and calendar UI for vendors.", "Should Have", "3.2.2"],
    ]),
    ("Document & Verification", [
        ["V-08", "Document Upload", "Vendor uploads business license, ID, certifications.", "Done", "Done", "Done", "POST /api/v1/vendor/documents with Supabase storage. Frontend DocumentUpload component.", "", "Must Have", "3.2.3, US-V-002"],
        ["V-09", "Document Delete", "Vendor can remove uploaded documents.", "Done", "Done", "Done", "DELETE /api/v1/vendor/documents/:documentId.", "", "Must Have", "3.2.3"],
        ["V-10", "Submit for Verification", "Vendor submits profile for admin review.", "Done", "Done", "Done", "POST /api/v1/vendor/profile/submit. Status transitions via status-machine.ts.", "", "Must Have", "3.2.3"],
        ["V-11", "Verification Status Display", "Vendor sees their current verification status.", "Done", "Done", "Done", "VendorStatusBanner component shows pending/verified/rejected state.", "", "Must Have", "3.2.3, US-V-003"],
        ["V-12", "Resubmission After Rejection", "Vendor fixes issues and resubmits with admin feedback.", "Done", "Done", "Done", "Status machine allows draft->pending_verification transition after rejection.", "", "Should Have", "3.2.3, US-V-004"],
    ]),
    ("Admin Vendor Actions", [
        ["V-13", "List All Vendors (Admin)", "Admin views all registered vendors with status.", "Done", "Done", "Done", "GET /api/v1/admin/vendors. Admin vendors list page with status badges.", "", "Must Have", "3.4.2 UC14"],
        ["V-14", "Approve Vendor", "Admin approves vendor after document review.", "Done", "Done", "Done", "POST /api/v1/admin/vendors/:id/approve. Admin detail page with approve button.", "", "Must Have", "3.4.2 UC14"],
        ["V-15", "Reject Vendor", "Admin rejects vendor with reason.", "Done", "Done", "Done", "POST /api/v1/admin/vendors/:id/reject with reason. ReasonModal component.", "", "Must Have", "3.4.2 UC14"],
        ["V-16", "Suspend Vendor", "Admin suspends vendor for policy violations.", "Done", "Done", "Done", "POST /api/v1/admin/vendors/:id/suspend. ConfirmDialog component.", "", "Must Have", "4.11.3"],
        ["V-17", "Reinstate Vendor", "Admin reactivates suspended vendor.", "Done", "Done", "Done", "POST /api/v1/admin/vendors/:id/reinstate.", "", "Should Have", "4.11.3"],
        ["V-18", "Deactivate Vendor", "Admin permanently deactivates vendor.", "Done", "Done", "Done", "POST /api/v1/admin/vendors/:id/deactivate.", "", "Should Have", "4.11.3"],
    ]),
    ("Public Vendor Discovery", [
        ["V-19", "Public Vendor Listing", "Couples browse verified vendors with search/filter.", "Done", "Done", "Done", "GET /api/v1/vendors with category, location, search, pagination, sort filters. Frontend /vendors page.", "", "Must Have", "3.2.2, US-C-004"],
        ["V-20", "Vendor Detail Page", "Couples view full vendor profile, portfolio, reviews.", "Done", "Done", "Done", "GET /api/v1/vendors/:id. Frontend /vendors/[id] page with full profile.", "", "Must Have", "3.4.2 UC8"],
        ["V-21", "Vendor Filtering by Category", "Filter vendors by service type (photographer, caterer, etc.).", "Done", "Done", "Done", "category query param in public vendor listing API.", "", "Must Have", "US-C-004"],
        ["V-22", "Vendor Filtering by Location", "Filter vendors by city/area.", "Done", "Done", "Done", "location query param in public vendor listing API.", "", "Must Have", "US-C-004"],
        ["V-23", "Vendor Filtering by Verification Status", "Show only verified vendors to couples.", "Done", "Done", "Done", "Public API only returns verified vendors by default.", "", "Must Have", "US-C-007"],
        ["V-24", "Save Vendors to Favorites", "Couple saves vendors for later comparison.", "Not Started", "Not Started", "Not Started", "No favorites/saved vendors implementation.", "Build favorites table, API endpoints, and heart/save button in vendor cards.", "Could Have", "US-C-008"],
    ]),
]

REALTIME_DATA = [
    ("Socket.IO Infrastructure", [
        ["R-01", "Socket.IO Server Setup", "WebSocket server with authentication.", "Done", "Done", "Done", "createSocketServer in socket-server.ts with cookie-based auth middleware.", "", "Must Have", "3.2.6"],
        ["R-02", "Socket.IO Client Connection", "Frontend establishes authenticated WS connection.", "Done", "Done", "Done", "socket-client.ts singleton with withCredentials. SocketProvider manages lifecycle.", "", "Must Have", "3.2.6"],
        ["R-03", "Room-based Architecture", "Users join/leave rooms for targeted messaging.", "Done", "Done", "Done", "Users join user:{id} room on connect. Chat rooms joined/left via hooks.", "", "Must Have", "3.2.6"],
    ]),
    ("Chat / Messaging", [
        ["R-04", "Create Conversation", "Couple initiates conversation with vendor.", "Done", "Done", "Done", "POST /api/v1/conversations with participantId. Frontend starts conversation from vendor detail page.", "", "Must Have", "3.2.6, US-C-015"],
        ["R-05", "List Conversations", "User sees all their active conversations.", "Done", "Done", "Done", "GET /api/v1/conversations. ChatList component shows conversations.", "", "Must Have", "3.2.6"],
        ["R-06", "Send Text Message", "Real-time text message delivery.", "Done", "Done", "Done", "chat:send event via Socket.IO. Message persisted in chat_message table.", "", "Must Have", "3.2.6"],
        ["R-07", "Message History", "Load previous messages for a conversation.", "Done", "Done", "Done", "GET /api/v1/conversations/:id/messages. useChat hook loads on room join.", "", "Must Have", "3.2.6"],
        ["R-08", "Typing Indicators", "Show when other user is typing.", "Done", "Done", "Done", "chat:typing and chat:stopTyping events. UI shows typing state.", "", "Should Have", "Scenario 4"],
        ["R-09", "Read Receipts", "Mark messages as read, show double ticks.", "Done", "Done", "Done", "chat:markRead event updates isRead. UI shows read status.", "", "Should Have", "Scenario 4"],
        ["R-10", "File Attachments in Chat", "Send contracts, invoices, images in chat.", "Not Started", "Not Started", "Not Started", "No file attachment support in chat messages.", "Add file upload endpoint for chat, message type field, file preview in chat UI.", "Should Have", "US-C-017"],
    ]),
    ("Notifications", [
        ["R-11", "In-App Notifications", "Real-time notification delivery to connected users.", "Done", "Done", "Done", "notification:new event via Socket.IO. NotificationDropdown shows list.", "", "Must Have", "3.2.6"],
        ["R-12", "Notification REST API", "Fetch notifications, mark as read.", "Done", "Done", "Done", "GET /api/v1/notifications, PATCH /:id/read, PATCH /read-all.", "", "Must Have", "3.2.6"],
        ["R-13", "Notification Bell/Badge", "Visual indicator of unread notifications.", "Done", "Done", "Done", "NotificationBell component with unread count badge.", "", "Must Have", "3.2.6"],
        ["R-14", "Push Notifications (Mobile/Browser)", "Notifications when app is closed or in background.", "Not Started", "Not Started", "Not Started", "Only in-app socket notifications exist.", "Integrate Web Push API / Firebase Cloud Messaging for background notifications.", "Should Have", "US-C-018"],
        ["R-15", "Email Notifications", "Email alerts for critical events (booking, payment).", "Not Started", "Not Started", "Not Started", "Email service exists but only for auth flows.", "Extend email service for booking confirmations, payment receipts, reminders.", "Could Have", "3.2.6"],
    ]),
    ("Presence", [
        ["R-16", "User Online/Offline Status", "Show if user is currently online.", "Done", "Done", "Done", "presence:update event. PresenceDot component shows green/gray dot.", "", "Could Have", "3.2.6"],
    ]),
]

BUDGET_DATA = [
    ("Budget Planning", [
        ["B-01", "Create Wedding Budget", "Couple defines total wedding budget amount.", "Not Started", "Not Started", "Not Started", "No budget implementation exists.", "Create budget table, POST /api/v1/budgets endpoint, budget creation UI page.", "Must Have", "3.2.5, US-C-009"],
        ["B-02", "Budget Category Allocation", "Allocate budget into categories (venue, catering, photo, etc.).", "Not Started", "Not Started", "Not Started", "No budget categories implementation.", "Create budget_categories table, CRUD API, category allocation UI with sliders/inputs.", "Must Have", "3.2.5, US-C-009"],
        ["B-03", "AI-Suggested Budget Allocation", "System suggests category distribution based on total budget and wedding type.", "Not Started", "Not Started", "Not Started", "No AI budget suggestion implementation.", "Build recommendation logic for budget distribution, API endpoint, suggestion UI.", "Should Have", "3.2.5, US-C-010"],
    ]),
    ("Expense Tracking", [
        ["B-04", "Add Expense Entry", "Record individual expenses against budget categories.", "Not Started", "Not Started", "Not Started", "No expense implementation.", "Create expenses table, POST /api/v1/expenses endpoint, expense form UI.", "Must Have", "3.2.5, US-C-011"],
        ["B-05", "Real-Time Budget Totals", "Dashboard shows live spending vs allocated amounts.", "Not Started", "Not Started", "Not Started", "No budget dashboard.", "Build budget summary calculations, real-time update UI with progress bars.", "Must Have", "3.2.5, US-C-011"],
        ["B-06", "Budget Alerts & Limits", "Warnings when spending approaches/exceeds category limits.", "Not Started", "Not Started", "Not Started", "No alert system for budgets.", "Add threshold logic, notification triggers, alert banners in budget UI.", "Should Have", "3.2.5, US-C-012"],
        ["B-07", "Receipt Upload", "Attach receipt images to expense entries.", "Not Started", "Not Started", "Not Started", "No receipt upload functionality.", "Add receipt storage (Supabase), upload API, receipt preview in expense detail.", "Could Have", "US-C-013"],
        ["B-08", "Budget Summary Reports", "Generate exportable budget reports.", "Not Started", "Not Started", "Not Started", "No reporting functionality.", "Build report generation API, PDF/CSV export, report view page.", "Could Have", "US-C-014"],
    ]),
]

AI_DATA = [
    ("Recommendation Engine", [
        ["AI-01", "Content-Based Filtering", "Recommend vendors using attributes (category, location, price, tags).", "Not Started", "Not Started", "Not Started", "No AI/ML implementation. vendor_profiles has category, location, price fields for future use.", "Build vendor vectorization, similarity scoring, content-based recommendation API.", "Must Have", "3.2.4, 4.6.3"],
        ["AI-02", "Collaborative Filtering", "Recommend vendors based on similar couples' behavior.", "Not Started", "Not Started", "Not Started", "No interaction tracking or matrix factorization.", "Build user-vendor interaction tracking, matrix factorization model, training pipeline.", "Should Have", "3.2.4, 4.6.3"],
        ["AI-03", "Hybrid Weighted Approach", "Combine content + collaborative with adaptive weights based on data maturity.", "Not Started", "Not Started", "Not Started", "No hybrid engine.", "Implement weighted scoring: 100% content early, shift to 70% collaborative at maturity.", "Should Have", "3.2.4, 4.6.3"],
        ["AI-04", "Cold-Start Handling", "Ensure meaningful recommendations with zero interaction data.", "Not Started", "Not Started", "Not Started", "No cold-start strategy implemented.", "Default to content-based only. Use vendor metadata + couple preferences for initial scoring.", "Must Have", "3.2.4, 4.6.3"],
        ["AI-05", "User Preference Collection", "Gather couple preferences (theme, location, budget, style).", "Not Started", "Not Started", "Not Started", "No preference collection UI or storage.", "Build preference wizard during couple onboarding, store in user_preferences table.", "Must Have", "3.2.4"],
    ]),
    ("Caching & Performance", [
        ["AI-06", "Redis Recommendation Cache", "Cache computed recommendations for 24h per user.", "Not Started", "Not Started", "Not Started", "No Redis setup in project.", "Add Redis dependency, cache layer for recommendations, TTL-based invalidation.", "Must Have", "4.6.3"],
        ["AI-07", "Cache Invalidation Rules", "Invalidate cache on preference update, new booking, or vendor profile change.", "Not Started", "Not Started", "Not Started", "No cache infrastructure.", "Implement cache key strategy, event-driven invalidation hooks.", "Should Have", "4.6.3"],
        ["AI-08", "Sub-2-Second Response", "Recommendation results within 2 seconds.", "Not Started", "Not Started", "Not Started", "No performance benchmarks.", "Optimize query paths, pre-compute scores, use Redis for instant retrieval.", "Must Have", "3.3.1"],
    ]),
    ("Recommendation UI", [
        ["AI-09", "Recommendation Badges", "Highlight top-matched vendors with AI badges.", "Not Started", "Not Started", "Not Started", "No recommendation UI elements.", "Add badge component, match score display, 'Recommended for You' section.", "Should Have", "Scenario 5"],
        ["AI-10", "Personalized Vendor Search", "AI-ranked search results based on couple profile.", "Not Started", "Not Started", "Not Started", "Current search is basic filter-only.", "Integrate AI scoring into vendor search pipeline, re-rank results.", "Must Have", "Scenario 5, US-C-006"],
    ]),
]

BOOKING_DATA = [
    ("Booking Workflow", [
        ["BK-01", "Send Booking Request", "Couple submits booking request with date and requirements.", "Not Started", "Not Started", "Not Started", "No booking implementation.", "Create bookings table, POST /api/v1/bookings endpoint, booking request form UI.", "Must Have", "3.4.2 UC9, US-C-016"],
        ["BK-02", "Vendor Reviews Booking", "Vendor sees incoming booking requests on dashboard.", "Not Started", "Not Started", "Not Started", "No booking list for vendors.", "Build vendor bookings list API, incoming requests page, accept/decline UI.", "Must Have", "3.4.2 UC9"],
        ["BK-03", "Accept/Decline Booking", "Vendor accepts or declines with optional message.", "Not Started", "Not Started", "Not Started", "No booking status workflow.", "Add booking status machine (pending/accepted/declined/completed), status update API.", "Must Have", "3.4.2 UC9"],
        ["BK-04", "Booking Status Tracking", "Both parties see current booking status.", "Not Started", "Not Started", "Not Started", "No booking status UI.", "Build booking detail page with status timeline, notifications on status change.", "Must Have", "3.4.2 UC9"],
        ["BK-05", "Booking Confirmation Notification", "Notify couple when vendor accepts/declines.", "Not Started", "Not Started", "Not Started", "Notification system exists but no booking triggers.", "Wire booking status changes to notification service.", "Must Have", "US-C-018"],
    ]),
    ("Scheduling & Calendar", [
        ["BK-06", "Vendor Availability Calendar", "Vendors set available/blocked dates.", "Not Started", "Not Started", "Not Started", "No calendar implementation.", "Build availability_slots table, calendar API, interactive calendar UI (vendor side).", "Should Have", "3.2.2"],
        ["BK-07", "Date Conflict Detection", "Prevent double-booking on same date.", "Not Started", "Not Started", "Not Started", "No conflict logic.", "Add unique constraint on vendor+date, validation in booking API.", "Should Have", "3.2.2"],
        ["BK-08", "Booking History", "List of past and upcoming bookings for both roles.", "Not Started", "Not Started", "Not Started", "No booking history.", "Build booking list API with status filters, history pages for couple and vendor.", "Must Have", "US-C-021"],
    ]),
]

PAYMENT_DATA = [
    ("Payment Integration", [
        ["P-01", "Chapa Payment Gateway", "Integrate Chapa API for accepting payments.", "Not Started", "Not Started", "Not Started", "No payment dependencies or code.", "Add Chapa SDK, payment initiation API, webhook handler, payment model.", "Must Have", "3.2.7"],
        ["P-02", "Telebirr Integration", "Support Telebirr mobile money payments.", "Not Started", "Not Started", "Not Started", "No Telebirr integration.", "Research Telebirr API, implement payment adapter, add to checkout flow.", "Should Have", "3.2.7"],
        ["P-03", "CBE Birr Integration", "Support CBE Birr payments.", "Not Started", "Not Started", "Not Started", "No CBE Birr integration.", "Research CBE Birr API, implement payment adapter.", "Could Have", "3.2.7"],
    ]),
    ("Payment Workflow", [
        ["P-04", "Advance Deposit Payment", "Couple makes deposit to secure booking.", "Not Started", "Not Started", "Not Started", "No deposit workflow.", "Build deposit flow: initiate -> redirect to gateway -> confirm via webhook -> update booking.", "Must Have", "3.2.7, US-C-020"],
        ["P-05", "Final Payment", "Complete remaining payment after service delivery.", "Not Started", "Not Started", "Not Started", "No final payment flow.", "Build final payment trigger, amount calculation (total - deposit), payment confirmation.", "Must Have", "3.2.7"],
        ["P-06", "Escrow-Style Workflow", "Hold funds until service is delivered/confirmed.", "Not Started", "Not Started", "Not Started", "No escrow mechanism.", "Design escrow state machine, hold/release/refund flows, dispute handling.", "Should Have", "3.2.7"],
        ["P-07", "Payment History", "View all transactions linked to bookings.", "Not Started", "Not Started", "Not Started", "No payment records.", "Build payments table, transaction history API, payment history page.", "Must Have", "3.2.7, US-C-021"],
        ["P-08", "Payment-Expense Linkage", "Auto-record payments as expenses in budget.", "Not Started", "Not Started", "Not Started", "No budget-payment integration.", "Link payment confirmation to expense creation, auto-categorize by vendor type.", "Should Have", "3.2.7"],
    ]),
]

REVIEW_DATA = [
    ("Review Submission", [
        ["RV-01", "Submit Rating & Review", "Couple rates vendor with stars (1-5) and text comment after completed booking.", "Not Started", "Not Started", "Not Started", "vendor_profiles has rating and review_count fields but no review submission system.", "Create reviews table, POST /api/v1/reviews endpoint, review form UI on completed bookings.", "Must Have", "3.2.8, US-C-022"],
        ["RV-02", "One Review Per Booking", "Prevent duplicate reviews for same booking.", "Not Started", "Not Started", "Not Started", "No review logic.", "Add unique constraint on booking_id in reviews table, check in API.", "Must Have", "3.2.8, Scenario 3"],
        ["RV-03", "Photo Upload with Review", "Couple uploads wedding photos as visual feedback.", "Not Started", "Not Started", "Not Started", "No review photo upload.", "Add review_photos table, image upload to Supabase, photo gallery in review display.", "Could Have", "US-C-023"],
    ]),
    ("Review Management", [
        ["RV-04", "Vendor Rating Aggregation", "Auto-calculate average vendor rating from reviews.", "Not Started", "Not Started", "Not Started", "Fields exist in vendor_profiles but no aggregation logic.", "Build trigger/service to recalculate avg rating on new review, update vendor record.", "Must Have", "3.2.8, Scenario 3"],
        ["RV-05", "Admin Review Moderation", "Admin approves or removes inappropriate reviews.", "Not Started", "Not Started", "Not Started", "No review moderation.", "Add isApproved field, admin review queue API, moderation UI page.", "Must Have", "3.4.2 UC15"],
        ["RV-06", "Review Display on Vendor Profile", "Show approved reviews on public vendor page.", "Not Started", "Not Started", "Not Started", "Vendor detail page exists but shows no reviews.", "Fetch reviews for vendor, display review cards with stars, text, photos, date.", "Must Have", "3.2.8"],
        ["RV-07", "Review Feeds AI Recommendations", "Use review data as signal for collaborative filtering.", "Not Started", "Not Started", "Not Started", "No AI integration.", "Include review scores in recommendation scoring matrix.", "Should Have", "3.2.4, 4.6.3"],
        ["RV-08", "Vendor Notification on New Review", "Notify vendor when they receive a review.", "Not Started", "Not Started", "Not Started", "Notification system exists but no review trigger.", "Wire review creation to notification service.", "Should Have", "Scenario 3"],
    ]),
]

ADMIN_DATA = [
    ("User Management", [
        ["AD-01", "View All Users", "Admin sees all registered users (couples, vendors, admins).", "In Progress", "In Progress", "Not Started", "Better Auth Admin plugin provides user management API. No custom admin UI.", "Build admin users list page with role filter, search, pagination.", "Must Have", "US-A-001"],
        ["AD-02", "Suspend/Ban User", "Admin suspends or bans accounts violating policies.", "In Progress", "Done", "Not Started", "Admin plugin has ban/suspend capabilities. No frontend UI.", "Build user action buttons, confirmation dialogs, banned state display.", "Must Have", "US-A-002"],
        ["AD-03", "Reactivate User", "Admin reactivates suspended accounts after review.", "In Progress", "Done", "Not Started", "Admin plugin supports reactivation. No frontend UI.", "Add reactivate button in user detail, confirmation flow.", "Should Have", "US-A-003"],
        ["AD-04", "Impersonate User Session", "Admin views platform as specific user for troubleshooting.", "In Progress", "Done", "Not Started", "Admin plugin supports impersonation with audit logging. No UI.", "Build impersonation button, session swap UI, exit-impersonation banner.", "Could Have", "US-A-004"],
    ]),
    ("Vendor Verification (Admin)", [
        ["AD-05", "Vendor Verification Dashboard", "Dedicated panel for reviewing vendor applications.", "Done", "Done", "Done", "Full implementation: list vendors, view documents, approve/reject/suspend.", "", "Must Have", "3.2.3"],
        ["AD-06", "Document Review Interface", "View uploaded vendor documents inline.", "Done", "Done", "Done", "Admin vendor detail page shows document list with download links.", "", "Must Have", "3.2.3"],
    ]),
    ("Analytics & Reporting", [
        ["AD-07", "Platform Analytics Dashboard", "Overview of user counts, vendor stats, booking volume.", "Not Started", "Not Started", "Not Started", "No analytics implementation.", "Build analytics queries, stats API endpoint, dashboard with charts.", "Should Have", "4.6.8"],
        ["AD-08", "Revenue Reports", "Track platform revenue from payments.", "Not Started", "Not Started", "Not Started", "No revenue tracking.", "Build revenue aggregation queries, report generation, export functionality.", "Could Have", "4.6.8"],
        ["AD-09", "User Activity Reports", "Monitor engagement, registration trends.", "Not Started", "Not Started", "Not Started", "No activity tracking.", "Build activity logging, trend calculations, report UI.", "Could Have", "4.6.8"],
        ["AD-10", "System Health Monitoring", "Uptime, response times, error rates.", "Not Started", "Not Started", "Not Started", "No monitoring setup.", "Integrate monitoring service (e.g., health endpoint exists), build status page.", "Should Have", "3.3.2"],
    ]),
    ("Content Moderation", [
        ["AD-11", "Review Moderation Queue", "Admin reviews and approves/rejects user reviews.", "Not Started", "Not Started", "Not Started", "No review system built yet.", "Build with review system: moderation queue, approve/reject actions.", "Must Have", "3.4.2 UC15"],
        ["AD-12", "Content Moderator Role", "Assign limited admin privileges to moderators.", "Not Started", "Not Started", "Not Started", "Admin plugin supports roles but no moderator distinction in UI.", "Add role assignment UI, moderator-specific nav/permissions.", "Could Have", "4.11.3"],
    ]),
]

COUPLE_DATA = [
    ("Wedding Project", [
        ["C-01", "Create Wedding Plan", "Initialize wedding with date, location, theme, guest estimate.", "Not Started", "Not Started", "Not Started", "No wedding project entity or API.", "Create wedding_projects table, CRUD API, wedding plan setup wizard.", "Must Have", "3.4.2 UC4, US-C-002"],
        ["C-02", "Wedding Profile Customization", "Set and update wedding preferences at any time.", "Not Started", "Not Started", "Not Started", "No preferences UI or storage.", "Build preferences page with theme, style, and location fields.", "Must Have", "US-C-003"],
        ["C-03", "Wedding Dashboard", "Central overview of planning progress.", "In Progress", "In Progress", "In Progress", "Couple dashboard exists with placeholder stat cards (Budget, Checklist, Guests show 'ready: false').", "Connect dashboard stats to real data from budget, checklist, and guest list modules.", "Must Have", "3.4.1"],
    ]),
    ("Checklist & Timeline", [
        ["C-04", "Wedding Checklist", "Pre-built and custom task checklist for planning.", "Not Started", "Not Started", "Not Started", "Couple nav has 'Checklist' link but no page or API.", "Build checklist_items table, CRUD API, checklist page with categories and deadlines.", "Must Have", "1.5.1"],
        ["C-05", "Task Deadlines & Reminders", "Set due dates for checklist items with reminders.", "Not Started", "Not Started", "Not Started", "No reminder/scheduling system.", "Add due_date field, reminder notification triggers, overdue highlighting.", "Should Have", "1.5.1"],
        ["C-06", "Planning Timeline/Milestones", "Visual timeline of key milestones.", "Not Started", "Not Started", "Not Started", "No timeline implementation.", "Build milestone model, timeline visualization component.", "Could Have", "1.5.1"],
    ]),
    ("Guest Management", [
        ["C-07", "Guest List", "Manage invited guests with contact details.", "Not Started", "Not Started", "Not Started", "Couple nav has 'Guest List' link but no page or API.", "Build guests table, CRUD API, guest list page with add/edit/delete.", "Must Have", "1.5.1"],
        ["C-08", "RSVP Tracking", "Track guest responses (attending/not attending/pending).", "Not Started", "Not Started", "Not Started", "No RSVP system.", "Add RSVP status field, guest-facing RSVP page/link, status tracking UI.", "Should Have", "1.5.1"],
        ["C-09", "Guest Count & Seating", "Track total guests and optional seating arrangement.", "Not Started", "Not Started", "Not Started", "No seating management.", "Build seating plan model, table assignment UI.", "Could Have", "1.5.1"],
    ]),
    ("Discovery & Favorites", [
        ["C-10", "Vendor Favorites List", "Save vendors to compare later.", "Not Started", "Not Started", "Not Started", "No favorites implementation.", "Build favorites table, toggle API, favorites page, heart icons on vendor cards.", "Could Have", "US-C-008"],
        ["C-11", "Vendor Comparison", "Side-by-side comparison of saved vendors.", "Not Started", "Not Started", "Not Started", "No comparison feature.", "Build comparison view with key metrics (price, rating, location) side by side.", "Could Have", "US-C-005"],
    ]),
]

TESTING_DATA = [
    ("Unit Tests - Auth Module", [
        ["T-01", "Register with valid credentials", "POST /api/auth/sign-up with valid email, password, and accountType returns 201 and creates user record.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: valid registration creates user with correct role, hashed password, emailVerified=false.", "Must Have", "3.2.1"],
        ["T-02", "Register with duplicate email", "POST /api/auth/sign-up with existing email returns 409 Conflict.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: duplicate email returns 409 with appropriate error message.", "Must Have", "3.2.1"],
        ["T-03", "Register with invalid data", "POST /api/auth/sign-up with missing/invalid fields returns 400.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: missing email, short password, invalid role all return 400.", "Must Have", "3.2.1"],
        ["T-04", "Login with correct credentials", "POST /api/auth/sign-in with valid email/password returns session cookie.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: correct login sets httpOnly cookie and returns user data.", "Must Have", "3.4.2 UC2"],
        ["T-05", "Login with wrong password", "POST /api/auth/sign-in with incorrect password returns 401.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: wrong password returns 401 Unauthorized.", "Must Have", "3.4.2 UC2"],
        ["T-06", "Login with unverified email", "POST /api/auth/sign-in before email verification returns appropriate error.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: unverified user login blocked with verification prompt.", "Must Have", "3.2.1"],
        ["T-07", "Session validation middleware", "requireAuth middleware blocks unauthenticated requests with 401.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: request without cookie returns 401; valid cookie passes.", "Must Have", "4.11.1"],
        ["T-08", "Role-based middleware", "requireRole('vendor') blocks couple users with 403.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: couple accessing vendor routes gets 403; vendor passes.", "Must Have", "4.11.4"],
        ["T-09", "Password reset request", "Forgot password sends reset email with valid token.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: valid email triggers email send; invalid email returns 404.", "Must Have", "3.2.1"],
        ["T-10", "Password reset completion", "Reset password with valid token updates password hash.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: valid token + new password updates hash; expired token fails.", "Must Have", "3.2.1"],
    ]),
    ("Unit Tests - Vendor Module", [
        ["T-11", "Create vendor profile", "POST /api/v1/vendor/profile with valid data creates profile.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: valid vendor profile creation returns 201 with profile data.", "Must Have", "3.2.2"],
        ["T-12", "Create duplicate profile", "POST /api/v1/vendor/profile when profile exists returns 409.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: second profile creation attempt returns conflict error.", "Must Have", "3.2.2"],
        ["T-13", "Update vendor profile", "PATCH /api/v1/vendor/profile updates fields correctly.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: partial update changes only specified fields, returns updated profile.", "Must Have", "3.2.2"],
        ["T-14", "Get vendor profile", "GET /api/v1/vendor/profile returns current user's profile.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: authenticated vendor gets own profile with all fields.", "Must Have", "3.2.2"],
        ["T-15", "Upload vendor document", "POST /api/v1/vendor/documents uploads to Supabase and creates record.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: file upload creates document record with storage URL.", "Must Have", "3.2.3"],
        ["T-16", "Delete vendor document", "DELETE /api/v1/vendor/documents/:id removes from storage and DB.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: delete removes storage file and DB record; 404 for invalid ID.", "Must Have", "3.2.3"],
        ["T-17", "Submit for verification", "POST /api/v1/vendor/profile/submit transitions status to pending_verification.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: draft status transitions to pending; already pending returns error.", "Must Have", "3.2.3"],
        ["T-18", "Vendor status machine transitions", "Status machine enforces valid transitions only.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: valid transitions succeed; invalid transitions throw error (e.g. draft -> verified).", "Must Have", "3.2.3"],
        ["T-19", "Public vendor listing filters", "GET /api/v1/vendors with filters returns correct subset.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: category filter, location filter, search, pagination all work correctly.", "Must Have", "US-C-004"],
        ["T-20", "Public vendor detail", "GET /api/v1/vendors/:id returns full profile for verified vendor.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: valid ID returns profile; unverified vendor returns 404.", "Must Have", "3.4.2 UC8"],
    ]),
    ("Unit Tests - Admin Vendor Module", [
        ["T-21", "Admin list vendors", "GET /api/v1/admin/vendors returns all vendors with statuses.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: admin gets full vendor list; non-admin gets 403.", "Must Have", "US-A-005"],
        ["T-22", "Admin approve vendor", "POST /api/v1/admin/vendors/:id/approve changes status to verified.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: pending vendor transitions to verified; non-pending fails.", "Must Have", "3.4.2 UC14"],
        ["T-23", "Admin reject vendor", "POST /api/v1/admin/vendors/:id/reject with reason changes status.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: rejection stores reason; missing reason returns 400.", "Must Have", "3.4.2 UC14"],
        ["T-24", "Admin suspend vendor", "POST /api/v1/admin/vendors/:id/suspend deactivates listing.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: verified vendor transitions to suspended; public listing excludes them.", "Must Have", "4.11.3"],
        ["T-25", "Admin reinstate vendor", "POST /api/v1/admin/vendors/:id/reinstate reactivates vendor.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: suspended vendor transitions back to verified.", "Should Have", "4.11.3"],
    ]),
    ("Unit Tests - Realtime/Chat Module", [
        ["T-26", "Create conversation", "POST /api/v1/conversations creates conversation between two users.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: valid participantId creates conversation; duplicate pair returns existing.", "Must Have", "3.2.6"],
        ["T-27", "List user conversations", "GET /api/v1/conversations returns user's conversations.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: returns only conversations where user is participant.", "Must Have", "3.2.6"],
        ["T-28", "Get conversation messages", "GET /api/v1/conversations/:id/messages returns message history.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: returns messages in chronological order; unauthorized user gets 403.", "Must Have", "3.2.6"],
        ["T-29", "Socket authentication", "Socket.IO connection requires valid session cookie.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: connection without cookie rejected; valid cookie allows connection.", "Must Have", "3.2.6"],
        ["T-30", "Send message via socket", "chat:send event persists message and broadcasts to room.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: emit chat:send creates DB record and partner receives chat:newMessage.", "Must Have", "3.2.6"],
        ["T-31", "Mark message as read", "chat:markRead updates message isRead field.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: markRead updates DB; sender receives chat:messageRead event.", "Should Have", "Scenario 4"],
    ]),
    ("Unit Tests - Notification Module", [
        ["T-32", "Get notifications", "GET /api/v1/notifications returns user's notifications.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: returns only current user's notifications ordered by date.", "Must Have", "3.2.6"],
        ["T-33", "Mark notification read", "PATCH /api/v1/notifications/:id/read updates read status.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: updates isRead to true; 404 for invalid ID.", "Must Have", "3.2.6"],
        ["T-34", "Mark all notifications read", "PATCH /api/v1/notifications/read-all updates all user notifications.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: all user notifications marked read; other users unaffected.", "Must Have", "3.2.6"],
        ["T-35", "Real-time notification delivery", "notification:new socket event delivered to target user.", "Not Tested", "Not Tested", "Not Tested", "No test suite exists.", "Write Jest test: notification creation emits event to correct user room.", "Must Have", "3.2.6"],
    ]),
    ("Unit Tests - Budget Module (Future)", [
        ["T-36", "Create budget", "POST /api/v1/budgets creates budget with total amount.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for budget creation.", "Must Have", "3.2.5"],
        ["T-37", "Add budget category", "POST /api/v1/budgets/:id/categories adds category with allocation.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for category allocation.", "Must Have", "3.2.5"],
        ["T-38", "Add expense entry", "POST /api/v1/expenses records expense against category.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for expense creation.", "Must Have", "3.2.5"],
        ["T-39", "Budget total calculation", "Budget summary correctly sums expenses per category.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for aggregation accuracy.", "Must Have", "3.2.5"],
        ["T-40", "Budget alert trigger", "System triggers alert when expense exceeds 80% of category limit.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for threshold detection.", "Should Have", "US-C-012"],
    ]),
    ("Unit Tests - Booking Module (Future)", [
        ["T-41", "Create booking request", "POST /api/v1/bookings creates pending booking.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for booking creation.", "Must Have", "3.4.2 UC9"],
        ["T-42", "Accept booking", "PATCH /api/v1/bookings/:id/accept transitions to accepted.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for acceptance workflow.", "Must Have", "3.4.2 UC9"],
        ["T-43", "Decline booking", "PATCH /api/v1/bookings/:id/decline transitions to declined.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for decline workflow.", "Must Have", "3.4.2 UC9"],
        ["T-44", "Prevent double booking", "Booking on same date for same vendor returns conflict.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for conflict detection.", "Should Have", "3.2.2"],
        ["T-45", "Complete booking", "PATCH /api/v1/bookings/:id/complete marks as done.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for completion flow.", "Must Have", "3.4.2 UC9"],
    ]),
    ("Unit Tests - Payment Module (Future)", [
        ["T-46", "Initiate payment", "POST /api/v1/payments/initiate calls Chapa API and returns checkout URL.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test with mocked Chapa API.", "Must Have", "3.2.7"],
        ["T-47", "Payment webhook verification", "POST /api/v1/payments/webhook validates Chapa signature.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for valid/invalid webhook signatures.", "Must Have", "3.2.7"],
        ["T-48", "Payment status update", "Webhook success updates payment record and booking status.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for status transition on payment confirm.", "Must Have", "3.2.7"],
        ["T-49", "Payment history retrieval", "GET /api/v1/payments returns user's transaction history.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for correct user scoping.", "Must Have", "US-C-021"],
    ]),
    ("Unit Tests - Review Module (Future)", [
        ["T-50", "Submit review", "POST /api/v1/reviews creates review for completed booking.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for review creation.", "Must Have", "3.2.8"],
        ["T-51", "Prevent duplicate review", "POST /api/v1/reviews for same booking returns 409.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for uniqueness constraint.", "Must Have", "Scenario 3"],
        ["T-52", "Review before booking complete", "POST /api/v1/reviews for non-completed booking returns 400.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for status validation.", "Must Have", "3.2.8"],
        ["T-53", "Rating aggregation update", "New review recalculates vendor average rating.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for avg calculation accuracy.", "Must Have", "Scenario 3"],
        ["T-54", "Admin moderate review", "PATCH /api/v1/admin/reviews/:id/approve changes isApproved.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write Jest test for moderation workflow.", "Must Have", "3.4.2 UC15"],
    ]),
    ("Unit Tests - AI Recommendation (Future)", [
        ["T-55", "Content-based recommendation", "Recommendation API returns vendors ranked by attribute similarity.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write test for correct ranking by preferences.", "Must Have", "3.2.4"],
        ["T-56", "Cold-start fallback", "System returns content-based results when no interaction history exists.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write test for graceful cold-start handling.", "Must Have", "4.6.3"],
        ["T-57", "Cache hit returns fast", "Cached recommendations retrieved without AI computation.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write test: second call uses cache, response < 100ms.", "Must Have", "4.6.3"],
        ["T-58", "Cache invalidation on preference change", "User preference update clears recommendation cache.", "Not Tested", "Not Tested", "Not Tested", "Module not implemented yet.", "Implement module first, then write test for cache key deletion on update.", "Should Have", "4.6.3"],
    ]),
    ("Integration Tests - End-to-End Flows", [
        ["T-59", "Full registration flow", "Register -> verify email -> login -> access dashboard.", "Not Tested", "Not Tested", "Not Tested", "No integration tests.", "Write Supertest flow: register, simulate email verify, login, call protected endpoint.", "Must Have", "Scenario 1"],
        ["T-60", "Vendor onboarding flow", "Register as vendor -> create profile -> upload docs -> submit -> admin approves.", "Not Tested", "Not Tested", "Not Tested", "No integration tests.", "Write Supertest flow: full vendor lifecycle from registration to verified status.", "Must Have", "Scenario 1, Fig 9"],
        ["T-61", "Chat flow end-to-end", "Create conversation -> send message -> receive -> mark read.", "Not Tested", "Not Tested", "Not Tested", "No integration tests.", "Write Socket.IO + Supertest test: two users exchange messages with read receipts.", "Must Have", "Scenario 4"],
        ["T-62", "Vendor search and discovery", "Search vendors -> filter by category -> view detail -> start conversation.", "Not Tested", "Not Tested", "Not Tested", "No integration tests.", "Write Supertest flow: public listing filters correctly, detail returns full data.", "Must Have", "Scenario 2"],
        ["T-63", "Booking and review flow (future)", "Book vendor -> complete booking -> leave review -> rating updates.", "Not Tested", "Not Tested", "Not Tested", "Depends on booking + review modules.", "Implement modules first, then write full flow integration test.", "Must Have", "Scenario 3"],
        ["T-64", "Payment flow (future)", "Initiate payment -> Chapa webhook -> booking confirmed -> expense recorded.", "Not Tested", "Not Tested", "Not Tested", "Depends on payment + budget modules.", "Implement modules first, then write webhook-driven integration test.", "Must Have", "3.2.7"],
        ["T-65", "Budget lifecycle (future)", "Create budget -> allocate categories -> add expenses -> check alerts.", "Not Tested", "Not Tested", "Not Tested", "Depends on budget module.", "Implement module first, then write full budget lifecycle integration test.", "Must Have", "3.2.5"],
        ["T-66", "AI recommendation flow (future)", "Set preferences -> request recommendations -> verify ranking.", "Not Tested", "Not Tested", "Not Tested", "Depends on AI module.", "Implement module first, then write recommendation accuracy integration test.", "Must Have", "Scenario 5"],
        ["T-67", "Admin vendor moderation flow", "View pending vendors -> approve/reject -> verify status change in public listing.", "Not Tested", "Not Tested", "Not Tested", "No integration tests.", "Write Supertest flow: admin approves, then public API includes vendor.", "Must Have", "3.2.3"],
    ]),
    ("Integration Tests - Security & Edge Cases", [
        ["T-68", "Cross-role access prevention", "Couple cannot access vendor/admin endpoints; vendor cannot access admin.", "Not Tested", "Not Tested", "Not Tested", "No integration tests.", "Write Supertest tests: each role attempts all other role endpoints, expect 403.", "Must Have", "4.11.4"],
        ["T-69", "Expired session handling", "Requests with expired session return 401.", "Not Tested", "Not Tested", "Not Tested", "No integration tests.", "Write test: manipulate session expiry, verify 401 response.", "Must Have", "4.11.1"],
        ["T-70", "Concurrent chat stress test", "Multiple users sending messages simultaneously.", "Not Tested", "Not Tested", "Not Tested", "No integration tests.", "Write Socket.IO test: 10+ concurrent connections sending messages, verify delivery.", "Should Have", "3.2.6"],
        ["T-71", "SQL injection prevention", "Malicious input in search/filter parameters doesn't execute.", "Not Tested", "Not Tested", "Not Tested", "No integration tests.", "Write test: SQL injection payloads in vendor search, verify parameterized safety.", "Must Have", "3.3.3"],
        ["T-72", "File upload validation", "Reject oversized files and unsupported MIME types.", "Not Tested", "Not Tested", "Not Tested", "No integration tests.", "Write test: upload > max size returns 413; invalid type returns 400.", "Must Have", "3.2.3"],
        ["T-73", "Rate limiting", "Excessive requests from single IP get throttled.", "Not Tested", "Not Tested", "Not Tested", "No rate limiting implemented.", "Implement rate limiter, then write test: 100+ rapid requests trigger 429.", "Should Have", "3.3.3"],
    ]),
    ("User Acceptance Tests (UAT) - Couple", [
        ["T-74", "Couple registers and logs in", "New couple can register, verify email, and access dashboard.", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: complete registration flow, verify email link works, dashboard loads.", "Must Have", "Scenario 1"],
        ["T-75", "Couple searches vendors", "Couple finds vendors by category/location with correct results.", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: apply different filters, verify results match criteria.", "Must Have", "Scenario 2"],
        ["T-76", "Couple views vendor profile", "Full vendor profile displays with all sections (bio, portfolio, pricing).", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: open vendor detail, verify all data sections render correctly.", "Must Have", "3.4.2 UC8"],
        ["T-77", "Couple starts conversation", "Couple initiates chat with vendor from profile page.", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: click 'Message' on vendor profile, conversation starts, messages send/receive.", "Must Have", "Scenario 4"],
        ["T-78", "Couple receives notifications", "Couple sees real-time notifications for new messages.", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: send message from vendor, verify bell shows unread count.", "Must Have", "3.2.6"],
        ["T-79", "Couple creates budget (future)", "Couple sets total budget and allocates categories.", "Not Tested", "Not Tested", "Not Tested", "Budget module not built.", "After implementation: manual test budget creation and category allocation.", "Must Have", "US-C-009"],
        ["T-80", "Couple books vendor (future)", "Couple sends booking request and tracks status.", "Not Tested", "Not Tested", "Not Tested", "Booking module not built.", "After implementation: manual test full booking lifecycle.", "Must Have", "3.4.2 UC9"],
        ["T-81", "Couple makes payment (future)", "Couple completes payment via Chapa for booking.", "Not Tested", "Not Tested", "Not Tested", "Payment module not built.", "After implementation: manual test payment flow with Chapa sandbox.", "Must Have", "US-C-019"],
        ["T-82", "Couple leaves review (future)", "Couple submits star rating and text after completed booking.", "Not Tested", "Not Tested", "Not Tested", "Review module not built.", "After implementation: manual test review submission and display.", "Must Have", "US-C-022"],
        ["T-83", "Couple uses AI recommendations (future)", "Couple receives personalized vendor suggestions.", "Not Tested", "Not Tested", "Not Tested", "AI module not built.", "After implementation: manual test recommendation quality with test preferences.", "Must Have", "US-C-006"],
    ]),
    ("User Acceptance Tests (UAT) - Vendor", [
        ["T-84", "Vendor registers and sets up profile", "New vendor creates account and completes profile setup wizard.", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: register as vendor, fill setup form, verify profile saved.", "Must Have", "US-V-001"],
        ["T-85", "Vendor uploads documents", "Vendor uploads business license and ID for verification.", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: upload multiple document types, verify they appear in profile.", "Must Have", "US-V-002"],
        ["T-86", "Vendor submits for verification", "Vendor submits profile and sees pending status.", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: click submit, verify status banner changes to 'Pending Verification'.", "Must Have", "3.2.3"],
        ["T-87", "Vendor receives verification result", "Vendor sees approved/rejected status with feedback.", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: after admin action, verify status update and any rejection reason.", "Must Have", "US-V-003"],
        ["T-88", "Vendor manages chat", "Vendor responds to couple inquiries in real-time.", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: receive message, respond, verify typing indicators and read receipts.", "Must Have", "Scenario 4"],
        ["T-89", "Vendor views booking requests (future)", "Vendor sees incoming booking requests on dashboard.", "Not Tested", "Not Tested", "Not Tested", "Booking module not built.", "After implementation: manual test vendor booking list and accept/decline.", "Must Have", "3.4.2 UC9"],
        ["T-90", "Vendor receives payment (future)", "Vendor sees payment confirmation after couple pays.", "Not Tested", "Not Tested", "Not Tested", "Payment module not built.", "After implementation: manual test payment notification and history.", "Must Have", "3.2.7"],
    ]),
    ("User Acceptance Tests (UAT) - Admin", [
        ["T-91", "Admin reviews pending vendors", "Admin sees list of vendors awaiting verification.", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: login as admin, navigate to vendor list, filter by pending status.", "Must Have", "3.2.3"],
        ["T-92", "Admin approves vendor", "Admin reviews documents and approves vendor.", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: open vendor detail, review docs, click approve, verify status change.", "Must Have", "3.4.2 UC14"],
        ["T-93", "Admin rejects vendor with reason", "Admin rejects vendor and provides feedback.", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: click reject, enter reason, verify vendor sees rejection reason.", "Must Have", "3.4.2 UC14"],
        ["T-94", "Admin suspends vendor", "Admin suspends active vendor for policy violation.", "Not Tested", "Not Tested", "Not Tested", "No UAT performed.", "Manual test: suspend vendor, verify removed from public listing.", "Must Have", "4.11.3"],
        ["T-95", "Admin manages users (future)", "Admin views, bans, and reactivates user accounts.", "Not Tested", "Not Tested", "Not Tested", "User management UI not built.", "After implementation: manual test full user management lifecycle.", "Must Have", "US-A-001"],
        ["T-96", "Admin moderates reviews (future)", "Admin approves or removes inappropriate reviews.", "Not Tested", "Not Tested", "Not Tested", "Review module not built.", "After implementation: manual test review moderation queue.", "Must Have", "3.4.2 UC15"],
        ["T-97", "Admin views analytics (future)", "Admin sees platform statistics dashboard.", "Not Tested", "Not Tested", "Not Tested", "Analytics not built.", "After implementation: manual test charts and metrics accuracy.", "Should Have", "4.6.8"],
    ]),
    ("Performance Tests", [
        ["T-98", "API response time under load", "95% of API responses complete within 1 second under 100 concurrent users.", "Not Tested", "Not Tested", "Not Tested", "No performance testing done.", "Run k6/Artillery load test on key endpoints, measure p95 latency.", "Must Have", "3.3.1"],
        ["T-99", "WebSocket message latency", "Messages delivered within 500ms under normal conditions.", "Not Tested", "Not Tested", "Not Tested", "No performance testing done.", "Run Socket.IO load test, measure message delivery time.", "Should Have", "3.3.1"],
        ["T-100", "Database query performance", "Complex queries (vendor search, budget sums) execute < 500ms.", "Not Tested", "Not Tested", "Not Tested", "No performance testing done.", "Profile slow queries with EXPLAIN ANALYZE, verify index usage.", "Should Have", "3.3.1"],
        ["T-101", "Concurrent user capacity", "System handles 5000 simultaneous connections without degradation.", "Not Tested", "Not Tested", "Not Tested", "No performance testing done.", "Run stress test with k6, monitor memory/CPU/response times at scale.", "Should Have", "3.3.1"],
        ["T-102", "Recommendation response time (future)", "AI recommendations return within 800ms with Redis caching.", "Not Tested", "Not Tested", "Not Tested", "AI module not built.", "After implementation: benchmark recommendation API with/without cache.", "Should Have", "3.3.1"],
    ]),
    ("Security Tests", [
        ["T-103", "XSS prevention", "User-generated content (chat, reviews) sanitized against XSS.", "Not Tested", "Not Tested", "Not Tested", "No security testing done.", "Test with XSS payloads in chat messages, vendor descriptions, review text.", "Must Have", "3.3.3"],
        ["T-104", "CSRF protection", "Cross-site requests blocked by SameSite cookie + origin check.", "Not Tested", "Not Tested", "Not Tested", "No security testing done.", "Test cross-origin POST requests are rejected by session cookie policy.", "Must Have", "3.3.3"],
        ["T-105", "Authentication bypass attempts", "Cannot access protected resources without valid session.", "Not Tested", "Not Tested", "Not Tested", "No security testing done.", "Test: forged cookies, missing cookies, expired cookies all return 401.", "Must Have", "4.11.1"],
        ["T-106", "Privilege escalation prevention", "Users cannot modify their role or access higher-privilege endpoints.", "Not Tested", "Not Tested", "Not Tested", "No security testing done.", "Test: couple attempts admin endpoints, vendor attempts couple-only routes.", "Must Have", "4.11.4"],
        ["T-107", "File upload security", "Uploaded files scanned for malicious content, path traversal blocked.", "Not Tested", "Not Tested", "Not Tested", "No security testing done.", "Test: upload with path traversal filenames, executable extensions, oversized files.", "Must Have", "3.3.3"],
        ["T-108", "Data privacy compliance", "Users can only access their own data (conversations, budget, etc.).", "Not Tested", "Not Tested", "Not Tested", "No security testing done.", "Test: user A attempts to read user B conversations, budget, profile.", "Must Have", "3.3.3"],
    ]),
]

NONFUNCTIONAL_DATA = [
    ("Performance", [
        ["NF-01", "Sub-2-Second Response Time", "95% of requests respond within 1 second.", "In Progress", "In Progress", "In Progress", "Basic API responses are fast. No load testing or optimization done.", "Add performance monitoring, load testing, optimize slow queries, add caching.", "Must Have", "3.3.1"],
        ["NF-02", "5000 Concurrent Users", "System supports 5000+ concurrent active users.", "Not Started", "Not Started", "Not Started", "No load testing or scaling configuration.", "Load test with k6/Artillery, configure horizontal scaling, optimize WebSocket capacity.", "Should Have", "3.3.1"],
        ["NF-03", "Recommendation < 800ms", "AI recommendations return within 800ms.", "Not Started", "Not Started", "Not Started", "No AI system to benchmark.", "Will depend on Redis caching strategy once AI engine is built.", "Should Have", "3.3.1"],
    ]),
    ("Security", [
        ["NF-04", "HTTPS/TLS Encryption", "All client-server communication encrypted.", "In Progress", "In Progress", "In Progress", "Development uses HTTP. Production will require TLS setup.", "Configure TLS certificates for production deployment.", "Must Have", "3.3.3"],
        ["NF-05", "Data Encryption at Rest", "Sensitive fields encrypted in database.", "Not Started", "Not Started", "Not Started", "No field-level encryption implemented.", "Identify PII fields, implement encryption/decryption layer.", "Should Have", "3.3.3, 4.8.5"],
        ["NF-06", "Input Validation (OWASP)", "All API inputs validated against injection.", "In Progress", "In Progress", "In Progress", "Some validation in controllers. No comprehensive validation framework.", "Add Zod/Joi schema validation on all endpoints, sanitize inputs.", "Must Have", "3.3.3"],
        ["NF-07", "SQL Injection Prevention", "Parameterized queries throughout.", "Done", "Done", "Done", "Repository layer uses parameterized queries with pg client.", "", "Must Have", "3.3.3"],
    ]),
    ("Reliability & Availability", [
        ["NF-08", "99.5% Monthly Uptime", "System available 99.5% of the time.", "Not Started", "Not Started", "Not Started", "No uptime monitoring or redundancy.", "Configure health checks, auto-restart, monitoring alerts.", "Should Have", "3.3.2"],
        ["NF-09", "Daily Automated Backups", "PostgreSQL backup with RPO <= 24 hours.", "Not Started", "Not Started", "Not Started", "Neon provides some backup but no custom strategy.", "Configure Neon automated backups, verify PITR capability.", "Should Have", "3.3.2, 4.8.4"],
        ["NF-10", "Disaster Recovery (RTO <= 4h)", "Recover from critical failures within 4 hours.", "Not Started", "Not Started", "Not Started", "No DR plan.", "Document recovery procedures, test backup restoration.", "Should Have", "3.3.2, 4.8.4"],
    ]),
    ("Scalability & Infrastructure", [
        ["NF-11", "Horizontal Backend Scaling", "Backend supports multiple instances.", "In Progress", "In Progress", "Not Started", "Stateless Express API supports scaling. Socket.IO needs Redis adapter for multi-instance.", "Add Socket.IO Redis adapter, configure load balancer.", "Should Have", "3.3.6"],
        ["NF-12", "Database Indexing & Caching", "Efficient queries with proper indexes.", "In Progress", "In Progress", "Not Started", "vendor_profiles has btree_gist index. No Redis caching layer.", "Add Redis for frequently accessed data, optimize query plans.", "Should Have", "3.3.6"],
        ["NF-13", "Cloud Deployment", "Deploy to AWS/GCP/Vercel.", "Not Started", "Not Started", "Not Started", "Running locally only.", "Configure CI/CD pipeline, deploy backend to cloud, frontend to Vercel.", "Should Have", "5.5"],
    ]),
    ("Mobile & Accessibility", [
        ["NF-14", "Flutter Mobile App", "Cross-platform mobile application.", "Not Started", "Not Started", "Not Started", "No Flutter/mobile codebase.", "Initialize Flutter project, implement core screens, connect to same backend API.", "Should Have", "1.11"],
        ["NF-15", "Responsive Web UI", "Works on mobile browsers and tablets.", "In Progress", "Done", "In Progress", "Tailwind CSS used with responsive classes. Not all pages fully tested on mobile.", "Test all pages on mobile viewports, fix responsive issues.", "Must Have", "3.3.4"],
        ["NF-16", "Accessibility Standards", "WCAG compliance for core interactions.", "Not Started", "Not Started", "Not Started", "No accessibility audit done.", "Add ARIA labels, keyboard navigation, screen reader testing.", "Should Have", "3.3.4"],
    ]),
]


def create_overview_sheet(wb):
    ws = wb.active
    ws.title = "Overview"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16

    row = 1
    ws.cell(row=row, column=2, value="TWEDAR - Project Progress Tracker").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=2, value="AI-Enabled Smart Wedding Planning Platform").font = SUBTITLE_FONT
    row += 2

    ws.cell(row=row, column=2, value="HOW TO USE THIS SPREADSHEET").font = BOLD_BODY
    row += 1
    instructions = [
        "1. Each tab represents a major feature module from the project document.",
        "2. Status colors: GREEN = Done, YELLOW = In Progress, RED = Not Started.",
        "3. Column D shows overall status. Columns E & F show backend/frontend separately.",
        "4. Column H describes exactly what remains to be built for each feature.",
        "5. Column I shows priority (Must Have / Should Have / Could Have).",
        "6. Column J references the section in the project document PDF.",
        "7. Use the summary table below to see completion % per module.",
    ]
    for instr in instructions:
        ws.cell(row=row, column=2, value=instr).font = BODY_FONT
        row += 1

    row += 2
    ws.cell(row=row, column=2, value="COLOR LEGEND").font = BOLD_BODY
    row += 1
    legend = [("Done", DONE_FILL, DONE_FONT), ("In Progress", PROGRESS_FILL, PROGRESS_FONT), ("Not Started", NOT_STARTED_FILL, NOT_STARTED_FONT)]
    for label, fill, font in legend:
        cell = ws.cell(row=row, column=2, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 2
    ws.cell(row=row, column=2, value="PROJECT SUMMARY").font = BOLD_BODY
    row += 1

    summary_headers = ["Module", "Total", "Done", "In Progress", "Not Started", "Completion %"]
    for col_idx, h in enumerate(summary_headers, 2):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    row += 1

    all_sheets_data = [
        ("1. Auth & User Mgmt", AUTH_DATA),
        ("2. Vendor Management", VENDOR_DATA),
        ("3. Real-Time Comms", REALTIME_DATA),
        ("4. Budgeting & Expense", BUDGET_DATA),
        ("5. AI Recommendations", AI_DATA),
        ("6. Booking & Scheduling", BOOKING_DATA),
        ("7. Payment Integration", PAYMENT_DATA),
        ("8. Review & Feedback", REVIEW_DATA),
        ("9. Admin & Analytics", ADMIN_DATA),
        ("10. Couple Planning", COUPLE_DATA),
        ("11. Non-Functional", NONFUNCTIONAL_DATA),
    ]

    grand_total = 0
    grand_done = 0
    grand_progress = 0
    grand_not_started = 0

    for module_name, sections in all_sheets_data:
        total = sum(len(items) for _, items in sections)
        done = sum(1 for _, items in sections for item in items if item[3] == "Done")
        progress = sum(1 for _, items in sections for item in items if item[3] == "In Progress")
        not_started = sum(1 for _, items in sections for item in items if item[3] == "Not Started")
        pct = round((done / total) * 100) if total > 0 else 0

        grand_total += total
        grand_done += done
        grand_progress += progress
        grand_not_started += not_started

        data_row = [module_name, total, done, progress, not_started, f"{pct}%"]
        for col_idx, val in enumerate(data_row, 2):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left")
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL

        pct_cell = ws.cell(row=row, column=7)
        if pct >= 80:
            pct_cell.fill = DONE_FILL
            pct_cell.font = DONE_FONT
        elif pct >= 30:
            pct_cell.fill = PROGRESS_FILL
            pct_cell.font = PROGRESS_FONT
        else:
            pct_cell.fill = NOT_STARTED_FILL
            pct_cell.font = NOT_STARTED_FONT

        row += 1

    # Grand total row
    grand_pct = round((grand_done / grand_total) * 100) if grand_total > 0 else 0
    totals_row = ["TOTAL", grand_total, grand_done, grand_progress, grand_not_started, f"{grand_pct}%"]
    for col_idx, val in enumerate(totals_row, 2):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="1B2A4A")
        cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left")
        cell.border = THIN_BORDER
        cell.fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")

    row += 3
    ws.cell(row=row, column=2, value="PROJECT METADATA").font = BOLD_BODY
    row += 1
    metadata = [
        ("Project:", "Twedar - AI-Enabled Smart Wedding Planning Platform"),
        ("Team:", "Chera Mihretu, Abdi Esayas, Surafel Takele, Tamirat Kebede, Hanamariam Mesfin"),
        ("University:", "Adama Science and Technology University"),
        ("Advisor:", "Mr. Anteneh"),
        ("Generated:", "April 2026"),
        ("Document:", "Twedar GC Project (110 pages)"),
    ]
    for label, value in metadata:
        ws.cell(row=row, column=2, value=label).font = BOLD_BODY
        ws.cell(row=row, column=3, value=value).font = BODY_FONT
        row += 1


def main():
    wb = Workbook()

    create_overview_sheet(wb)

    sheets_config = [
        ("1. Auth & User Mgmt", AUTH_DATA),
        ("2. Vendor Management", VENDOR_DATA),
        ("3. Real-Time Comms", REALTIME_DATA),
        ("4. Budgeting & Expense", BUDGET_DATA),
        ("5. AI Recommendations", AI_DATA),
        ("6. Booking & Scheduling", BOOKING_DATA),
        ("7. Payment Integration", PAYMENT_DATA),
        ("8. Review & Feedback", REVIEW_DATA),
        ("9. Admin & Analytics", ADMIN_DATA),
        ("10. Couple Planning", COUPLE_DATA),
        ("11. Non-Functional", NONFUNCTIONAL_DATA),
    ]

    for sheet_name, data in sheets_config:
        ws = wb.create_sheet(title=sheet_name)
        populate_feature_sheet(ws, data)

    output_path = "/home/chera/Public/my_stuffs/astu/twedar/Twedar_Progress_Tracker.xlsx"
    wb.save(output_path)
    print(f"Workbook saved to: {output_path}")
    print(f"Sheets: {wb.sheetnames}")

    total_items = 0
    done_items = 0
    for _, data in sheets_config:
        for _, items in data:
            total_items += len(items)
            done_items += sum(1 for item in items if item[3] == "Done")
    print(f"Total features tracked: {total_items}")
    print(f"Completed: {done_items} ({round(done_items/total_items*100)}%)")
    print(f"Remaining: {total_items - done_items}")


if __name__ == "__main__":
    main()
