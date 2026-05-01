#!/usr/bin/env python3
"""
Twedar Manual Test Assignment Generator
Generates a comprehensive manual test plan Excel workbook assigned to Sura & Hanamariam.
Includes dropdowns for test result (Pass/Fail/Not Implemented), expected results,
how to test instructions, and notes columns.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# === Styles ===
HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
SUBHEADER_FONT = Font(name="Segoe UI", size=10, bold=True, color="1B2A4A")

SURA_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HANA_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

PRIORITY_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PRIORITY_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PRIORITY_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALT_ROW_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
DEADLINE_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
DEADLINE_FONT = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")

TITLE_FONT = Font(name="Segoe UI", size=16, bold=True, color="1B2A4A")
SUBTITLE_FONT = Font(name="Segoe UI", size=12, bold=False, color="4472C4")
BODY_FONT = Font(name="Segoe UI", size=10, color="333333")
BOLD_BODY = Font(name="Segoe UI", size=10, bold=True, color="333333")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

ASSIGNEE_STYLES = {
    "Sura": SURA_FILL,
    "Hanamariam": HANA_FILL,
}

COLUMNS = [
    "Test ID", "Module", "Test Case", "Description", "Preconditions",
    "How to Test (Steps)", "Expected Result", "Tester", "Priority",
    "Deadline", "Result", "Notes"
]
COL_WIDTHS = [10, 18, 32, 45, 35, 55, 45, 14, 12, 14, 18, 35]

# ========== TEST CASES ==========

AUTH_TESTS = [
    ("Registration", [
        ["TC-A-001", "Email Registration - Valid", "Register a new couple account with valid email and password",
         "No existing account with test email", 
         "1. Navigate to /register\n2. Select 'Couple' role\n3. Enter valid email, name, and password (8+ chars, uppercase, number)\n4. Click Register\n5. Verify success message",
         "Account created successfully, verification email sent, user redirected to verify page", "High", "May 2"],
        ["TC-A-002", "Email Registration - Duplicate Email", "Attempt registration with existing email",
         "An account with test email already exists",
         "1. Navigate to /register\n2. Enter email that is already registered\n3. Fill all other fields correctly\n4. Click Register",
         "Error message: 'Email already exists' or similar, no new account created", "High", "May 2"],
        ["TC-A-003", "Email Registration - Weak Password", "Attempt registration with a weak password",
         "No existing account",
         "1. Navigate to /register\n2. Enter valid email\n3. Enter weak password (e.g., '123')\n4. Click Register",
         "Validation error showing password requirements (min 8 chars, uppercase, number)", "High", "May 2"],
        ["TC-A-004", "Email Registration - Empty Fields", "Submit registration with missing required fields",
         "None",
         "1. Navigate to /register\n2. Leave name, email, or password empty\n3. Click Register",
         "Form validation errors shown for each empty required field", "Medium", "May 2"],
        ["TC-A-005", "Vendor Registration", "Register a new vendor account",
         "No existing vendor account",
         "1. Navigate to /register\n2. Select 'Vendor' role\n3. Enter business name, email, password\n4. Click Register\n5. Verify organization auto-created",
         "Vendor account created, organization auto-created, redirected to vendor profile setup", "High", "May 2"],
        ["TC-A-006", "Email Verification", "Verify email after registration",
         "Newly registered account, verification email received",
         "1. Register a new account\n2. Check email inbox for verification link\n3. Click the verification link\n4. Verify account status changes",
         "Email marked as verified, user can now fully access the platform", "High", "May 2"],
    ]),
    ("Login", [
        ["TC-A-007", "Email Login - Valid Credentials", "Login with correct email and password",
         "Registered and verified account exists",
         "1. Navigate to /login\n2. Enter valid email and password\n3. Click Login\n4. Verify redirection based on role",
         "User authenticated, session cookie set, redirected to role-based dashboard (couple→/dashboard, vendor→/vendor/dashboard)", "High", "May 2"],
        ["TC-A-008", "Email Login - Wrong Password", "Login with incorrect password",
         "Registered account exists",
         "1. Navigate to /login\n2. Enter valid email but wrong password\n3. Click Login",
         "Error message: 'Invalid credentials' or similar, no session created", "High", "May 2"],
        ["TC-A-009", "Email Login - Non-existent Email", "Login with email not in system",
         "None",
         "1. Navigate to /login\n2. Enter email that doesn't exist\n3. Enter any password\n4. Click Login",
         "Error message: 'Invalid credentials', no session created", "High", "May 2"],
        ["TC-A-010", "Google OAuth Login", "Login via Google social login",
         "Google account available, app configured with Google OAuth",
         "1. Navigate to /login\n2. Click 'Sign in with Google'\n3. Complete Google OAuth flow\n4. Verify redirect back to app",
         "User authenticated via Google, session created, redirected to couple dashboard", "High", "May 2"],
        ["TC-A-011", "Apple OAuth Login", "Login via Apple social login",
         "Apple account available, app configured with Apple OAuth",
         "1. Navigate to /login\n2. Click 'Sign in with Apple'\n3. Complete Apple OAuth flow\n4. Verify redirect back to app",
         "User authenticated via Apple, session created, redirected to couple dashboard", "Medium", "May 2"],
    ]),
    ("Session Management", [
        ["TC-A-012", "Session Persistence", "Session persists across page refreshes",
         "User is logged in",
         "1. Login successfully\n2. Refresh the page (F5)\n3. Verify user is still authenticated\n4. Close and reopen the browser tab",
         "User remains logged in after refresh, session cookie still valid", "High", "May 2"],
        ["TC-A-013", "Session Expiry", "Session expires after timeout",
         "User is logged in, know session timeout duration",
         "1. Login successfully\n2. Wait for session timeout (or manually expire via dev tools)\n3. Attempt to navigate to a protected page",
         "User redirected to login page, session expired gracefully", "Medium", "May 2"],
        ["TC-A-014", "Logout", "User logs out successfully",
         "User is logged in",
         "1. Click logout button/link\n2. Verify redirect to login page\n3. Attempt to access protected route directly via URL",
         "Session destroyed, cookie cleared, user cannot access protected routes", "High", "May 2"],
        ["TC-A-015", "401 Global Handling", "Unauthorized access redirects to login",
         "User is not logged in",
         "1. Clear all cookies\n2. Navigate directly to a protected route (e.g., /dashboard)\n3. Verify redirect",
         "User redirected to /login with appropriate message", "High", "May 2"],
    ]),
    ("Password Reset", [
        ["TC-A-016", "Forgot Password - Valid Email", "Request password reset for valid email",
         "Account with test email exists",
         "1. Navigate to /login\n2. Click 'Forgot Password'\n3. Enter registered email\n4. Click Submit",
         "Success message: 'Reset link sent', email received with reset link", "High", "May 2"],
        ["TC-A-017", "Forgot Password - Invalid Email", "Request reset for non-existent email",
         "No account with test email",
         "1. Navigate to forgot password page\n2. Enter email not in system\n3. Click Submit",
         "Generic success message (no information leak about account existence)", "Medium", "May 2"],
        ["TC-A-018", "Password Reset - Use Link", "Reset password using email link",
         "Valid reset link received",
         "1. Click reset link from email\n2. Enter new password meeting requirements\n3. Confirm new password\n4. Submit\n5. Try logging in with new password",
         "Password updated successfully, can login with new password, old password no longer works", "High", "May 2"],
    ]),
    ("Role-Based Access", [
        ["TC-A-019", "Couple Access Control", "Couple can only access couple routes",
         "Logged in as couple user",
         "1. Login as couple\n2. Access /dashboard (should work)\n3. Try /vendor/dashboard (should fail)\n4. Try /admin/dashboard (should fail)",
         "Access granted for couple routes, 403/redirect for vendor/admin routes", "High", "May 2"],
        ["TC-A-020", "Vendor Access Control", "Vendor can only access vendor routes",
         "Logged in as vendor user",
         "1. Login as vendor\n2. Access /vendor/dashboard (should work)\n3. Try /dashboard (should fail)\n4. Try /admin/dashboard (should fail)",
         "Access granted for vendor routes, 403/redirect for couple/admin routes", "High", "May 2"],
        ["TC-A-021", "Admin Access Control", "Admin can access admin routes",
         "Logged in as admin user",
         "1. Login as admin\n2. Access /admin/dashboard (should work)\n3. Verify full system access",
         "Access granted for admin routes and system management features", "High", "May 2"],
        ["TC-A-022", "RBAC Middleware Enforcement", "API rejects unauthorized role access",
         "Logged in as couple user",
         "1. Login as couple\n2. Use browser dev tools or Postman to call vendor-only API endpoint\n3. Check response",
         "API returns 403 Forbidden with appropriate error message", "High", "May 2"],
    ]),
]

VENDOR_TESTS = [
    ("Vendor Profile", [
        ["TC-V-001", "Create Business Profile", "Vendor creates initial business profile",
         "Logged in as vendor owner, no profile exists yet",
         "1. Navigate to /vendor/profile/setup\n2. Fill in business name, category, description\n3. Add contact details\n4. Click Save",
         "Profile created, visible in vendor dashboard, all fields persisted", "High", "May 3"],
        ["TC-V-002", "Update Business Profile", "Edit existing vendor profile",
         "Vendor profile exists",
         "1. Navigate to vendor profile page\n2. Edit business name or description\n3. Save changes\n4. Refresh and verify",
         "Changes saved and displayed correctly after refresh", "High", "May 3"],
        ["TC-V-003", "Service & Pricing Packages", "Add and edit service packages",
         "Vendor profile exists",
         "1. Navigate to vendor profile\n2. Add a new service package (name, price, description)\n3. Save\n4. Edit the package\n5. Delete a package",
         "Packages created, edited, and deleted correctly", "High", "May 3"],
        ["TC-V-004", "Location with Map", "Set vendor location with map picker",
         "Vendor profile exists",
         "1. Navigate to vendor profile location section\n2. Use map picker to set location\n3. Verify lat/long saved\n4. Check location displays on public profile",
         "Map displays correctly, coordinates saved, location visible on public profile", "High", "May 3"],
        ["TC-V-005", "Social Links", "Add social media links",
         "Vendor profile exists",
         "1. Navigate to profile settings\n2. Add Facebook, Instagram, TikTok links\n3. Save\n4. Verify links display on public profile",
         "Social links saved and displayed as clickable icons/links", "Low", "May 3"],
    ]),
    ("Document & Verification", [
        ["TC-V-006", "Upload Business Document", "Vendor uploads verification documents",
         "Logged in as vendor, profile exists",
         "1. Navigate to verification section\n2. Click upload document\n3. Select a PDF/image file\n4. Verify file uploaded successfully",
         "Document uploaded, shown in documents list with filename and timestamp", "High", "May 3"],
        ["TC-V-007", "Delete Uploaded Document", "Remove a previously uploaded document",
         "At least one document uploaded",
         "1. Navigate to documents section\n2. Click delete on a document\n3. Confirm deletion\n4. Verify removed from list",
         "Document removed from list and storage, cannot be recovered", "Medium", "May 3"],
        ["TC-V-008", "Submit for Verification", "Submit profile for admin review",
         "Profile complete, documents uploaded, status is 'draft' or 'rejected'",
         "1. Navigate to verification section\n2. Click 'Submit for Verification'\n3. Verify status changes to 'pending'",
         "Status changes to 'pending_verification', submit button disabled, status displayed", "High", "May 3"],
        ["TC-V-009", "Verification Status Display", "Vendor sees current verification status",
         "Profile submitted for verification",
         "1. Navigate to vendor dashboard\n2. Check verification status badge/indicator\n3. Verify correct status shown",
         "Current status (pending/approved/rejected) clearly visible with appropriate styling", "Medium", "May 3"],
        ["TC-V-010", "Resubmission After Rejection", "Resubmit after admin rejection",
         "Profile was rejected by admin",
         "1. View rejection reason\n2. Update profile/documents to address issues\n3. Click 'Resubmit'\n4. Verify status changes back to pending",
         "Rejection reason visible, profile editable, resubmission changes status to pending", "High", "May 3"],
    ]),
    ("Admin Vendor Actions", [
        ["TC-V-011", "Admin Lists All Vendors", "Admin views vendor list with status",
         "Logged in as admin, vendors exist in system",
         "1. Navigate to admin vendor management\n2. View vendor list\n3. Check filter by status works",
         "All vendors listed with name, category, status; filtering works correctly", "High", "May 4"],
        ["TC-V-012", "Admin Approves Vendor", "Admin approves a pending vendor",
         "Vendor in 'pending_verification' status",
         "1. Navigate to admin vendor list\n2. Click on pending vendor\n3. Review documents\n4. Click 'Approve'",
         "Vendor status changes to 'approved', vendor receives notification", "High", "May 4"],
        ["TC-V-013", "Admin Rejects Vendor", "Admin rejects vendor with reason",
         "Vendor in 'pending_verification' status",
         "1. Navigate to admin vendor list\n2. Click on pending vendor\n3. Click 'Reject'\n4. Enter rejection reason\n5. Submit",
         "Status changes to 'rejected', rejection reason saved, vendor notified", "High", "May 4"],
        ["TC-V-014", "Admin Suspends Vendor", "Admin suspends an active vendor",
         "Vendor in 'approved' status",
         "1. Navigate to admin vendor management\n2. Select active vendor\n3. Click 'Suspend'\n4. Confirm action",
         "Vendor status changes to 'suspended', vendor cannot access dashboard, hidden from public", "High", "May 4"],
        ["TC-V-015", "Admin Reinstates Vendor", "Admin reactivates suspended vendor",
         "Vendor in 'suspended' status",
         "1. Navigate to admin vendor management\n2. Select suspended vendor\n3. Click 'Reinstate'\n4. Confirm",
         "Vendor status changes back to 'approved', vendor regains access", "Medium", "May 4"],
    ]),
    ("Public Vendor Discovery", [
        ["TC-V-016", "Public Vendor Listing Page", "Browse verified vendors as couple",
         "Logged in as couple, verified vendors exist",
         "1. Navigate to /vendors\n2. View vendor cards\n3. Verify only verified vendors shown",
         "Vendor listing page shows only approved/verified vendors with cards (photo, name, category, rating)", "High", "May 4"],
        ["TC-V-017", "Vendor Detail Page", "View full vendor profile",
         "Verified vendors exist",
         "1. From vendor listing, click on a vendor\n2. View full profile details\n3. Check all sections render",
         "Full profile visible: description, services, pricing, location, portfolio, social links", "High", "May 4"],
        ["TC-V-018", "Filter by Category", "Filter vendors by service type",
         "Multiple vendors in different categories",
         "1. Go to vendor listing\n2. Select a category filter (e.g., Photography)\n3. Verify results",
         "Only vendors matching selected category shown, count updates", "High", "May 4"],
        ["TC-V-019", "Filter by Location", "Filter vendors by city/area",
         "Vendors in multiple locations",
         "1. Go to vendor listing\n2. Select a location filter\n3. Verify results match location",
         "Only vendors in selected location shown", "High", "May 4"],
        ["TC-V-020", "Filter by Verification Status", "Show only verified vendors",
         "Mix of verified and unverified vendors",
         "1. Go to vendor listing\n2. Toggle 'Verified Only' filter\n3. Verify all shown vendors are verified",
         "Only verified vendors displayed, unverified hidden", "Medium", "May 4"],
    ]),
]

REALTIME_TESTS = [
    ("Chat / Messaging", [
        ["TC-R-001", "Create New Conversation", "Couple initiates chat with vendor",
         "Logged in as couple, verified vendor exists",
         "1. Navigate to vendor detail page\n2. Click 'Message' or 'Chat' button\n3. Verify conversation created",
         "New conversation created, chat window opens, both parties can see it", "High", "May 4"],
        ["TC-R-002", "Send Text Message", "Send a real-time message",
         "Active conversation between couple and vendor",
         "1. Open conversation\n2. Type a message\n3. Click send\n4. Verify message appears immediately\n5. Check other party receives it",
         "Message sent instantly, appears in both parties' chat windows without refresh", "High", "May 4"],
        ["TC-R-003", "Message History", "Load previous messages on opening chat",
         "Conversation with previous messages exists",
         "1. Close and reopen the chat window\n2. Scroll through messages\n3. Verify old messages loaded",
         "Previous messages loaded in correct chronological order with timestamps", "High", "May 4"],
        ["TC-R-004", "Typing Indicators", "Show typing indicator",
         "Active conversation open on both sides",
         "1. Start typing in chat\n2. Check other party's screen for typing indicator\n3. Stop typing, verify indicator disappears",
         "Typing indicator shows when other user is typing, disappears when they stop", "Medium", "May 5"],
        ["TC-R-005", "Read Receipts", "Messages marked as read",
         "Unread messages in conversation",
         "1. Send a message from one party\n2. Other party opens and views the message\n3. Check sender's view for read receipt",
         "Messages show read status (checkmarks or similar) once viewed by recipient", "Medium", "May 5"],
        ["TC-R-006", "Conversation List", "View all conversations",
         "Multiple conversations exist",
         "1. Navigate to messages page\n2. View conversation list\n3. Verify all conversations shown with last message preview",
         "All conversations listed with name, last message, timestamp, unread count", "High", "May 4"],
    ]),
    ("Notifications", [
        ["TC-R-007", "In-App Notification Received", "Real-time notification delivery",
         "User is online",
         "1. Trigger an event that sends notification (e.g., new message)\n2. Check notification bell for new notification\n3. Verify content is correct",
         "Notification appears in real-time without page refresh, badge count increments", "High", "May 5"],
        ["TC-R-008", "Notification Bell Badge", "Unread notification count shows",
         "User has unread notifications",
         "1. Observe notification bell icon\n2. Verify badge shows unread count\n3. Open notifications\n4. Verify count decreases when read",
         "Badge shows correct unread count, decreases to 0 when all read", "High", "May 5"],
        ["TC-R-009", "Mark Notification as Read", "Mark individual notification read",
         "Unread notifications exist",
         "1. Open notification dropdown\n2. Click on a notification\n3. Verify it's marked as read\n4. Badge count decreases",
         "Notification visual changes to 'read' state, unread count decremented", "Medium", "May 5"],
        ["TC-R-010", "Notification History via REST", "Fetch notifications via API",
         "Notifications exist for user",
         "1. Navigate to notifications page\n2. Verify all past notifications loaded\n3. Check pagination if many exist",
         "All notifications listed with timestamp, message, and read/unread status", "Medium", "May 5"],
    ]),
    ("Online Status", [
        ["TC-R-011", "User Online Status", "See if user is online",
         "Two users with a conversation",
         "1. Both users login\n2. Check online status indicator on chat\n3. One user logs out\n4. Check status updates",
         "Online/offline indicator shows correctly, updates in real-time when user connects/disconnects", "Low", "May 5"],
    ]),
]

BUDGET_TESTS = [
    ("Budget Management", [
        ["TC-B-001", "Create Wedding Budget", "Create initial budget with total amount",
         "Logged in as couple",
         "1. Navigate to budget section\n2. Click 'Create Budget'\n3. Enter total budget amount\n4. Save",
         "Budget created with total amount displayed, ready for category allocation", "High", "May 6"],
        ["TC-B-002", "Budget Category Allocation", "Allocate budget to categories",
         "Budget exists",
         "1. Open budget\n2. Add categories (Venue, Photography, Catering, etc.)\n3. Allocate amounts\n4. Verify total doesn't exceed budget",
         "Categories created with allocated amounts, total allocation shown vs overall budget", "High", "May 6"],
        ["TC-B-003", "Add Expense Entry", "Record an expense",
         "Budget with categories exists",
         "1. Navigate to expenses\n2. Click 'Add Expense'\n3. Select category, enter amount, description\n4. Save",
         "Expense recorded under category, budget totals updated in real-time", "High", "May 6"],
        ["TC-B-004", "Real-Time Budget Totals", "Budget updates after expenses",
         "Budget with categories and expenses",
         "1. View budget dashboard\n2. Check remaining amounts per category\n3. Add new expense\n4. Verify totals update",
         "Spent/remaining amounts update immediately, progress bars reflect changes", "High", "May 6"],
        ["TC-B-005", "Budget Alerts", "Warning when approaching limit",
         "Category nearly at allocation limit",
         "1. Set a category budget to 1000 ETB\n2. Add expenses totaling 900 ETB\n3. Check for warning indicator\n4. Exceed budget, check alert",
         "Warning shown at 80% threshold, alert/error at 100%", "Medium", "May 7"],
    ]),
]

AI_TESTS = [
    ("Recommendations", [
        ["TC-AI-001", "Preference Collection Wizard", "Couple fills out preference survey",
         "Logged in as couple, no preferences set",
         "1. Navigate to recommendations or first-time wizard\n2. Fill preference form (budget range, category, location, style)\n3. Submit",
         "Preferences saved, system begins generating recommendations", "High", "May 8"],
        ["TC-AI-002", "Content-Based Recommendations", "Get vendor recommendations based on preferences",
         "Preferences set, verified vendors exist",
         "1. Navigate to recommended vendors page\n2. View AI-suggested vendors\n3. Verify relevance to preferences",
         "Vendors shown match couple's preferences (budget, location, category)", "High", "May 8"],
        ["TC-AI-003", "Cold-Start Handling", "New user gets meaningful suggestions",
         "New couple account with no preferences or history",
         "1. Login as new couple\n2. Navigate to vendor listing\n3. Check if any suggestions shown",
         "System shows popular/top-rated vendors as default, prompts user to set preferences", "Medium", "May 8"],
        ["TC-AI-004", "Recommendation Response Time", "Recommendations load within 2 seconds",
         "Preferences set, system operational",
         "1. Navigate to recommendations page\n2. Time the load from click to display\n3. Repeat 3 times",
         "Recommendations appear within 2 seconds on average", "High", "May 8"],
        ["TC-AI-005", "Recommendation Badges", "Top-match badges on vendor cards",
         "Recommendations generated",
         "1. View vendor listing\n2. Check for 'Top Match' or match score badges\n3. Verify badge on AI-recommended vendors",
         "Recommended vendors display match badge/score, helping users identify best matches", "Medium", "May 9"],
    ]),
]

BOOKING_TESTS = [
    ("Booking Workflow", [
        ["TC-BK-001", "Send Booking Request", "Couple sends booking request",
         "Logged in as couple, verified vendor exists",
         "1. Navigate to vendor detail page\n2. Click 'Book Now'\n3. Select date and service package\n4. Add message\n5. Submit request",
         "Booking request created with 'pending' status, vendor notified", "High", "May 7"],
        ["TC-BK-002", "Vendor Views Booking Request", "Vendor sees incoming request",
         "Booking request sent to vendor",
         "1. Login as vendor\n2. Navigate to bookings/requests\n3. View incoming request details",
         "Request shown with couple info, requested date, service, and message", "High", "May 7"],
        ["TC-BK-003", "Vendor Accepts Booking", "Vendor accepts the request",
         "Pending booking request exists",
         "1. Open booking request\n2. Click 'Accept'\n3. Verify status changes\n4. Check couple receives notification",
         "Booking status changes to 'accepted/confirmed', couple notified", "High", "May 7"],
        ["TC-BK-004", "Vendor Declines Booking", "Vendor declines the request",
         "Pending booking request exists",
         "1. Open booking request\n2. Click 'Decline'\n3. Optionally add reason\n4. Verify status changes",
         "Booking status changes to 'declined', couple notified with reason", "High", "May 7"],
        ["TC-BK-005", "Booking Status Tracking", "Both parties see status timeline",
         "Booking exists in any state",
         "1. Navigate to booking detail (as couple or vendor)\n2. View status timeline\n3. Verify history shown",
         "Status timeline shows all transitions with timestamps", "Medium", "May 7"],
        ["TC-BK-006", "Booking History", "View past bookings",
         "Multiple bookings in various states exist",
         "1. Navigate to booking history page\n2. View past and upcoming bookings\n3. Filter by status",
         "All bookings listed with correct statuses, filterable", "Medium", "May 7"],
    ]),
    ("Scheduling", [
        ["TC-BK-007", "Vendor Availability Calendar", "Set available dates",
         "Logged in as vendor",
         "1. Navigate to schedule/calendar\n2. Mark available dates/times\n3. Block off unavailable dates\n4. Save",
         "Calendar shows availability clearly, couples see only available dates when booking", "Medium", "May 8"],
        ["TC-BK-008", "Date Conflict Prevention", "Cannot double-book same date",
         "Vendor has one booking accepted for a date",
         "1. As another couple, try to book same vendor on same date\n2. Submit request",
         "System prevents double-booking, shows date as unavailable or warns conflict", "High", "May 8"],
    ]),
]

PAYMENT_TESTS = [
    ("Payment Processing", [
        ["TC-P-001", "Initiate Payment via Chapa", "Start payment for booking",
         "Accepted booking exists, Chapa configured",
         "1. Navigate to booking detail\n2. Click 'Pay Now' or 'Pay Deposit'\n3. Verify redirect to Chapa gateway",
         "User redirected to Chapa payment page with correct amount", "High", "May 9"],
        ["TC-P-002", "Successful Payment", "Complete payment successfully",
         "On Chapa payment page",
         "1. Enter test payment details\n2. Complete payment\n3. Verify redirect back to app\n4. Check booking status updated",
         "Payment confirmed, booking status updated to 'paid', receipt generated", "High", "May 9"],
        ["TC-P-003", "Failed Payment", "Payment fails or is cancelled",
         "On Chapa payment page",
         "1. Cancel payment or use failing test card\n2. Verify redirect back to app\n3. Check booking status unchanged",
         "User returned to app, error displayed, booking status remains 'pending payment'", "High", "May 9"],
        ["TC-P-004", "Payment History", "View transaction history",
         "Payments have been made",
         "1. Navigate to payment history page\n2. View all transactions\n3. Verify amounts and dates",
         "All transactions listed with date, amount, status, and reference", "Medium", "May 9"],
        ["TC-P-005", "Deposit Payment Flow", "Pay advance deposit",
         "Booking accepted, deposit required",
         "1. View booking\n2. See deposit amount (e.g., 30% of total)\n3. Click 'Pay Deposit'\n4. Complete payment",
         "Partial payment recorded, remaining balance shown, status updated", "High", "May 10"],
    ]),
]

REVIEW_TESTS = [
    ("Reviews", [
        ["TC-RV-001", "Submit Review with Rating", "Couple leaves review after service",
         "Completed booking exists",
         "1. Navigate to completed booking\n2. Click 'Leave Review'\n3. Select star rating (1-5)\n4. Write review text\n5. Submit",
         "Review submitted, visible on vendor profile, rating updated", "High", "May 10"],
        ["TC-RV-002", "One Review Per Booking", "Cannot submit duplicate review",
         "Review already submitted for booking",
         "1. Navigate to same completed booking\n2. Try to leave another review",
         "System prevents second review, shows existing review instead", "High", "May 10"],
        ["TC-RV-003", "Review Display on Vendor Page", "Reviews visible on vendor profile",
         "Vendor has reviews",
         "1. Navigate to vendor public profile\n2. Scroll to reviews section\n3. Verify reviews shown with stars",
         "Reviews displayed with star rating, text, reviewer name, and date", "High", "May 10"],
        ["TC-RV-004", "Vendor Rating Aggregation", "Average rating calculated correctly",
         "Vendor has multiple reviews with different ratings",
         "1. View vendor profile\n2. Check overall rating\n3. Manually calculate average\n4. Compare",
         "Displayed average matches calculated average of all review ratings", "Medium", "May 10"],
        ["TC-RV-005", "Admin Review Moderation", "Admin can approve/remove reviews",
         "Logged in as admin, reviews exist",
         "1. Navigate to admin review moderation\n2. View pending reviews\n3. Approve or reject a review\n4. Verify visibility changes",
         "Approved reviews visible publicly, rejected reviews hidden", "High", "May 10"],
    ]),
]

ADMIN_TESTS = [
    ("User Management", [
        ["TC-AD-001", "View All Users", "Admin views user list",
         "Logged in as admin, users exist",
         "1. Navigate to admin user management\n2. View user list\n3. Search/filter by role\n4. Check pagination",
         "All users listed with name, email, role, status; search and filter functional", "High", "May 5"],
        ["TC-AD-002", "Suspend User", "Admin suspends a user account",
         "Active user exists",
         "1. Find user in list\n2. Click 'Suspend'\n3. Confirm action\n4. Verify user status changes",
         "User marked as suspended, cannot login, visual indicator in admin list", "High", "May 5"],
        ["TC-AD-003", "Ban User", "Admin bans a user permanently",
         "Active user exists",
         "1. Find user\n2. Click 'Ban'\n3. Enter ban reason\n4. Confirm\n5. Verify banned user cannot login",
         "User banned with reason stored, login blocked, ban reason visible to admin", "High", "May 5"],
        ["TC-AD-004", "Reactivate User", "Admin reactivates suspended user",
         "Suspended user exists",
         "1. Find suspended user\n2. Click 'Reactivate'\n3. Confirm\n4. Verify user can login again",
         "User status back to active, can login and access platform normally", "Medium", "May 5"],
        ["TC-AD-005", "Impersonate User", "Admin views as another user",
         "Admin logged in, target user exists",
         "1. Find user in admin panel\n2. Click 'Impersonate'\n3. Verify you see platform as that user\n4. Exit impersonation",
         "Admin sees platform from user's perspective, exit banner shown, audit logged", "Low", "May 6"],
    ]),
    ("Vendor Verification Admin", [
        ["TC-AD-006", "Vendor Verification Dashboard", "Admin reviews vendor submissions",
         "Pending vendors exist",
         "1. Navigate to vendor verification section\n2. View pending submissions\n3. Click on one to review",
         "List of pending vendors with submission date, documents viewable inline", "High", "May 5"],
        ["TC-AD-007", "Document Review", "Admin reviews uploaded documents",
         "Vendor with documents pending verification",
         "1. Open vendor verification details\n2. View each uploaded document\n3. Check document quality",
         "Documents render inline or downloadable, all submitted docs accessible", "High", "May 5"],
    ]),
]

COUPLE_TESTS = [
    ("Wedding Planning", [
        ["TC-C-001", "Create Wedding Plan", "Initialize wedding project",
         "Logged in as couple, no wedding plan exists",
         "1. Navigate to dashboard or setup wizard\n2. Enter wedding date\n3. Enter location/theme preferences\n4. Save",
         "Wedding project created with date, location, and theme stored", "High", "May 6"],
        ["TC-C-002", "Wedding Dashboard Overview", "View planning progress",
         "Wedding plan exists with some activity",
         "1. Navigate to /dashboard\n2. View overview cards (budget, checklist progress, vendor count)\n3. Verify data accuracy",
         "Dashboard shows correct stats for budget, tasks, vendors, days until wedding", "High", "May 6"],
        ["TC-C-003", "Wedding Checklist", "Use wedding checklist",
         "Wedding plan exists",
         "1. Navigate to checklist\n2. View pre-built tasks\n3. Add custom task\n4. Mark a task as done\n5. Delete a task",
         "Checklist works: add/edit/delete/complete tasks, progress percentage updates", "High", "May 7"],
        ["TC-C-004", "Guest List Management", "Manage wedding guests",
         "Wedding plan exists",
         "1. Navigate to guest list\n2. Add a guest (name, email, phone, group)\n3. Edit guest\n4. Delete guest\n5. Check total count",
         "Guest CRUD works correctly, total count accurate", "High", "May 7"],
        ["TC-C-005", "RSVP Tracking", "Track guest RSVP responses",
         "Guests exist in guest list",
         "1. View guest list\n2. Change RSVP status for a guest (Confirmed/Declined/Pending)\n3. Check RSVP summary counts",
         "RSVP statuses tracked per guest, summary shows confirmed/declined/pending counts", "Medium", "May 7"],
    ]),
]

NONFUNCTIONAL_TESTS = [
    ("Performance", [
        ["TC-NF-001", "Page Load Time", "Main pages load within 2 seconds",
         "Application deployed and accessible",
         "1. Clear cache\n2. Navigate to home page, measure load time\n3. Navigate to vendor listing, measure\n4. Navigate to dashboard, measure",
         "All main pages load within 2 seconds (first contentful paint)", "High", "May 10"],
        ["TC-NF-002", "API Response Time", "API endpoints respond quickly",
         "Backend running",
         "1. Use browser dev tools Network tab\n2. Trigger various API calls\n3. Measure response times",
         "95% of API calls respond within 1 second", "High", "May 10"],
    ]),
    ("Security", [
        ["TC-NF-003", "HTTPS Enforcement", "All traffic over HTTPS",
         "Application deployed",
         "1. Try accessing via HTTP\n2. Verify redirect to HTTPS\n3. Check SSL certificate valid",
         "HTTP redirects to HTTPS, valid SSL certificate, no mixed content warnings", "High", "May 10"],
        ["TC-NF-004", "XSS Prevention", "Script injection blocked",
         "Any text input field",
         "1. Enter `<script>alert('xss')</script>` in text fields\n2. Submit form\n3. Verify script not executed on display",
         "Script tags escaped or sanitized, no alert popup, content displayed safely", "High", "May 10"],
        ["TC-NF-005", "SQL Injection Prevention", "SQL injection blocked",
         "Any input that hits the database",
         "1. Enter `'; DROP TABLE users; --` in search/input fields\n2. Submit\n3. Verify database unaffected",
         "Application handles input safely, no database error, parameterized queries used", "High", "May 10"],
        ["TC-NF-006", "Session Cookie Security", "Cookies are secure and HTTP-only",
         "User is logged in",
         "1. Open browser dev tools → Application → Cookies\n2. Inspect session cookie\n3. Check flags",
         "Cookie has HttpOnly, Secure, and SameSite flags set appropriately", "High", "May 10"],
    ]),
    ("Responsiveness", [
        ["TC-NF-007", "Mobile Responsive - Login", "Login page works on mobile",
         "Mobile device or responsive mode",
         "1. Open login page in mobile view (375px width)\n2. Check layout\n3. Try logging in\n4. Verify all elements accessible",
         "Login page fully functional and readable on mobile screen sizes", "High", "May 10"],
        ["TC-NF-008", "Mobile Responsive - Vendor List", "Vendor listing works on mobile",
         "Mobile view, verified vendors exist",
         "1. Open vendor listing on mobile\n2. Check card layout\n3. Try filters\n4. Scroll through results",
         "Vendor cards stack properly, filters accessible, scrolling smooth", "High", "May 10"],
        ["TC-NF-009", "Mobile Responsive - Chat", "Chat works on mobile",
         "Mobile view, conversation exists",
         "1. Open chat on mobile\n2. Send a message\n3. Check layout of conversation list and messages",
         "Chat fully functional on mobile, messages readable, input accessible", "Medium", "May 10"],
    ]),
    ("Browser Compatibility", [
        ["TC-NF-010", "Chrome Compatibility", "App works in Chrome",
         "Chrome browser available",
         "1. Open app in latest Chrome\n2. Navigate through key flows\n3. Check for visual/functional issues",
         "Full functionality in Chrome, no console errors, correct rendering", "High", "May 10"],
        ["TC-NF-011", "Firefox Compatibility", "App works in Firefox",
         "Firefox browser available",
         "1. Open app in latest Firefox\n2. Navigate through key flows\n3. Check for differences",
         "Full functionality in Firefox, consistent with Chrome behavior", "Medium", "May 10"],
        ["TC-NF-012", "Safari/Mobile Browser", "App works in Safari",
         "Safari browser available (macOS/iOS)",
         "1. Open app in Safari\n2. Test login, navigation, chat\n3. Check for Safari-specific issues",
         "Functional in Safari, no WebKit-specific bugs blocking core features", "Medium", "May 10"],
    ]),
]

# ========== ASSIGNMENT DISTRIBUTION ==========
# Sura: Auth (urgent - due tomorrow), Vendor, Real-Time, Admin
# Hanamariam: Budget, AI, Booking, Payment, Review, Couple Planning, Non-Functional

ALL_MODULES = {
    "1. Auth & User Mgmt": ("Sura", AUTH_TESTS),
    "2. Vendor Management": ("Sura", VENDOR_TESTS),
    "3. Real-Time Comms": ("Sura", REALTIME_TESTS),
    "4. Budgeting & Expense": ("Hanamariam", BUDGET_TESTS),
    "5. AI Recommendations": ("Hanamariam", AI_TESTS),
    "6. Booking & Scheduling": ("Hanamariam", BOOKING_TESTS),
    "7. Payment Integration": ("Hanamariam", PAYMENT_TESTS),
    "8. Review & Feedback": ("Hanamariam", REVIEW_TESTS),
    "9. Admin & Analytics": ("Sura", ADMIN_TESTS),
    "10. Couple Planning": ("Hanamariam", COUPLE_TESTS),
    "11. Non-Functional": ("Hanamariam", NONFUNCTIONAL_TESTS),
}


def style_cell(cell, fill=None, font=None, align=None, border=THIN_BORDER):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border


def write_test_sheet(wb, sheet_name, assignee, sections):
    ws = wb.create_sheet(title=sheet_name)

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col_idx, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        style_cell(cell, HEADER_FILL, HEADER_FONT,
                   Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    # Result dropdown validation
    result_dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,Not Implemented,Blocked,Skipped"',
        allow_blank=True
    )
    result_dv.error = "Please select: Pass, Fail, Not Implemented, Blocked, or Skipped"
    result_dv.errorTitle = "Invalid Result"
    result_dv.prompt = "Select test result"
    result_dv.promptTitle = "Test Result"
    ws.add_data_validation(result_dv)

    # Priority dropdown validation
    priority_dv = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(priority_dv)

    row = 2
    for section_title, test_cases in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
        cell = ws.cell(row=row, column=1, value=section_title)
        style_cell(cell, SUBHEADER_FILL, SUBHEADER_FONT, Alignment(vertical="center"))
        for col in range(2, len(COLUMNS) + 1):
            style_cell(ws.cell(row=row, column=col), fill=SUBHEADER_FILL)
        row += 1

        for tc in test_cases:
            test_id, test_case_name, description, preconditions, steps, expected, priority, deadline = tc
            values = [test_id, sheet_name.split(". ", 1)[-1] if ". " in sheet_name else sheet_name,
                      test_case_name, description, preconditions, steps, expected,
                      assignee, priority, deadline, "", ""]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = THIN_BORDER

            ws.row_dimensions[row].height = 60

            if row % 2 == 0:
                for col_idx in range(1, len(COLUMNS) + 1):
                    ws.cell(row=row, column=col_idx).fill = ALT_ROW_FILL

            # Assignee coloring
            assignee_cell = ws.cell(row=row, column=8)
            if assignee in ASSIGNEE_STYLES:
                assignee_cell.fill = ASSIGNEE_STYLES[assignee]
            assignee_cell.font = BOLD_BODY
            assignee_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Priority coloring
            priority_cell = ws.cell(row=row, column=9)
            if priority == "High":
                priority_cell.fill = PRIORITY_HIGH
            elif priority == "Medium":
                priority_cell.fill = PRIORITY_MED
            else:
                priority_cell.fill = PRIORITY_LOW
            priority_cell.font = BOLD_BODY
            priority_cell.alignment = Alignment(horizontal="center", vertical="center")
            priority_dv.add(priority_cell)

            # Deadline coloring for auth (May 2 = tomorrow)
            deadline_cell = ws.cell(row=row, column=10)
            if "May 2" in deadline:
                deadline_cell.fill = DEADLINE_FILL
                deadline_cell.font = DEADLINE_FONT
            deadline_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add result dropdown to result column
            result_cell = ws.cell(row=row, column=11)
            result_dv.add(result_cell)
            result_cell.alignment = Alignment(horizontal="center", vertical="center")

            row += 1
        row += 1


def create_overview(wb):
    ws = wb.active
    ws.title = "Overview"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 14

    row = 1
    ws.cell(row=row, column=2, value="TWEDAR - Manual Test Plan").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=2, value="Comprehensive Manual Testing Assignment for Sura & Hanamariam").font = SUBTITLE_FONT
    row += 2

    # Urgent notice
    cell = ws.cell(row=row, column=2, value="⚠️  URGENT: Auth tests (TC-A-001 to TC-A-022) must be completed by May 2!")
    cell.font = Font(name="Segoe UI", size=11, bold=True, color="FF0000")
    row += 2

    # Instructions
    ws.cell(row=row, column=2, value="HOW TO USE THIS SHEET").font = BOLD_BODY
    row += 1
    instructions = [
        "1. Navigate to your assigned module sheet(s)",
        "2. Follow the 'How to Test' steps for each test case",
        "3. Compare actual result with 'Expected Result' column",
        "4. Select result from the DROPDOWN in 'Result' column: Pass / Fail / Not Implemented / Blocked / Skipped",
        "5. Add any observations in the 'Notes' column",
        "6. High priority tests should be completed first",
    ]
    for instr in instructions:
        ws.cell(row=row, column=2, value=instr).font = BODY_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="TESTER LEGEND").font = BOLD_BODY
    row += 1
    for name, fill in ASSIGNEE_STYLES.items():
        cell = ws.cell(row=row, column=2, value=name)
        cell.fill = fill
        cell.font = BOLD_BODY
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="RESULT OPTIONS").font = BOLD_BODY
    row += 1
    result_options = [
        ("Pass", "Test case passed - actual matches expected", PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")),
        ("Fail", "Test case failed - actual differs from expected", PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")),
        ("Not Implemented", "Feature not yet built - cannot test", PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")),
        ("Blocked", "Cannot test due to dependency or environment issue", PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")),
        ("Skipped", "Intentionally skipped with justification in Notes", PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")),
    ]
    for label, desc, fill in result_options:
        cell = ws.cell(row=row, column=2, value=label)
        cell.fill = fill
        cell.font = BOLD_BODY
        cell.border = THIN_BORDER
        ws.cell(row=row, column=3, value=desc).font = BODY_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="PRIORITY LEVELS").font = BOLD_BODY
    row += 1
    for label, fill in [("High", PRIORITY_HIGH), ("Medium", PRIORITY_MED), ("Low", PRIORITY_LOW)]:
        cell = ws.cell(row=row, column=2, value=label)
        cell.fill = fill
        cell.font = BOLD_BODY
        cell.border = THIN_BORDER
        row += 1

    row += 2
    ws.cell(row=row, column=2, value="MODULE ASSIGNMENT SUMMARY").font = BOLD_BODY
    row += 1

    headers = ["Module", "Tester", "Total Tests", "High Priority", "Deadline", "Status"]
    for col_idx, h in enumerate(headers, 2):
        cell = ws.cell(row=row, column=col_idx, value=h)
        style_cell(cell, HEADER_FILL, HEADER_FONT, Alignment(horizontal="center"))
    row += 1

    total_tests = 0
    for sheet_name, (assignee, sections) in ALL_MODULES.items():
        count = sum(len(tcs) for _, tcs in sections)
        high = sum(1 for _, tcs in sections for tc in tcs if tc[6] == "High")
        total_tests += count
        
        # Determine earliest deadline
        deadlines = [tc[7] for _, tcs in sections for tc in tcs]
        earliest = sorted(deadlines)[0] if deadlines else ""
        
        values = [sheet_name, assignee, count, high, earliest, "Not Started"]
        for col_idx, val in enumerate(values, 2):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL

        # Assignee coloring
        ac = ws.cell(row=row, column=3)
        if assignee in ASSIGNEE_STYLES:
            ac.fill = ASSIGNEE_STYLES[assignee]
        ac.font = BOLD_BODY

        # Deadline urgency
        dl_cell = ws.cell(row=row, column=6)
        if "May 2" in earliest:
            dl_cell.fill = DEADLINE_FILL
            dl_cell.font = DEADLINE_FONT

        row += 1

    # Total row
    values = ["TOTAL", "", total_tests, "", "", ""]
    for col_idx, val in enumerate(values, 2):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="1B2A4A")
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        cell.fill = TOTAL_FILL

    row += 2
    ws.cell(row=row, column=2, value="PER-TESTER SUMMARY").font = BOLD_BODY
    row += 1

    p_headers = ["Tester", "Modules Assigned", "Total Tests", "High Priority", "", ""]
    for col_idx, h in enumerate(p_headers, 2):
        cell = ws.cell(row=row, column=col_idx, value=h)
        style_cell(cell, HEADER_FILL, HEADER_FONT, Alignment(horizontal="center"))
    row += 1

    for name in ["Sura", "Hanamariam"]:
        modules = [k for k, (a, _) in ALL_MODULES.items() if a == name]
        count = sum(sum(len(tcs) for _, tcs in secs) for k, (a, secs) in ALL_MODULES.items() if a == name)
        high = sum(sum(1 for tc in tcs if tc[6] == "High") for k, (a, secs) in ALL_MODULES.items() if a == name for _, tcs in secs)
        
        values = [name, len(modules), count, high, "", ""]
        for col_idx, val in enumerate(values, 2):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        nc = ws.cell(row=row, column=2)
        if name in ASSIGNEE_STYLES:
            nc.fill = ASSIGNEE_STYLES[name]
        nc.font = BOLD_BODY
        row += 1


def main():
    wb = Workbook()
    create_overview(wb)

    for sheet_name, (assignee, sections) in ALL_MODULES.items():
        write_test_sheet(wb, sheet_name, assignee, sections)

    output_path = "/home/chera/Public/my_stuffs/astu/twedar/Twedar_Manual_Test_Plan.xlsx"
    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Sheets: {wb.sheetnames}")
    print()

    total = 0
    for name, (assignee, sections) in ALL_MODULES.items():
        count = sum(len(tcs) for _, tcs in sections)
        total += count
        high = sum(1 for _, tcs in sections for tc in tcs if tc[6] == "High")
        print(f"  {name} -> {assignee} ({count} tests, {high} high-priority)")
    
    print(f"\n  TOTAL: {total} test cases")
    print(f"  Sura: {sum(sum(len(tcs) for _, tcs in s) for _, (a, s) in ALL_MODULES.items() if a == 'Sura')} tests")
    print(f"  Hanamariam: {sum(sum(len(tcs) for _, tcs in s) for _, (a, s) in ALL_MODULES.items() if a == 'Hanamariam')} tests")


if __name__ == "__main__":
    main()
