#!/usr/bin/env python3
"""
Twedar Test Plan - Separate Excel Workbook Generator
Generates a comprehensive test plan with all test cases organized by type
(Unit, Integration, UAT, Performance, Security).
Assigned to Sura and Hanamariam with dropdown for Pass/Fail/Not Implemented results.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# === Color Definitions ===
HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
SUBHEADER_FONT = Font(name="Segoe UI", size=10, bold=True, color="1B2A4A")

PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PASS_FONT = Font(name="Segoe UI", size=10, color="006100")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FAIL_FONT = Font(name="Segoe UI", size=10, color="9C0006")
NOT_TESTED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NOT_TESTED_FONT = Font(name="Segoe UI", size=10, color="9C0006")
BLOCKED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLOCKED_FONT = Font(name="Segoe UI", size=10, color="9C5700")
NOT_IMPL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
NOT_IMPL_FONT = Font(name="Segoe UI", size=10, color="595959")

SURA_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HANA_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TITLE_FONT = Font(name="Segoe UI", size=16, bold=True, color="1B2A4A")
SUBTITLE_FONT = Font(name="Segoe UI", size=12, bold=False, color="4472C4")
BODY_FONT = Font(name="Segoe UI", size=10, color="333333")
BOLD_BODY = Font(name="Segoe UI", size=10, bold=True, color="333333")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

STATUS_COLORS = {
    "Not Tested": (NOT_TESTED_FILL, NOT_TESTED_FONT),
    "Passed": (PASS_FILL, PASS_FONT),
    "Failed": (FAIL_FILL, FAIL_FONT),
    "Blocked": (BLOCKED_FILL, BLOCKED_FONT),
    "Not Implemented": (NOT_IMPL_FILL, NOT_IMPL_FONT),
}

ASSIGNEE_STYLES = {
    "Sura": SURA_FILL,
    "Hanamariam": HANA_FILL,
}

TEST_COLUMNS = ["Test ID", "Test Case Name", "Description / Steps", "Expected Result",
                "Tester", "Result", "Notes", "Test Type", "Module", "Priority", "Prerequisites", "Doc Reference"]
TEST_COL_WIDTHS = [9, 32, 55, 45, 14, 18, 35, 14, 20, 13, 35, 14]

# Tester assignment per sheet
SHEET_ASSIGNMENTS = {
    "Unit - Auth": "Sura",
    "Unit - Vendor": "Sura",
    "Unit - Admin": "Sura",
    "Unit - Realtime": "Sura",
    "Unit - Budget": "Hanamariam",
    "Unit - Booking": "Hanamariam",
    "Unit - Payment": "Hanamariam",
    "Unit - Review": "Hanamariam",
    "Integration": "Sura",
    "UAT": "Hanamariam",
    "Performance": "Hanamariam",
    "Security": "Sura",
}


def style_header_row(ws, cols, row=1):
    for col_idx in range(1, len(cols) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_test_headers(ws):
    for col_idx, name in enumerate(TEST_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=name)
    style_header_row(ws, TEST_COLUMNS)
    set_col_widths(ws, TEST_COL_WIDTHS)
    ws.freeze_panes = "A2"


def write_section_header(ws, row_idx, title, num_cols):
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=num_cols)
    cell = ws.cell(row=row_idx, column=1, value=title)
    cell.fill = SUBHEADER_FILL
    cell.font = SUBHEADER_FONT
    cell.alignment = Alignment(vertical="center")
    for col in range(2, num_cols + 1):
        ws.cell(row=row_idx, column=col).fill = SUBHEADER_FILL
    return row_idx + 1


def write_test_row(ws, row_idx, data, tester):
    """Write a test row with tester assignment and result/notes columns."""
    # data format: [id, name, desc, expected, old_status, test_type, module, priority, prereqs, doc_ref]
    # New column layout: ID, Name, Desc, Expected, Tester, Result, Notes, Type, Module, Priority, Prereqs, DocRef
    new_row = [
        data[0],  # Test ID
        data[1],  # Test Case Name
        data[2],  # Description / Steps
        data[3],  # Expected Result
        tester,   # Tester (assigned)
        "",       # Result (dropdown - empty initially)
        "",       # Notes (free text)
        data[5],  # Test Type
        data[6],  # Module
        data[7],  # Priority
        data[8],  # Prerequisites
        data[9],  # Doc Reference
    ]

    for col_idx, value in enumerate(new_row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    if row_idx % 2 == 0:
        for col_idx in range(1, len(TEST_COLUMNS) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = ALT_ROW_FILL

    # Tester coloring (column 5)
    tester_cell = ws.cell(row=row_idx, column=5)
    if tester in ASSIGNEE_STYLES:
        tester_cell.fill = ASSIGNEE_STYLES[tester]
    tester_cell.font = BOLD_BODY
    tester_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Result column styling (column 6) - initially empty with center alignment
    result_cell = ws.cell(row=row_idx, column=6)
    result_cell.alignment = Alignment(horizontal="center", vertical="center")


def populate_test_sheet(ws, sections, tester):
    write_test_headers(ws)
    row = 2
    for section_title, items in sections:
        row = write_section_header(ws, row, section_title, len(TEST_COLUMNS))
        for item in items:
            write_test_row(ws, row, item, tester)
            row += 1
        row += 1

    # Result dropdown (column 6)
    result_dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,Not Implemented,Blocked,Skipped"',
        allow_blank=True
    )
    result_dv.error = "Please select: Pass, Fail, Not Implemented, Blocked, or Skipped"
    result_dv.errorTitle = "Invalid Result"
    result_dv.prompt = "Select test result"
    result_dv.promptTitle = "Test Result"
    ws.add_data_validation(result_dv)
    for r in range(2, row):
        result_dv.add(ws.cell(row=r, column=6))


# ========== TEST DATA ==========

UNIT_AUTH_TESTS = [
    ("Authentication - Registration", [
        ["UT-01", "Register with valid credentials", "1. Send POST /api/auth/sign-up with valid email, password, accountType='couple'\n2. Verify response status\n3. Check database for new user record", "201 Created. User record exists with hashed password, emailVerified=false, role='couple'.", "Not Tested", "Unit", "Auth", "Must Have", "Database connection, email service mock", "3.2.1"],
        ["UT-02", "Register with duplicate email", "1. Create user with email X\n2. Attempt POST /api/auth/sign-up with same email X", "409 Conflict with error message 'Email already exists'.", "Not Tested", "Unit", "Auth", "Must Have", "Existing user in database", "3.2.1"],
        ["UT-03", "Register with missing email", "1. Send POST /api/auth/sign-up without email field", "400 Bad Request with validation error for email field.", "Not Tested", "Unit", "Auth", "Must Have", "None", "3.2.1"],
        ["UT-04", "Register with weak password", "1. Send POST /api/auth/sign-up with password '123'", "400 Bad Request with password complexity error.", "Not Tested", "Unit", "Auth", "Must Have", "None", "3.2.1"],
        ["UT-05", "Register with invalid role", "1. Send POST /api/auth/sign-up with accountType='superadmin'", "400 Bad Request. Only 'couple' and 'vendor' allowed at registration.", "Not Tested", "Unit", "Auth", "Must Have", "None", "3.2.1"],
        ["UT-06", "Vendor registration creates organization", "1. Register with accountType='vendor'\n2. Check organization table", "User created with role='vendor'. Organization auto-created with user as owner.", "Not Tested", "Unit", "Auth", "Must Have", "Database connection", "4.11.2"],
    ]),
    ("Authentication - Login", [
        ["UT-07", "Login with correct credentials", "1. Register and verify user\n2. POST /api/auth/sign-in with correct email/password", "200 OK. Set-Cookie header with httpOnly session cookie. Response contains user data.", "Not Tested", "Unit", "Auth", "Must Have", "Verified user in database", "3.4.2 UC2"],
        ["UT-08", "Login with wrong password", "1. POST /api/auth/sign-in with valid email but wrong password", "401 Unauthorized. No cookie set.", "Not Tested", "Unit", "Auth", "Must Have", "Existing user", "3.4.2 UC2"],
        ["UT-09", "Login with non-existent email", "1. POST /api/auth/sign-in with unregistered email", "401 Unauthorized with generic error (no user enumeration).", "Not Tested", "Unit", "Auth", "Must Have", "None", "3.4.2 UC2"],
        ["UT-10", "Login with unverified email", "1. Register user (don't verify)\n2. Attempt login", "Error response indicating email verification required.", "Not Tested", "Unit", "Auth", "Must Have", "Unverified user", "3.2.1"],
        ["UT-11", "Login with banned account", "1. Ban user via admin\n2. Attempt login", "401/403 with message indicating account is banned.", "Not Tested", "Unit", "Auth", "Must Have", "Banned user record", "4.11.3"],
    ]),
    ("Authentication - Email Verification", [
        ["UT-12", "Verify email with valid token", "1. Register user (generates verification token)\n2. GET /api/auth/verify?token=<valid_token>", "200 OK. User emailVerified field updated to true.", "Not Tested", "Unit", "Auth", "Must Have", "User with pending verification", "3.2.1"],
        ["UT-13", "Verify email with expired token", "1. Use an expired/invalid verification token", "400 Bad Request with 'Invalid or expired token' error.", "Not Tested", "Unit", "Auth", "Must Have", "Expired token", "3.2.1"],
        ["UT-14", "Verify already verified email", "1. Verify email\n2. Use same token again", "Already verified or token consumed - no error but no state change.", "Not Tested", "Unit", "Auth", "Should Have", "Verified user", "3.2.1"],
    ]),
    ("Authentication - Password Reset", [
        ["UT-15", "Request password reset for valid email", "1. POST /api/auth/forgot-password with registered email", "200 OK. Reset email sent with valid token link.", "Not Tested", "Unit", "Auth", "Must Have", "Verified user, email mock", "3.2.1"],
        ["UT-16", "Request password reset for non-existent email", "1. POST /api/auth/forgot-password with unregistered email", "200 OK (no user enumeration). No email sent.", "Not Tested", "Unit", "Auth", "Must Have", "Email mock", "3.2.1"],
        ["UT-17", "Reset password with valid token", "1. Request reset\n2. POST /api/auth/reset-password with valid token + new password", "200 OK. Password hash updated. Old password no longer works.", "Not Tested", "Unit", "Auth", "Must Have", "Valid reset token", "3.2.1"],
        ["UT-18", "Reset password with expired token", "1. Use an expired reset token", "400 Bad Request with 'Token expired' error.", "Not Tested", "Unit", "Auth", "Must Have", "Expired token", "3.2.1"],
    ]),
    ("Authentication - Session & Middleware", [
        ["UT-19", "Access protected route with valid session", "1. Login (get cookie)\n2. GET /api/v1/vendor/profile with cookie", "200 OK. Request passes through requireAuth middleware.", "Not Tested", "Unit", "Auth", "Must Have", "Authenticated session", "4.11.1"],
        ["UT-20", "Access protected route without session", "1. GET /api/v1/vendor/profile without cookie", "401 Unauthorized.", "Not Tested", "Unit", "Auth", "Must Have", "None", "4.11.1"],
        ["UT-21", "Access vendor route as couple", "1. Login as couple\n2. GET /api/v1/vendor/profile", "403 Forbidden. Role mismatch.", "Not Tested", "Unit", "Auth", "Must Have", "Couple session", "4.11.4"],
        ["UT-22", "Access admin route as vendor", "1. Login as vendor\n2. GET /api/v1/admin/vendors", "403 Forbidden.", "Not Tested", "Unit", "Auth", "Must Have", "Vendor session", "4.11.4"],
        ["UT-23", "Session cookie properties", "1. Login\n2. Inspect Set-Cookie header", "Cookie is httpOnly, SameSite=lax, path=/. Secure flag in production.", "Not Tested", "Unit", "Auth", "Must Have", "None", "4.11.1"],
    ]),
]

UNIT_VENDOR_TESTS = [
    ("Vendor Profile CRUD", [
        ["UT-24", "Create vendor profile - valid data", "1. Login as vendor\n2. POST /api/v1/vendor/profile with business_name, category, description, location", "201 Created. Profile record in database with status='draft'.", "Not Tested", "Unit", "Vendor", "Must Have", "Authenticated vendor", "3.2.2"],
        ["UT-25", "Create vendor profile - duplicate", "1. Vendor already has profile\n2. POST /api/v1/vendor/profile again", "409 Conflict. Only one profile per vendor.", "Not Tested", "Unit", "Vendor", "Must Have", "Vendor with existing profile", "3.2.2"],
        ["UT-26", "Update vendor profile", "1. PATCH /api/v1/vendor/profile with partial fields (e.g., description only)", "200 OK. Only specified fields updated; others unchanged.", "Not Tested", "Unit", "Vendor", "Must Have", "Vendor with profile", "3.2.2"],
        ["UT-27", "Get own vendor profile", "1. GET /api/v1/vendor/profile as authenticated vendor", "200 OK. Returns full profile with all fields including status.", "Not Tested", "Unit", "Vendor", "Must Have", "Vendor with profile", "3.2.2"],
        ["UT-28", "Get profile as non-vendor", "1. Login as couple\n2. GET /api/v1/vendor/profile", "403 Forbidden.", "Not Tested", "Unit", "Vendor", "Must Have", "Couple session", "4.11.4"],
        ["UT-29", "Update profile with location coordinates", "1. PATCH /api/v1/vendor/profile with latitude, longitude", "200 OK. Geo coordinates stored correctly.", "Not Tested", "Unit", "Vendor", "Must Have", "Vendor with profile", "3.2.2"],
        ["UT-30", "Update profile with price range", "1. PATCH with price_range_min=5000, price_range_max=50000", "200 OK. Price range stored and retrievable.", "Not Tested", "Unit", "Vendor", "Must Have", "Vendor with profile", "3.2.2"],
    ]),
    ("Vendor Documents", [
        ["UT-31", "Upload document - valid file", "1. POST /api/v1/vendor/documents with PDF/image file and document_type", "201 Created. File uploaded to Supabase. Document record in DB with URL.", "Not Tested", "Unit", "Vendor", "Must Have", "Vendor with profile, Supabase mock", "3.2.3"],
        ["UT-32", "Upload document - invalid type", "1. POST /api/v1/vendor/documents with .exe file", "400 Bad Request. Unsupported file type.", "Not Tested", "Unit", "Vendor", "Must Have", "Vendor with profile", "3.2.3"],
        ["UT-33", "Upload document - oversized file", "1. POST /api/v1/vendor/documents with file > max size limit", "413 Payload Too Large.", "Not Tested", "Unit", "Vendor", "Must Have", "Vendor with profile", "3.2.3"],
        ["UT-34", "Delete own document", "1. DELETE /api/v1/vendor/documents/:id for own document", "200 OK. File removed from Supabase. DB record deleted.", "Not Tested", "Unit", "Vendor", "Must Have", "Vendor with uploaded document", "3.2.3"],
        ["UT-35", "Delete another vendor's document", "1. DELETE /api/v1/vendor/documents/:id for another vendor's doc", "403 or 404. Cannot delete others' documents.", "Not Tested", "Unit", "Vendor", "Must Have", "Two vendors with documents", "3.2.3"],
    ]),
    ("Vendor Verification Workflow", [
        ["UT-36", "Submit profile for verification", "1. POST /api/v1/vendor/profile/submit when status='draft' and documents uploaded", "200 OK. Status transitions to 'pending_verification'.", "Not Tested", "Unit", "Vendor", "Must Have", "Draft profile with documents", "3.2.3"],
        ["UT-37", "Submit without documents", "1. POST /api/v1/vendor/profile/submit with no documents uploaded", "400 Bad Request. Documents required for verification.", "Not Tested", "Unit", "Vendor", "Must Have", "Draft profile, no documents", "3.2.3"],
        ["UT-38", "Submit already pending profile", "1. Profile already in 'pending_verification'\n2. POST /api/v1/vendor/profile/submit again", "400 Bad Request. Invalid status transition.", "Not Tested", "Unit", "Vendor", "Must Have", "Pending profile", "3.2.3"],
        ["UT-39", "Resubmit after rejection", "1. Profile in 'rejected' status\n2. Update profile\n3. POST /api/v1/vendor/profile/submit", "200 OK. Status transitions from 'rejected' back to 'pending_verification'.", "Not Tested", "Unit", "Vendor", "Should Have", "Rejected profile", "3.2.3"],
        ["UT-40", "Status machine - invalid transitions", "1. Attempt transition from 'draft' directly to 'verified'", "Error thrown. Only valid transitions allowed (draft->pending, pending->verified/rejected).", "Not Tested", "Unit", "Vendor", "Must Have", "Draft profile", "3.2.3"],
    ]),
    ("Public Vendor Listing", [
        ["UT-41", "List vendors - no filters", "1. GET /api/v1/vendors without query params", "200 OK. Returns paginated list of verified vendors only.", "Not Tested", "Unit", "Vendor", "Must Have", "Verified vendors in DB", "US-C-004"],
        ["UT-42", "List vendors - category filter", "1. GET /api/v1/vendors?category=photographer", "200 OK. Only vendors with category='photographer' returned.", "Not Tested", "Unit", "Vendor", "Must Have", "Vendors with varied categories", "US-C-004"],
        ["UT-43", "List vendors - location filter", "1. GET /api/v1/vendors?location=Addis Ababa", "200 OK. Only vendors in specified location returned.", "Not Tested", "Unit", "Vendor", "Must Have", "Vendors with varied locations", "US-C-004"],
        ["UT-44", "List vendors - search keyword", "1. GET /api/v1/vendors?search=wedding photography", "200 OK. Vendors matching search in name/description returned.", "Not Tested", "Unit", "Vendor", "Must Have", "Vendors with varied descriptions", "US-C-004"],
        ["UT-45", "List vendors - pagination", "1. GET /api/v1/vendors?page=2&limit=5", "200 OK. Returns page 2 with max 5 items. Pagination metadata included.", "Not Tested", "Unit", "Vendor", "Must Have", "10+ verified vendors", "US-C-004"],
        ["UT-46", "Get vendor detail - valid ID", "1. GET /api/v1/vendors/:id for verified vendor", "200 OK. Full profile with all public fields.", "Not Tested", "Unit", "Vendor", "Must Have", "Verified vendor", "3.4.2 UC8"],
        ["UT-47", "Get vendor detail - unverified vendor", "1. GET /api/v1/vendors/:id for unverified vendor", "404 Not Found. Unverified vendors not publicly accessible.", "Not Tested", "Unit", "Vendor", "Must Have", "Unverified vendor", "US-C-007"],
    ]),
]

UNIT_ADMIN_TESTS = [
    ("Admin Vendor Management", [
        ["UT-48", "Admin list all vendors", "1. Login as admin\n2. GET /api/v1/admin/vendors", "200 OK. Returns all vendors regardless of status with status field.", "Not Tested", "Unit", "Admin", "Must Have", "Admin session, vendors in DB", "US-A-005"],
        ["UT-49", "Non-admin access admin vendors", "1. Login as couple\n2. GET /api/v1/admin/vendors", "403 Forbidden.", "Not Tested", "Unit", "Admin", "Must Have", "Couple session", "4.11.4"],
        ["UT-50", "Approve pending vendor", "1. POST /api/v1/admin/vendors/:id/approve for pending vendor", "200 OK. Vendor status changes to 'verified'. Vendor appears in public listing.", "Not Tested", "Unit", "Admin", "Must Have", "Admin session, pending vendor", "3.4.2 UC14"],
        ["UT-51", "Approve non-pending vendor", "1. POST /api/v1/admin/vendors/:id/approve for already verified vendor", "400 Bad Request. Invalid status transition.", "Not Tested", "Unit", "Admin", "Must Have", "Admin session, verified vendor", "3.4.2 UC14"],
        ["UT-52", "Reject vendor with reason", "1. POST /api/v1/admin/vendors/:id/reject with body { reason: 'Invalid license' }", "200 OK. Status changes to 'rejected'. Rejection reason stored.", "Not Tested", "Unit", "Admin", "Must Have", "Admin session, pending vendor", "3.4.2 UC14"],
        ["UT-53", "Reject vendor without reason", "1. POST /api/v1/admin/vendors/:id/reject without reason field", "400 Bad Request. Reason is required.", "Not Tested", "Unit", "Admin", "Must Have", "Admin session, pending vendor", "3.4.2 UC14"],
        ["UT-54", "Suspend verified vendor", "1. POST /api/v1/admin/vendors/:id/suspend for verified vendor", "200 OK. Status changes to 'suspended'. Vendor removed from public listing.", "Not Tested", "Unit", "Admin", "Must Have", "Admin session, verified vendor", "4.11.3"],
        ["UT-55", "Reinstate suspended vendor", "1. POST /api/v1/admin/vendors/:id/reinstate for suspended vendor", "200 OK. Status changes back to 'verified'. Vendor re-appears in public listing.", "Not Tested", "Unit", "Admin", "Should Have", "Admin session, suspended vendor", "4.11.3"],
        ["UT-56", "Deactivate vendor", "1. POST /api/v1/admin/vendors/:id/deactivate", "200 OK. Vendor permanently deactivated.", "Not Tested", "Unit", "Admin", "Should Have", "Admin session, verified vendor", "4.11.3"],
    ]),
]

UNIT_REALTIME_TESTS = [
    ("Conversation Management", [
        ["UT-57", "Create conversation - valid", "1. POST /api/v1/conversations with { participantId: vendor_user_id }", "201 Created. Conversation record with both participant IDs. Returns conversation object.", "Not Tested", "Unit", "Realtime", "Must Have", "Authenticated couple, existing vendor", "3.2.6"],
        ["UT-58", "Create duplicate conversation", "1. POST /api/v1/conversations with same participantId as existing conversation", "200 OK. Returns existing conversation (no duplicate created).", "Not Tested", "Unit", "Realtime", "Must Have", "Existing conversation pair", "3.2.6"],
        ["UT-59", "Create conversation with self", "1. POST /api/v1/conversations with own user ID as participantId", "400 Bad Request. Cannot create conversation with self.", "Not Tested", "Unit", "Realtime", "Must Have", "Authenticated user", "3.2.6"],
        ["UT-60", "List conversations", "1. GET /api/v1/conversations as authenticated user", "200 OK. Returns only conversations where user is a participant. Sorted by last message.", "Not Tested", "Unit", "Realtime", "Must Have", "User with conversations", "3.2.6"],
        ["UT-61", "Get messages for own conversation", "1. GET /api/v1/conversations/:id/messages for user's conversation", "200 OK. Returns messages in chronological order.", "Not Tested", "Unit", "Realtime", "Must Have", "Conversation with messages", "3.2.6"],
        ["UT-62", "Get messages for others' conversation", "1. GET /api/v1/conversations/:id/messages for conversation user is NOT part of", "403 Forbidden.", "Not Tested", "Unit", "Realtime", "Must Have", "Other users' conversation", "3.2.6"],
    ]),
    ("Socket.IO - Connection & Auth", [
        ["UT-63", "Socket connect with valid session", "1. Connect to Socket.IO server with valid session cookie", "Connection accepted. User joins user:{userId} room.", "Not Tested", "Unit", "Realtime", "Must Have", "Valid session cookie", "3.2.6"],
        ["UT-64", "Socket connect without session", "1. Connect to Socket.IO server without cookie", "Connection rejected with authentication error.", "Not Tested", "Unit", "Realtime", "Must Have", "None", "3.2.6"],
        ["UT-65", "Socket disconnect cleanup", "1. Connect\n2. Disconnect\n3. Verify room membership removed", "User removed from all rooms. Presence updated to offline.", "Not Tested", "Unit", "Realtime", "Must Have", "Connected socket", "3.2.6"],
    ]),
    ("Socket.IO - Chat Events", [
        ["UT-66", "Send message (chat:send)", "1. Emit chat:send with { conversationId, content }\n2. Check database and receiver", "Message saved to chat_message table. Receiver gets chat:newMessage event.", "Not Tested", "Unit", "Realtime", "Must Have", "Two connected users in conversation", "3.2.6"],
        ["UT-67", "Send message to invalid conversation", "1. Emit chat:send with non-existent conversationId", "Error callback. Message not saved.", "Not Tested", "Unit", "Realtime", "Must Have", "Connected user", "3.2.6"],
        ["UT-68", "Typing indicator (chat:typing)", "1. Emit chat:typing with { conversationId }\n2. Check other participant receives event", "Other participant receives chat:userTyping event.", "Not Tested", "Unit", "Realtime", "Should Have", "Two connected users in conversation", "Scenario 4"],
        ["UT-69", "Stop typing (chat:stopTyping)", "1. Emit chat:stopTyping with { conversationId }", "Other participant receives chat:userStoppedTyping event.", "Not Tested", "Unit", "Realtime", "Should Have", "Two connected users in conversation", "Scenario 4"],
        ["UT-70", "Mark as read (chat:markRead)", "1. Emit chat:markRead with { conversationId, messageId }\n2. Check database update", "Message isRead=true in DB. Sender receives chat:messageRead event.", "Not Tested", "Unit", "Realtime", "Should Have", "Unread message in conversation", "Scenario 4"],
    ]),
    ("Notifications", [
        ["UT-71", "Get notifications - authenticated", "1. GET /api/v1/notifications", "200 OK. Returns user's notifications ordered by createdAt desc.", "Not Tested", "Unit", "Realtime", "Must Have", "User with notifications", "3.2.6"],
        ["UT-72", "Mark single notification read", "1. PATCH /api/v1/notifications/:id/read", "200 OK. Notification isRead updated to true.", "Not Tested", "Unit", "Realtime", "Must Have", "Unread notification", "3.2.6"],
        ["UT-73", "Mark all notifications read", "1. PATCH /api/v1/notifications/read-all", "200 OK. All user's notifications marked as read.", "Not Tested", "Unit", "Realtime", "Must Have", "Multiple unread notifications", "3.2.6"],
        ["UT-74", "Real-time notification delivery", "1. Trigger notification (e.g., new message)\n2. Check connected user's socket receives event", "notification:new event emitted to user:{targetId} room with notification data.", "Not Tested", "Unit", "Realtime", "Must Have", "Connected target user", "3.2.6"],
    ]),
]

UNIT_BUDGET_TESTS = [
    ("Budget CRUD (Future Implementation)", [
        ["UT-75", "Create budget", "1. POST /api/v1/budgets with { totalAmount, currency }", "201 Created. Budget record with status 'active'.", "Not Tested", "Unit", "Budget", "Must Have", "Budget module implemented", "3.2.5"],
        ["UT-76", "Create duplicate budget", "1. Couple already has active budget\n2. POST /api/v1/budgets again", "409 Conflict. One active budget per couple.", "Not Tested", "Unit", "Budget", "Must Have", "Budget module implemented", "3.2.5"],
        ["UT-77", "Add category to budget", "1. POST /api/v1/budgets/:id/categories with { name, allocatedAmount }", "201 Created. Category linked to budget.", "Not Tested", "Unit", "Budget", "Must Have", "Budget module implemented", "US-C-009"],
        ["UT-78", "Category allocation exceeds total", "1. Add categories that sum > budget total", "400 Bad Request. Total allocations cannot exceed budget.", "Not Tested", "Unit", "Budget", "Must Have", "Budget module implemented", "3.2.5"],
        ["UT-79", "Add expense to category", "1. POST /api/v1/expenses with { categoryId, amount, description }", "201 Created. Expense linked to category. Budget totals recalculated.", "Not Tested", "Unit", "Budget", "Must Have", "Budget module implemented", "US-C-011"],
        ["UT-80", "Budget alert at 80% threshold", "1. Add expenses totaling 80%+ of category allocation", "Alert notification triggered for approaching limit.", "Not Tested", "Unit", "Budget", "Should Have", "Budget module implemented", "US-C-012"],
        ["UT-81", "Budget summary calculation", "1. GET /api/v1/budgets/:id/summary", "200 OK. Correct totals: spent, remaining, per-category breakdown.", "Not Tested", "Unit", "Budget", "Must Have", "Budget module implemented", "US-C-011"],
    ]),
]

UNIT_BOOKING_TESTS = [
    ("Booking Workflow (Future Implementation)", [
        ["UT-82", "Create booking request", "1. POST /api/v1/bookings with { vendorId, date, requirements }", "201 Created. Booking with status 'pending'. Vendor notified.", "Not Tested", "Unit", "Booking", "Must Have", "Booking module implemented", "3.4.2 UC9"],
        ["UT-83", "Book unverified vendor", "1. POST /api/v1/bookings with unverified vendorId", "400 Bad Request. Can only book verified vendors.", "Not Tested", "Unit", "Booking", "Must Have", "Booking module implemented", "3.4.2 UC9"],
        ["UT-84", "Vendor accepts booking", "1. PATCH /api/v1/bookings/:id/accept as vendor", "200 OK. Status -> 'accepted'. Couple notified.", "Not Tested", "Unit", "Booking", "Must Have", "Booking module implemented", "3.4.2 UC9"],
        ["UT-85", "Vendor declines booking", "1. PATCH /api/v1/bookings/:id/decline with { reason }", "200 OK. Status -> 'declined'. Couple notified with reason.", "Not Tested", "Unit", "Booking", "Must Have", "Booking module implemented", "3.4.2 UC9"],
        ["UT-86", "Double booking prevention", "1. Vendor has accepted booking for date X\n2. Another couple books same vendor for date X", "409 Conflict. Vendor unavailable on that date.", "Not Tested", "Unit", "Booking", "Should Have", "Booking module implemented", "3.2.2"],
        ["UT-87", "Complete booking", "1. PATCH /api/v1/bookings/:id/complete", "200 OK. Status -> 'completed'. Review now allowed.", "Not Tested", "Unit", "Booking", "Must Have", "Booking module implemented", "3.4.2 UC9"],
        ["UT-88", "Booking history for couple", "1. GET /api/v1/bookings as couple", "200 OK. Returns all bookings for this couple with statuses.", "Not Tested", "Unit", "Booking", "Must Have", "Booking module implemented", "US-C-021"],
    ]),
]

UNIT_PAYMENT_TESTS = [
    ("Payment Processing (Future Implementation)", [
        ["UT-89", "Initiate Chapa payment", "1. POST /api/v1/payments/initiate with { bookingId, amount }", "200 OK. Returns Chapa checkout URL. Payment record with status 'pending'.", "Not Tested", "Unit", "Payment", "Must Have", "Payment module implemented, Chapa mock", "3.2.7"],
        ["UT-90", "Payment for non-accepted booking", "1. POST /api/v1/payments/initiate for pending/declined booking", "400 Bad Request. Payment only allowed for accepted bookings.", "Not Tested", "Unit", "Payment", "Must Have", "Payment module implemented", "3.2.7"],
        ["UT-91", "Chapa webhook - success", "1. POST /api/v1/payments/webhook with valid Chapa signature and success payload", "200 OK. Payment status -> 'completed'. Booking updated.", "Not Tested", "Unit", "Payment", "Must Have", "Payment module implemented, valid webhook", "3.2.7"],
        ["UT-92", "Chapa webhook - invalid signature", "1. POST /api/v1/payments/webhook with tampered signature", "401 Unauthorized. Payment unchanged.", "Not Tested", "Unit", "Payment", "Must Have", "Payment module implemented", "3.2.7"],
        ["UT-93", "Payment history", "1. GET /api/v1/payments as authenticated user", "200 OK. Returns user's payment transactions with booking references.", "Not Tested", "Unit", "Payment", "Must Have", "Payment module implemented", "US-C-021"],
    ]),
]

UNIT_REVIEW_TESTS = [
    ("Review System (Future Implementation)", [
        ["UT-94", "Submit review for completed booking", "1. POST /api/v1/reviews with { bookingId, rating: 5, comment: 'Great!' }", "201 Created. Review saved with isApproved=false. Vendor notified.", "Not Tested", "Unit", "Review", "Must Have", "Review module implemented, completed booking", "3.2.8"],
        ["UT-95", "Submit review for non-completed booking", "1. POST /api/v1/reviews for booking with status != 'completed'", "400 Bad Request. Reviews only for completed bookings.", "Not Tested", "Unit", "Review", "Must Have", "Review module implemented", "3.2.8"],
        ["UT-96", "Submit duplicate review", "1. Review already exists for booking\n2. POST /api/v1/reviews for same booking", "409 Conflict. One review per booking.", "Not Tested", "Unit", "Review", "Must Have", "Review module implemented", "Scenario 3"],
        ["UT-97", "Rating out of range", "1. POST /api/v1/reviews with rating=6 or rating=0", "400 Bad Request. Rating must be 1-5.", "Not Tested", "Unit", "Review", "Must Have", "Review module implemented", "3.2.8"],
        ["UT-98", "Vendor rating aggregation", "1. Submit multiple reviews for same vendor\n2. Check vendor avg_rating", "Average correctly calculated. review_count incremented.", "Not Tested", "Unit", "Review", "Must Have", "Review module implemented", "Scenario 3"],
        ["UT-99", "Admin approve review", "1. PATCH /api/v1/admin/reviews/:id/approve", "200 OK. isApproved=true. Review visible on public vendor profile.", "Not Tested", "Unit", "Review", "Must Have", "Review module implemented", "3.4.2 UC15"],
        ["UT-100", "Admin reject review", "1. PATCH /api/v1/admin/reviews/:id/reject with { reason }", "200 OK. Review removed or hidden from public view.", "Not Tested", "Unit", "Review", "Must Have", "Review module implemented", "3.4.2 UC15"],
    ]),
]

INTEGRATION_TESTS = [
    ("End-to-End User Flows", [
        ["IT-01", "Full couple registration flow", "1. Register couple\n2. Verify email via token\n3. Login\n4. Access /dashboard\n5. Verify session cookie", "Complete flow works. User ends up authenticated with couple dashboard access.", "Not Tested", "Integration", "Auth + Frontend", "Must Have", "Clean database, email mock", "Scenario 1"],
        ["IT-02", "Full vendor onboarding flow", "1. Register vendor\n2. Verify email\n3. Login\n4. Create profile\n5. Upload documents\n6. Submit for verification\n7. Admin approves\n8. Vendor appears in public listing", "Complete lifecycle from registration to public visibility.", "Not Tested", "Integration", "Auth + Vendor + Admin", "Must Have", "Clean database, admin account", "Scenario 1, Fig 9"],
        ["IT-03", "Vendor rejection and resubmission", "1. Vendor submits\n2. Admin rejects with reason\n3. Vendor sees reason\n4. Vendor updates and resubmits\n5. Admin approves", "Rejection reason visible. Resubmission transitions correctly.", "Not Tested", "Integration", "Vendor + Admin", "Should Have", "Pending vendor, admin account", "US-V-004"],
        ["IT-04", "Couple discovers and contacts vendor", "1. Couple searches vendors\n2. Filters by category\n3. Views vendor detail\n4. Starts conversation\n5. Sends message\n6. Vendor receives message", "Search to chat flow works end-to-end.", "Not Tested", "Integration", "Vendor + Realtime", "Must Have", "Verified vendor, couple account", "Scenario 2"],
        ["IT-05", "Real-time chat exchange", "1. Couple sends message\n2. Vendor receives in real-time\n3. Vendor replies\n4. Couple receives reply\n5. Read receipts update", "Bidirectional real-time messaging with persistence and receipts.", "Not Tested", "Integration", "Realtime", "Must Have", "Two connected users, existing conversation", "Scenario 4"],
        ["IT-06", "Notification lifecycle", "1. Action triggers notification\n2. User receives in real-time\n3. User marks as read\n4. Notification status persists", "Notifications created, delivered, and persisted correctly.", "Not Tested", "Integration", "Realtime", "Must Have", "Connected users", "3.2.6"],
    ]),
    ("End-to-End Booking Flow (Future)", [
        ["IT-07", "Complete booking lifecycle", "1. Couple books vendor\n2. Vendor accepts\n3. Couple pays\n4. Service delivered\n5. Booking marked complete\n6. Couple reviews vendor\n7. Rating updates", "Full booking-to-review cycle with payment.", "Not Tested", "Integration", "Booking + Payment + Review", "Must Have", "All modules implemented", "Scenario 2, 3"],
        ["IT-08", "Budget-payment integration", "1. Create budget\n2. Make payment for booking\n3. Payment auto-added as expense\n4. Budget totals reflect payment", "Payment automatically appears in budget tracking.", "Not Tested", "Integration", "Payment + Budget", "Should Have", "Payment + Budget modules implemented", "3.2.7"],
        ["IT-09", "AI recommendation with interaction data", "1. Couple books vendor A\n2. Couple reviews vendor A\n3. Request recommendations\n4. Similar vendors ranked higher", "Collaborative filtering incorporates booking/review history.", "Not Tested", "Integration", "AI + Booking + Review", "Should Have", "AI + Booking + Review modules implemented", "Scenario 5"],
    ]),
    ("Cross-Role Security", [
        ["IT-10", "Couple cannot access vendor endpoints", "1. Login as couple\n2. Attempt all /api/v1/vendor/* endpoints", "All return 403 Forbidden.", "Not Tested", "Integration", "Auth", "Must Have", "Couple account", "4.11.4"],
        ["IT-11", "Vendor cannot access admin endpoints", "1. Login as vendor\n2. Attempt all /api/v1/admin/* endpoints", "All return 403 Forbidden.", "Not Tested", "Integration", "Auth", "Must Have", "Vendor account", "4.11.4"],
        ["IT-12", "Couple cannot access other couple's data", "1. Login as couple A\n2. Attempt to read couple B's conversations/budget", "403 or 404. No cross-user data leakage.", "Not Tested", "Integration", "Auth + Realtime", "Must Have", "Two couple accounts", "3.3.3"],
        ["IT-13", "Vendor cannot read other vendor's documents", "1. Login as vendor A\n2. Attempt to access vendor B's documents", "403 or 404.", "Not Tested", "Integration", "Auth + Vendor", "Must Have", "Two vendor accounts", "3.3.3"],
    ]),
]

UAT_TESTS = [
    ("UAT - Couple Experience", [
        ["UAT-01", "Couple registration to dashboard", "1. Open registration page\n2. Fill form (name, email, password, role=couple)\n3. Submit\n4. Check email for verification link\n5. Click link\n6. Login\n7. Verify dashboard loads with nav items", "Smooth registration, verification email arrives, dashboard loads with Budget/Checklist/Guests/Vendors/Messages nav.", "Not Tested", "UAT", "Auth + UI", "Must Have", "Running frontend + backend", "Scenario 1"],
        ["UAT-02", "Google social login", "1. Click 'Continue with Google' on login page\n2. Complete Google OAuth\n3. Verify redirected to couple dashboard", "OAuth flow completes. User created as couple. Dashboard loads.", "Not Tested", "UAT", "Auth + UI", "Must Have", "Google OAuth configured", "3.2.1"],
        ["UAT-03", "Browse vendor directory", "1. Navigate to 'Find Vendors'\n2. See vendor cards grid\n3. Use category filter\n4. Use location filter\n5. Use search bar\n6. Verify results update", "Filters work correctly. Only verified vendors shown. Cards display name, category, rating, location.", "Not Tested", "UAT", "Vendor + UI", "Must Have", "Verified vendors in DB", "US-C-004"],
        ["UAT-04", "View vendor detail page", "1. Click on vendor card\n2. View full profile page\n3. See business info, portfolio, pricing, location map", "All vendor information displayed clearly. Map shows correct location. Pricing visible.", "Not Tested", "UAT", "Vendor + UI", "Must Have", "Verified vendor with complete profile", "3.4.2 UC8"],
        ["UAT-05", "Start and use chat", "1. Click 'Message' on vendor profile\n2. Type message\n3. Send\n4. Verify message appears\n5. Receive vendor reply\n6. See typing indicator\n7. See read receipts", "Chat is real-time. Messages persist on page refresh. Typing/read indicators work.", "Not Tested", "UAT", "Realtime + UI", "Must Have", "Couple + vendor accounts, both online", "Scenario 4"],
        ["UAT-06", "Receive notifications", "1. Vendor sends message\n2. Check notification bell shows badge\n3. Click bell\n4. See notification dropdown\n5. Click notification\n6. Mark as read", "Badge count updates in real-time. Dropdown shows notification details. Navigation works.", "Not Tested", "UAT", "Realtime + UI", "Must Have", "Connected couple, incoming message", "3.2.6"],
        ["UAT-07", "Budget management (future)", "1. Navigate to Budget\n2. Set total budget\n3. Add categories\n4. Add expenses\n5. View summary with progress bars\n6. Receive alert on over-budget", "Intuitive budget creation. Real-time totals. Clear visual progress.", "Not Tested", "UAT", "Budget + UI", "Must Have", "Budget module implemented", "US-C-009"],
        ["UAT-08", "Booking flow (future)", "1. Click 'Book' on vendor profile\n2. Select date\n3. Add requirements\n4. Submit request\n5. Receive acceptance notification\n6. Make payment\n7. View booking status", "Booking form is clear. Status updates visible. Payment flow smooth.", "Not Tested", "UAT", "Booking + Payment + UI", "Must Have", "Booking + Payment modules implemented", "3.4.2 UC9"],
        ["UAT-09", "Leave review (future)", "1. Go to completed bookings\n2. Click 'Leave Review'\n3. Select star rating\n4. Write comment\n5. Submit\n6. See confirmation", "Review form intuitive. Stars clickable. Confirmation shown.", "Not Tested", "UAT", "Review + UI", "Must Have", "Review module implemented, completed booking", "US-C-022"],
        ["UAT-10", "AI recommendations (future)", "1. Set preferences during onboarding\n2. Navigate to vendor discovery\n3. See 'Recommended for You' section\n4. Verify suggestions match preferences", "Recommendations relevant to stated preferences. AI badges visible on matched vendors.", "Not Tested", "UAT", "AI + UI", "Must Have", "AI module implemented", "US-C-006"],
    ]),
    ("UAT - Vendor Experience", [
        ["UAT-11", "Vendor registration and setup", "1. Register as vendor\n2. Verify email\n3. Login\n4. Complete profile setup wizard (name, category, description, pricing, location)\n5. Verify all saved", "Setup wizard intuitive. All fields save correctly. Map picker works.", "Not Tested", "UAT", "Auth + Vendor + UI", "Must Have", "Running frontend + backend", "US-V-001"],
        ["UAT-12", "Upload verification documents", "1. Navigate to Documents section\n2. Upload business license (PDF)\n3. Upload ID (image)\n4. See uploaded files listed\n5. Delete one and re-upload", "Upload progress shown. Files listed with names. Delete works.", "Not Tested", "UAT", "Vendor + UI", "Must Have", "Vendor with profile", "US-V-002"],
        ["UAT-13", "Submit for verification", "1. After uploading docs\n2. Click 'Submit for Verification'\n3. See status banner change to 'Pending'\n4. Wait for admin decision\n5. See approved/rejected status", "Clear status indication throughout. Rejection shows reason.", "Not Tested", "UAT", "Vendor + UI", "Must Have", "Vendor with profile + docs", "3.2.3"],
        ["UAT-14", "Vendor messaging", "1. Login as vendor\n2. Navigate to Messages\n3. See conversation list\n4. Open conversation\n5. Reply to couple\n6. Send with typing indicator", "Messages page shows all conversations. Real-time delivery works both ways.", "Not Tested", "UAT", "Realtime + UI", "Must Have", "Vendor with existing conversations", "Scenario 4"],
        ["UAT-15", "Vendor manages bookings (future)", "1. See incoming booking requests\n2. Review request details\n3. Accept one\n4. Decline another with reason\n5. View booking calendar", "Booking requests clearly displayed. Accept/decline smooth. Calendar accurate.", "Not Tested", "UAT", "Booking + UI", "Must Have", "Booking module implemented", "3.4.2 UC9"],
    ]),
    ("UAT - Admin Experience", [
        ["UAT-16", "Admin reviews vendor applications", "1. Login as admin\n2. Navigate to Vendor Management\n3. See list of pending vendors\n4. Click one to view details\n5. View uploaded documents\n6. Approve vendor", "Pending vendors clearly listed. Documents viewable inline. Approve action immediate.", "Not Tested", "UAT", "Admin + UI", "Must Have", "Admin account, pending vendors", "3.2.3"],
        ["UAT-17", "Admin rejects vendor", "1. View vendor detail\n2. Click Reject\n3. Enter reason in modal\n4. Confirm\n5. Verify status updated", "Rejection modal appears. Reason required. Status changes immediately.", "Not Tested", "UAT", "Admin + UI", "Must Have", "Admin account, pending vendor", "3.4.2 UC14"],
        ["UAT-18", "Admin suspends vendor", "1. Find verified vendor\n2. Click Suspend\n3. Confirm dialog\n4. Verify vendor hidden from public", "Suspension confirmed. Vendor no longer in public listing.", "Not Tested", "UAT", "Admin + UI", "Must Have", "Admin account, verified vendor", "4.11.3"],
        ["UAT-19", "Admin manages users (future)", "1. Navigate to Users page\n2. Search for user\n3. View details\n4. Ban user\n5. Verify user cannot login", "User management intuitive. Ban effective immediately.", "Not Tested", "UAT", "Admin + UI", "Must Have", "Admin users page implemented", "US-A-001"],
        ["UAT-20", "Admin views analytics (future)", "1. Navigate to Analytics\n2. See dashboard with charts\n3. View user registration trends\n4. View vendor stats\n5. View revenue data", "Charts render correctly. Data accurate. Filters work.", "Not Tested", "UAT", "Admin + UI", "Should Have", "Analytics module implemented", "4.6.8"],
    ]),
]

PERFORMANCE_TESTS = [
    ("Load & Response Time", [
        ["PT-01", "API response time (normal load)", "1. Run k6 with 50 virtual users for 5 minutes\n2. Measure p50, p95, p99 response times for key endpoints", "p95 < 1 second for all endpoints. p99 < 2 seconds.", "Not Tested", "Performance", "All", "Must Have", "Deployed backend with test data", "3.3.1"],
        ["PT-02", "Vendor search response time", "1. Run 100 concurrent vendor search requests with filters\n2. Measure response time", "p95 < 800ms. Database indexes used effectively.", "Not Tested", "Performance", "Vendor", "Must Have", "100+ vendors in DB", "3.3.1"],
        ["PT-03", "WebSocket message latency", "1. Send 1000 messages between 50 concurrent connections\n2. Measure delivery time", "Average latency < 200ms. No dropped messages.", "Not Tested", "Performance", "Realtime", "Should Have", "Multiple connected users", "3.3.1"],
        ["PT-04", "Concurrent user capacity", "1. Gradually ramp to 5000 virtual users\n2. Monitor error rate and response times", "System maintains < 1% error rate at 5000 users. Graceful degradation.", "Not Tested", "Performance", "All", "Should Have", "Scaled deployment", "3.3.1"],
        ["PT-05", "Database query performance", "1. Run EXPLAIN ANALYZE on complex queries (vendor search, budget sums)\n2. Check execution plans", "No sequential scans on large tables. All queries < 500ms.", "Not Tested", "Performance", "Database", "Should Have", "Production-like dataset", "3.3.6"],
        ["PT-06", "AI recommendation latency (future)", "1. Request recommendations with warm cache\n2. Request with cold cache\n3. Measure both", "Warm cache: < 100ms. Cold cache: < 800ms.", "Not Tested", "Performance", "AI", "Should Have", "AI module + Redis implemented", "3.3.1"],
        ["PT-07", "File upload throughput", "1. Upload 10 files concurrently (2-5MB each)\n2. Measure upload time and success rate", "All uploads succeed within 10 seconds. No timeouts.", "Not Tested", "Performance", "Vendor", "Should Have", "Supabase storage configured", "3.2.3"],
    ]),
    ("Stress & Endurance", [
        ["PT-08", "Extended operation test", "1. Run moderate load (500 users) for 2 hours\n2. Monitor memory, CPU, connection pools", "No memory leaks. Stable CPU. Connection pools healthy.", "Not Tested", "Performance", "All", "Should Have", "Deployed system", "3.3.2"],
        ["PT-09", "Socket.IO connection storm", "1. Rapidly connect/disconnect 500 clients within 30 seconds", "Server handles gracefully. No crashes. Memory stable after connections close.", "Not Tested", "Performance", "Realtime", "Should Have", "Deployed Socket.IO server", "3.2.6"],
        ["PT-10", "Database connection pool exhaustion", "1. Send concurrent requests exceeding pool size\n2. Monitor behavior", "Requests queue gracefully. No crashes. Errors returned after timeout.", "Not Tested", "Performance", "Database", "Should Have", "High concurrent load", "3.3.6"],
    ]),
]

SECURITY_TESTS = [
    ("Authentication & Authorization", [
        ["ST-01", "Session cookie security flags", "1. Login\n2. Inspect cookie attributes in browser DevTools", "httpOnly=true, SameSite=lax (or strict), Secure=true in production, path=/.", "Not Tested", "Security", "Auth", "Must Have", "Running application", "4.11.1"],
        ["ST-02", "Password storage verification", "1. Check database user table\n2. Verify password field is hashed", "Password stored as bcrypt/argon2 hash. Raw password never stored.", "Not Tested", "Security", "Auth", "Must Have", "Database access", "3.3.3"],
        ["ST-03", "Brute force login prevention", "1. Attempt 20 failed logins rapidly for same account", "Account temporarily locked or requests throttled after N failures.", "Not Tested", "Security", "Auth", "Should Have", "Running application", "3.3.3"],
        ["ST-04", "Session fixation prevention", "1. Note session ID before login\n2. Login\n3. Check if session ID changed", "New session generated on login. Old session invalidated.", "Not Tested", "Security", "Auth", "Must Have", "Running application", "4.11.1"],
        ["ST-05", "Privilege escalation - modify role", "1. Intercept API requests\n2. Try to modify role field in request body", "Role changes rejected. Server-side role enforcement unchanged.", "Not Tested", "Security", "Auth", "Must Have", "Authenticated session", "4.11.4"],
    ]),
    ("Input Validation & Injection", [
        ["ST-06", "SQL injection in vendor search", "1. GET /api/v1/vendors?search=' OR 1=1 --\n2. Check if all vendors returned or error", "No SQL injection. Parameterized query returns empty or filtered results.", "Not Tested", "Security", "Vendor", "Must Have", "Running application", "3.3.3"],
        ["ST-07", "SQL injection in login", "1. POST /api/auth/sign-in with email: \"' OR '1'='1\"", "Login fails. No database manipulation.", "Not Tested", "Security", "Auth", "Must Have", "Running application", "3.3.3"],
        ["ST-08", "XSS in chat messages", "1. Send message: <script>alert('XSS')</script>\n2. Check if script executes in receiver's browser", "Script not executed. Message displayed as escaped text.", "Not Tested", "Security", "Realtime", "Must Have", "Running chat", "3.3.3"],
        ["ST-09", "XSS in vendor profile", "1. Set vendor business_name to '<img src=x onerror=alert(1)>'\n2. View profile in public listing", "HTML not rendered. Displayed as plain text.", "Not Tested", "Security", "Vendor", "Must Have", "Running application", "3.3.3"],
        ["ST-10", "Path traversal in file upload", "1. Upload file with name '../../../etc/passwd'\n2. Check storage location", "File stored with sanitized name. No path traversal possible.", "Not Tested", "Security", "Vendor", "Must Have", "Running application + Supabase", "3.3.3"],
        ["ST-11", "NoSQL/JSON injection in profile update", "1. PATCH /api/v1/vendor/profile with malicious JSON payloads", "Only whitelisted fields updated. Injected fields ignored.", "Not Tested", "Security", "Vendor", "Should Have", "Running application", "3.3.3"],
    ]),
    ("Data Privacy & Access Control", [
        ["ST-12", "User cannot read others' conversations", "1. Login as user A\n2. GET /api/v1/conversations/:id (owned by user B)", "403 Forbidden. No conversation data leaked.", "Not Tested", "Security", "Realtime", "Must Have", "Two users with separate conversations", "3.3.3"],
        ["ST-13", "User cannot read others' notifications", "1. Guess notification IDs\n2. PATCH /api/v1/notifications/:id/read for another user's notification", "403 or 404. Cannot modify others' notifications.", "Not Tested", "Security", "Realtime", "Must Have", "Two users with notifications", "3.3.3"],
        ["ST-14", "Vendor cannot access other vendor's profile data", "1. Login as vendor A\n2. Attempt to GET /api/v1/vendor/profile (returns only own data)", "Only own profile data returned. No IDOR vulnerability.", "Not Tested", "Security", "Vendor", "Must Have", "Two vendor accounts", "3.3.3"],
        ["ST-15", "Admin audit logging (future)", "1. Admin performs sensitive action (impersonate, ban)\n2. Check audit log", "Action logged with admin ID, target, timestamp, action type.", "Not Tested", "Security", "Admin", "Should Have", "Admin module with audit logging", "4.11.3"],
    ]),
    ("API Security", [
        ["ST-16", "CORS policy enforcement", "1. Make request from unauthorized origin\n2. Check if blocked", "Cross-origin requests from non-whitelisted domains blocked.", "Not Tested", "Security", "All", "Must Have", "Running application", "3.3.3"],
        ["ST-17", "Request size limits", "1. Send POST with extremely large body (>10MB JSON)", "413 Payload Too Large. Server not crashed.", "Not Tested", "Security", "All", "Must Have", "Running application", "3.3.3"],
        ["ST-18", "Rate limiting on auth endpoints", "1. Send 100 login requests in 10 seconds", "Rate limited after threshold. 429 Too Many Requests.", "Not Tested", "Security", "Auth", "Should Have", "Rate limiter configured", "3.3.3"],
        ["ST-19", "HTTPS enforcement (production)", "1. Attempt HTTP connection to production server", "Redirected to HTTPS. No data sent over plain HTTP.", "Not Tested", "Security", "All", "Must Have", "Production deployment", "3.3.3"],
        ["ST-20", "Sensitive data in error responses", "1. Trigger various errors\n2. Check response bodies", "No stack traces, internal paths, or sensitive info in error responses.", "Not Tested", "Security", "All", "Must Have", "Running application", "3.3.3"],
    ]),
]


def create_test_overview(wb):
    ws = wb.active
    ws.title = "Test Plan Overview"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 14

    row = 1
    ws.cell(row=row, column=2, value="TWEDAR - Test Plan Document").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=2, value="AI-Enabled Smart Wedding Planning Platform").font = SUBTITLE_FONT
    row += 2

    ws.cell(row=row, column=2, value="HOW TO USE THIS TEST PLAN").font = BOLD_BODY
    row += 1
    instructions = [
        "1. Each tab contains test cases for a specific test type (Unit, Integration, UAT, etc.).",
        "2. Tests are assigned to either Sura or Hanamariam (see Tester column).",
        "3. Use the DROPDOWN in the 'Result' column to record: Pass / Fail / Not Implemented / Blocked / Skipped.",
        "4. Column D (Expected Result) defines the pass/fail criteria.",
        "5. Use the 'Notes' column to record observations, issues, or deviations.",
        "6. Column K (Prerequisites) shows what must exist before running the test.",
        "7. Tests marked '(Future)' require modules to be implemented first - mark as 'Not Implemented'.",
        "8. Priority column helps decide execution order (Must Have first).",
        "9. Auth tests (Unit - Auth sheet) must be completed by May 2.",
    ]
    for instr in instructions:
        ws.cell(row=row, column=2, value=instr).font = BODY_FONT
        row += 1

    row += 1
    cell = ws.cell(row=row, column=2, value="⚠️  URGENT: Auth tests must be completed by May 2!")
    cell.font = Font(name="Segoe UI", size=11, bold=True, color="FF0000")
    row += 2

    ws.cell(row=row, column=2, value="TESTER ASSIGNMENT").font = BOLD_BODY
    row += 1
    for name, fill in ASSIGNEE_STYLES.items():
        cell = ws.cell(row=row, column=2, value=name)
        cell.fill = fill
        cell.font = BOLD_BODY
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="RESULT OPTIONS (Dropdown)").font = BOLD_BODY
    row += 1
    result_options = [
        ("Pass", "Test passed - actual matches expected", PASS_FILL, PASS_FONT),
        ("Fail", "Test failed - actual differs from expected", FAIL_FILL, FAIL_FONT),
        ("Not Implemented", "Feature not yet built - cannot test", NOT_IMPL_FILL, NOT_IMPL_FONT),
        ("Blocked", "Cannot test due to dependency or environment issue", BLOCKED_FILL, BLOCKED_FONT),
        ("Skipped", "Intentionally skipped (add justification in Notes)", PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"), Font(name="Segoe UI", size=10, color="1B2A4A")),
    ]
    for label, desc, fill, font in result_options:
        cell = ws.cell(row=row, column=2, value=label)
        cell.fill = fill
        cell.font = font
        cell.border = THIN_BORDER
        ws.cell(row=row, column=3, value=desc).font = BODY_FONT
        row += 1

    row += 2
    ws.cell(row=row, column=2, value="TEST SUMMARY BY SHEET").font = BOLD_BODY
    row += 1

    summary_headers = ["Test Type", "Tester", "Total", "High Priority", "Not Tested", "Deadline"]
    for col_idx, h in enumerate(summary_headers, 2):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    row += 1

    all_test_sheets = [
        ("Unit - Auth", UNIT_AUTH_TESTS),
        ("Unit - Vendor", UNIT_VENDOR_TESTS),
        ("Unit - Admin", UNIT_ADMIN_TESTS),
        ("Unit - Realtime", UNIT_REALTIME_TESTS),
        ("Unit - Budget (Future)", UNIT_BUDGET_TESTS),
        ("Unit - Booking (Future)", UNIT_BOOKING_TESTS),
        ("Unit - Payment (Future)", UNIT_PAYMENT_TESTS),
        ("Unit - Review (Future)", UNIT_REVIEW_TESTS),
        ("Integration Tests", INTEGRATION_TESTS),
        ("User Acceptance (UAT)", UAT_TESTS),
        ("Performance Tests", PERFORMANCE_TESTS),
        ("Security Tests", SECURITY_TESTS),
    ]

    sheet_name_map = {
        "Unit - Auth": "Unit - Auth",
        "Unit - Vendor": "Unit - Vendor",
        "Unit - Admin": "Unit - Admin",
        "Unit - Realtime": "Unit - Realtime",
        "Unit - Budget (Future)": "Unit - Budget",
        "Unit - Booking (Future)": "Unit - Booking",
        "Unit - Payment (Future)": "Unit - Payment",
        "Unit - Review (Future)": "Unit - Review",
        "Integration Tests": "Integration",
        "User Acceptance (UAT)": "UAT",
        "Performance Tests": "Performance",
        "Security Tests": "Security",
    }

    grand_total = 0
    for name, sections in all_test_sheets:
        total = sum(len(items) for _, items in sections)
        high = sum(1 for _, items in sections for i in items if i[7] == "Must Have")
        grand_total += total
        mapped = sheet_name_map.get(name, name)
        tester = SHEET_ASSIGNMENTS.get(mapped, "Sura")
        deadline = "May 2" if "Auth" in name else "May 10"

        data_row = [name, tester, total, high, total, deadline]
        for col_idx, val in enumerate(data_row, 2):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left")
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL

        # Tester coloring
        tc = ws.cell(row=row, column=3)
        if tester in ASSIGNEE_STYLES:
            tc.fill = ASSIGNEE_STYLES[tester]
        tc.font = BOLD_BODY

        # Deadline urgency
        dl_cell = ws.cell(row=row, column=7)
        if "May 2" in deadline:
            dl_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            dl_cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")

        row += 1

    totals = ["TOTAL", "", grand_total, "", grand_total, ""]
    for col_idx, val in enumerate(totals, 2):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="1B2A4A")
        cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left")
        cell.border = THIN_BORDER
        cell.fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")

    row += 2
    ws.cell(row=row, column=2, value="PER-TESTER SUMMARY").font = BOLD_BODY
    row += 1
    p_headers = ["Tester", "", "Sheets Assigned", "Total Tests", "", ""]
    for col_idx, h in enumerate(p_headers, 2):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    row += 1

    for person in ["Sura", "Hanamariam"]:
        sheets_count = sum(1 for v in SHEET_ASSIGNMENTS.values() if v == person)
        test_count = sum(
            sum(len(items) for _, items in secs)
            for name, secs in all_test_sheets
            if SHEET_ASSIGNMENTS.get(sheet_name_map.get(name, name), "Sura") == person
        )
        values = [person, "", sheets_count, test_count, "", ""]
        for col_idx, val in enumerate(values, 2):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        nc = ws.cell(row=row, column=2)
        if person in ASSIGNEE_STYLES:
            nc.fill = ASSIGNEE_STYLES[person]
        nc.font = BOLD_BODY
        row += 1

    row += 2
    ws.cell(row=row, column=2, value="TEST ENVIRONMENT").font = BOLD_BODY
    row += 1
    env_info = [
        ("Backend:", "Node.js + Express + TypeScript (localhost:5000)"),
        ("Frontend:", "Next.js 16 + React 19 (localhost:3000)"),
        ("Database:", "PostgreSQL on Supabase"),
        ("Auth:", "Better Auth with Organization + Admin plugins"),
        ("Realtime:", "Socket.IO (same backend server)"),
        ("Storage:", "Supabase (vendor documents)"),
        ("Testing Type:", "Manual Testing"),
        ("Load Testing:", "k6 or Artillery (performance)"),
    ]
    for label, value in env_info:
        ws.cell(row=row, column=2, value=label).font = BOLD_BODY
        ws.cell(row=row, column=3, value=value).font = BODY_FONT
        row += 1

    row += 2
    ws.cell(row=row, column=2, value="PROJECT INFO").font = BOLD_BODY
    row += 1
    meta = [
        ("Project:", "Twedar - AI-Enabled Smart Wedding Planning Platform"),
        ("Testers:", "Sura (Auth, Vendor, Admin, Realtime, Integration, Security) | Hanamariam (Budget, Booking, Payment, Review, UAT, Performance)"),
        ("Generated:", "May 2026"),
        ("Total Test Cases:", str(grand_total)),
        ("Current Pass Rate:", "0% (No tests executed)"),
    ]
    for label, value in meta:
        ws.cell(row=row, column=2, value=label).font = BOLD_BODY
        ws.cell(row=row, column=3, value=value).font = BODY_FONT
        row += 1


def main():
    wb = Workbook()
    create_test_overview(wb)

    sheets = [
        ("Unit - Auth", UNIT_AUTH_TESTS),
        ("Unit - Vendor", UNIT_VENDOR_TESTS),
        ("Unit - Admin", UNIT_ADMIN_TESTS),
        ("Unit - Realtime", UNIT_REALTIME_TESTS),
        ("Unit - Budget", UNIT_BUDGET_TESTS),
        ("Unit - Booking", UNIT_BOOKING_TESTS),
        ("Unit - Payment", UNIT_PAYMENT_TESTS),
        ("Unit - Review", UNIT_REVIEW_TESTS),
        ("Integration", INTEGRATION_TESTS),
        ("UAT", UAT_TESTS),
        ("Performance", PERFORMANCE_TESTS),
        ("Security", SECURITY_TESTS),
    ]

    for sheet_name, data in sheets:
        ws = wb.create_sheet(title=sheet_name)
        tester = SHEET_ASSIGNMENTS.get(sheet_name, "Sura")
        populate_test_sheet(ws, data, tester)

    output_path = "/home/chera/Public/my_stuffs/astu/twedar/Twedar_Test_Plan.xlsx"
    wb.save(output_path)
    print(f"Test Plan saved to: {output_path}")
    print(f"Sheets: {wb.sheetnames}")

    total = sum(sum(len(items) for _, items in sections) for _, sections in sheets)
    print(f"\nTotal test cases: {total}")
    print("\nAssignments:")
    for sheet_name, data in sheets:
        tester = SHEET_ASSIGNMENTS.get(sheet_name, "Sura")
        count = sum(len(items) for _, items in data)
        print(f"  {sheet_name} -> {tester} ({count} tests)")

    sura_count = sum(sum(len(items) for _, items in data) for name, data in sheets if SHEET_ASSIGNMENTS.get(name, "Sura") == "Sura")
    hana_count = sum(sum(len(items) for _, items in data) for name, data in sheets if SHEET_ASSIGNMENTS.get(name, "Sura") == "Hanamariam")
    print(f"\n  Sura: {sura_count} tests")
    print(f"  Hanamariam: {hana_count} tests")


if __name__ == "__main__":
    main()
