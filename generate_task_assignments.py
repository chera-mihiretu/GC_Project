#!/usr/bin/env python3
"""
Twedar Task Assignment Generator
Generates an Excel workbook assigning feature modules to team members,
with IDs aligned to the Progress Tracker spreadsheet.
Also prints GitHub issue bodies for easy creation.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === Styles ===
HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
SUBHEADER_FONT = Font(name="Segoe UI", size=10, bold=True, color="1B2A4A")

DONE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
DONE_FONT = Font(name="Segoe UI", size=10, color="006100")
PROGRESS_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PROGRESS_FONT = Font(name="Segoe UI", size=10, color="9C5700")
NOT_STARTED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NOT_STARTED_FONT = Font(name="Segoe UI", size=10, color="9C0006")

CHERA_FILL = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
ABDI_FILL = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")
TAMIRAT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")

TITLE_FONT = Font(name="Segoe UI", size=16, bold=True, color="1B2A4A")
SUBTITLE_FONT = Font(name="Segoe UI", size=12, bold=False, color="4472C4")
BODY_FONT = Font(name="Segoe UI", size=10, color="333333")
BOLD_BODY = Font(name="Segoe UI", size=10, bold=True, color="333333")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

STATUS_STYLES = {
    "Done": (DONE_FILL, DONE_FONT),
    "In Progress": (PROGRESS_FILL, PROGRESS_FONT),
    "Not Started": (NOT_STARTED_FILL, NOT_STARTED_FONT),
}

ASSIGNEE_STYLES = {
    "Chera": CHERA_FILL,
    "Abdi": ABDI_FILL,
    "Tamirat": TAMIRAT_FILL,
}

COLUMNS = ["ID", "Feature Name", "Description", "Assignee", "Status", "Backend",
           "Frontend", "Remaining Work", "Priority", "Doc Reference"]
COL_WIDTHS = [8, 28, 50, 12, 14, 14, 14, 45, 14, 16]

# ========== FEATURE DATA (same IDs as Progress Tracker) ==========

AUTH_DATA = [
    ("Registration & Login", [
        ["A-01", "Email/Password Registration", "Users register with email and password.", "Done", "Done", "Done", "", "Must Have", "3.2.1"],
        ["A-02", "Email Verification", "System sends verification email after registration.", "Done", "Done", "Done", "", "Must Have", "3.2.1"],
        ["A-03", "Email/Password Login", "Authenticated login with session cookie.", "Done", "Done", "Done", "", "Must Have", "3.4.2 UC2"],
        ["A-04", "Google OAuth (Couples)", "Social login via Google.", "Done", "Done", "Done", "", "Must Have", "3.2.1"],
        ["A-05", "Apple OAuth (Couples)", "Social login via Apple.", "Done", "Done", "Done", "", "Should Have", "3.2.1"],
        ["A-06", "Password Reset / Forgot Password", "Email-based password reset flow.", "Done", "Done", "Done", "", "Must Have", "3.2.1"],
    ]),
    ("Session & Security", [
        ["A-07", "HTTP-Only Session Cookies", "Secure session management.", "Done", "Done", "Done", "", "Must Have", "4.11.1"],
        ["A-08", "RBAC Middleware", "Role-based access control at API level.", "Done", "Done", "Done", "", "Must Have", "4.11.4"],
        ["A-09", "Session Validation on App Load", "Check session on initial page load.", "Done", "Done", "Done", "", "Must Have", "4.11.1"],
        ["A-10", "401 Global Handling", "Redirect to login on unauthorized.", "Done", "Done", "Done", "", "Must Have", "4.11.1"],
    ]),
    ("Role Management", [
        ["A-11", "Couple Role", "Standard user with wedding planning access.", "Done", "Done", "Done", "", "Must Have", "4.11.4"],
        ["A-12", "Vendor Owner Role", "Organization owner with full vendor dashboard.", "Done", "Done", "In Progress", "Add organization member list UI, staff management.", "Must Have", "4.11.4"],
        ["A-13", "Vendor Staff Role", "Organization member with limited access.", "In Progress", "Done", "Not Started", "Build vendor staff invitation UI, restricted staff dashboard.", "Should Have", "4.11.4"],
        ["A-14", "Super Admin Role", "Full system access for platform management.", "Done", "Done", "In Progress", "Build admin users list, reports, and settings pages.", "Must Have", "4.11.4"],
        ["A-15", "Content Moderator Role", "Limited admin privileges.", "In Progress", "In Progress", "Not Started", "Add moderator role assignment UI, restricted admin views.", "Could Have", "4.11.4"],
    ]),
    ("Profile Management", [
        ["A-16", "Couple Profile Update", "Couples update personal info, wedding prefs.", "In Progress", "In Progress", "Not Started", "Build couple profile page with wedding date, theme, location fields.", "Must Have", "3.4.2 UC3"],
        ["A-17", "Vendor Profile Management", "Vendors manage business info.", "Done", "Done", "Done", "", "Must Have", "3.4.2 UC3"],
        ["A-18", "Admin Profile/Settings", "Admin can update system settings.", "Not Started", "Not Started", "Not Started", "Build admin settings page and backend config API.", "Could Have", "3.4.2 UC3"],
    ]),
]

VENDOR_DATA = [
    ("Vendor Profile & Listing", [
        ["V-01", "Vendor Registration", "Vendor registers and auto-creates organization.", "Done", "Done", "Done", "", "Must Have", "3.2.2"],
        ["V-02", "Business Profile Creation", "Vendor creates profile with business details.", "Done", "Done", "Done", "", "Must Have", "3.2.2"],
        ["V-03", "Service & Pricing Packages", "Vendor defines service types and pricing.", "Done", "Done", "Done", "", "Must Have", "3.2.2"],
        ["V-04", "Location with Map", "Vendor sets location with lat/long on map.", "Done", "Done", "Done", "", "Must Have", "3.2.2"],
        ["V-05", "Portfolio Upload", "Vendor uploads images/videos.", "In Progress", "Done", "In Progress", "Build dedicated portfolio page with media grid.", "Must Have", "3.4.2 UC11"],
        ["V-06", "Social Links", "Vendor adds social media links.", "Done", "Done", "Done", "", "Could Have", "3.2.2"],
        ["V-07", "Vendor Availability/Schedule", "Vendor sets available dates and hours.", "Not Started", "Not Started", "Not Started", "Build availability schema, API endpoints, calendar UI.", "Should Have", "3.2.2"],
    ]),
    ("Document & Verification", [
        ["V-08", "Document Upload", "Vendor uploads business documents.", "Done", "Done", "Done", "", "Must Have", "3.2.3"],
        ["V-09", "Document Delete", "Vendor removes uploaded documents.", "Done", "Done", "Done", "", "Must Have", "3.2.3"],
        ["V-10", "Submit for Verification", "Vendor submits profile for admin review.", "Done", "Done", "Done", "", "Must Have", "3.2.3"],
        ["V-11", "Verification Status Display", "Vendor sees verification status.", "Done", "Done", "Done", "", "Must Have", "3.2.3"],
        ["V-12", "Resubmission After Rejection", "Vendor fixes and resubmits.", "Done", "Done", "Done", "", "Should Have", "3.2.3"],
    ]),
    ("Admin Vendor Actions", [
        ["V-13", "List All Vendors (Admin)", "Admin views all vendors with status.", "Done", "Done", "Done", "", "Must Have", "3.4.2 UC14"],
        ["V-14", "Approve Vendor", "Admin approves vendor.", "Done", "Done", "Done", "", "Must Have", "3.4.2 UC14"],
        ["V-15", "Reject Vendor", "Admin rejects vendor with reason.", "Done", "Done", "Done", "", "Must Have", "3.4.2 UC14"],
        ["V-16", "Suspend Vendor", "Admin suspends vendor.", "Done", "Done", "Done", "", "Must Have", "4.11.3"],
        ["V-17", "Reinstate Vendor", "Admin reactivates suspended vendor.", "Done", "Done", "Done", "", "Should Have", "4.11.3"],
        ["V-18", "Deactivate Vendor", "Admin permanently deactivates vendor.", "Done", "Done", "Done", "", "Should Have", "4.11.3"],
    ]),
    ("Public Vendor Discovery", [
        ["V-19", "Public Vendor Listing", "Couples browse verified vendors.", "Done", "Done", "Done", "", "Must Have", "3.2.2"],
        ["V-20", "Vendor Detail Page", "Couples view full vendor profile.", "Done", "Done", "Done", "", "Must Have", "3.4.2 UC8"],
        ["V-21", "Vendor Filtering by Category", "Filter by service type.", "Done", "Done", "Done", "", "Must Have", "US-C-004"],
        ["V-22", "Vendor Filtering by Location", "Filter by city/area.", "Done", "Done", "Done", "", "Must Have", "US-C-004"],
        ["V-23", "Vendor Filtering by Verification", "Show only verified vendors.", "Done", "Done", "Done", "", "Must Have", "US-C-007"],
        ["V-24", "Save Vendors to Favorites", "Couple saves vendors for later.", "Not Started", "Not Started", "Not Started", "Build favorites table, API, heart button in vendor cards.", "Could Have", "US-C-008"],
    ]),
]

REALTIME_DATA = [
    ("Socket.IO Infrastructure", [
        ["R-01", "Socket.IO Server Setup", "WebSocket server with auth.", "Done", "Done", "Done", "", "Must Have", "3.2.6"],
        ["R-02", "Socket.IO Client Connection", "Frontend authenticated WS connection.", "Done", "Done", "Done", "", "Must Have", "3.2.6"],
        ["R-03", "Room-based Architecture", "Users join/leave rooms.", "Done", "Done", "Done", "", "Must Have", "3.2.6"],
    ]),
    ("Chat / Messaging", [
        ["R-04", "Create Conversation", "Couple initiates conversation with vendor.", "Done", "Done", "Done", "", "Must Have", "3.2.6"],
        ["R-05", "List Conversations", "User sees all conversations.", "Done", "Done", "Done", "", "Must Have", "3.2.6"],
        ["R-06", "Send Text Message", "Real-time text message delivery.", "Done", "Done", "Done", "", "Must Have", "3.2.6"],
        ["R-07", "Message History", "Load previous messages.", "Done", "Done", "Done", "", "Must Have", "3.2.6"],
        ["R-08", "Typing Indicators", "Show when other user is typing.", "Done", "Done", "Done", "", "Should Have", "Scenario 4"],
        ["R-09", "Read Receipts", "Mark messages as read.", "Done", "Done", "Done", "", "Should Have", "Scenario 4"],
        ["R-10", "File Attachments in Chat", "Send files in chat.", "Not Started", "Not Started", "Not Started", "Add file upload for chat, message type field, file preview.", "Should Have", "US-C-017"],
    ]),
    ("Notifications", [
        ["R-11", "In-App Notifications", "Real-time notification delivery.", "Done", "Done", "Done", "", "Must Have", "3.2.6"],
        ["R-12", "Notification REST API", "Fetch and mark notifications.", "Done", "Done", "Done", "", "Must Have", "3.2.6"],
        ["R-13", "Notification Bell/Badge", "Visual unread indicator.", "Done", "Done", "Done", "", "Must Have", "3.2.6"],
        ["R-14", "Push Notifications", "Notifications when app is closed.", "Not Started", "Not Started", "Not Started", "Integrate Web Push API / FCM for background notifications.", "Should Have", "US-C-018"],
        ["R-15", "Email Notifications", "Email alerts for critical events.", "Not Started", "Not Started", "Not Started", "Extend email service for booking, payment, reminders.", "Could Have", "3.2.6"],
    ]),
    ("Presence", [
        ["R-16", "User Online/Offline Status", "Show if user is online.", "Done", "Done", "Done", "", "Could Have", "3.2.6"],
    ]),
]

BUDGET_DATA = [
    ("Budget Planning", [
        ["B-01", "Create Wedding Budget", "Couple defines total budget.", "Not Started", "Not Started", "Not Started", "Create budget table, API, budget creation UI.", "Must Have", "3.2.5"],
        ["B-02", "Budget Category Allocation", "Allocate budget into categories.", "Not Started", "Not Started", "Not Started", "Create budget_categories table, CRUD API, allocation UI.", "Must Have", "3.2.5"],
        ["B-03", "AI-Suggested Budget Allocation", "System suggests category distribution.", "Not Started", "Not Started", "Not Started", "Build recommendation logic, API, suggestion UI.", "Should Have", "3.2.5"],
    ]),
    ("Expense Tracking", [
        ["B-04", "Add Expense Entry", "Record expenses against categories.", "Not Started", "Not Started", "Not Started", "Create expenses table, API, expense form UI.", "Must Have", "3.2.5"],
        ["B-05", "Real-Time Budget Totals", "Live spending vs allocated amounts.", "Not Started", "Not Started", "Not Started", "Build budget summary calculations, progress bars UI.", "Must Have", "3.2.5"],
        ["B-06", "Budget Alerts & Limits", "Warnings when spending approaches limits.", "Not Started", "Not Started", "Not Started", "Add threshold logic, notification triggers, alert banners.", "Should Have", "3.2.5"],
        ["B-07", "Receipt Upload", "Attach receipt images to expenses.", "Not Started", "Not Started", "Not Started", "Add receipt storage, upload API, receipt preview.", "Could Have", "US-C-013"],
        ["B-08", "Budget Summary Reports", "Generate exportable reports.", "Not Started", "Not Started", "Not Started", "Build report API, PDF/CSV export, report page.", "Could Have", "US-C-014"],
    ]),
]

AI_DATA = [
    ("Recommendation Engine", [
        ["AI-01", "Content-Based Filtering", "Recommend vendors using attributes.", "Not Started", "Not Started", "Not Started", "Build vendor vectorization, similarity scoring, API.", "Must Have", "3.2.4"],
        ["AI-02", "Collaborative Filtering", "Recommend based on similar couples.", "Not Started", "Not Started", "Not Started", "Build interaction tracking, matrix factorization.", "Should Have", "3.2.4"],
        ["AI-03", "Hybrid Weighted Approach", "Combine content + collaborative.", "Not Started", "Not Started", "Not Started", "Implement weighted scoring with adaptive weights.", "Should Have", "3.2.4"],
        ["AI-04", "Cold-Start Handling", "Meaningful recommendations with zero data.", "Not Started", "Not Started", "Not Started", "Default to content-based, use metadata + preferences.", "Must Have", "3.2.4"],
        ["AI-05", "User Preference Collection", "Gather couple preferences.", "Not Started", "Not Started", "Not Started", "Build preference wizard, user_preferences table.", "Must Have", "3.2.4"],
    ]),
    ("Caching & Performance", [
        ["AI-06", "Redis Recommendation Cache", "Cache recommendations for 24h.", "Not Started", "Not Started", "Not Started", "Add Redis, cache layer, TTL-based invalidation.", "Must Have", "4.6.3"],
        ["AI-07", "Cache Invalidation Rules", "Invalidate on preference/booking/profile change.", "Not Started", "Not Started", "Not Started", "Implement cache key strategy, event-driven invalidation.", "Should Have", "4.6.3"],
        ["AI-08", "Sub-2-Second Response", "Results within 2 seconds.", "Not Started", "Not Started", "Not Started", "Optimize queries, pre-compute scores, use Redis.", "Must Have", "3.3.1"],
    ]),
    ("Recommendation UI", [
        ["AI-09", "Recommendation Badges", "Highlight top-matched vendors.", "Not Started", "Not Started", "Not Started", "Add badge component, match score display.", "Should Have", "Scenario 5"],
        ["AI-10", "Personalized Vendor Search", "AI-ranked search results.", "Not Started", "Not Started", "Not Started", "Integrate AI scoring into search pipeline.", "Must Have", "Scenario 5"],
    ]),
]

BOOKING_DATA = [
    ("Booking Workflow", [
        ["BK-01", "Send Booking Request", "Couple submits booking request.", "Not Started", "Not Started", "Not Started", "Create bookings table, API, booking request form.", "Must Have", "3.4.2 UC9"],
        ["BK-02", "Vendor Reviews Booking", "Vendor sees incoming requests.", "Not Started", "Not Started", "Not Started", "Build vendor bookings list, incoming requests page.", "Must Have", "3.4.2 UC9"],
        ["BK-03", "Accept/Decline Booking", "Vendor accepts or declines.", "Not Started", "Not Started", "Not Started", "Add booking status machine, status update API.", "Must Have", "3.4.2 UC9"],
        ["BK-04", "Booking Status Tracking", "Both parties see booking status.", "Not Started", "Not Started", "Not Started", "Build booking detail page with status timeline.", "Must Have", "3.4.2 UC9"],
        ["BK-05", "Booking Confirmation Notification", "Notify couple on accept/decline.", "Not Started", "Not Started", "Not Started", "Wire booking status changes to notification service.", "Must Have", "US-C-018"],
    ]),
    ("Scheduling & Calendar", [
        ["BK-06", "Vendor Availability Calendar", "Vendors set available/blocked dates.", "Not Started", "Not Started", "Not Started", "Build availability_slots table, calendar API and UI.", "Should Have", "3.2.2"],
        ["BK-07", "Date Conflict Detection", "Prevent double-booking.", "Not Started", "Not Started", "Not Started", "Add unique constraint on vendor+date, validation.", "Should Have", "3.2.2"],
        ["BK-08", "Booking History", "Past and upcoming bookings list.", "Not Started", "Not Started", "Not Started", "Build booking list API, history pages.", "Must Have", "US-C-021"],
    ]),
]

PAYMENT_DATA = [
    ("Payment Integration", [
        ["P-01", "Chapa Payment Gateway", "Integrate Chapa API.", "Not Started", "Not Started", "Not Started", "Add Chapa SDK, payment API, webhook handler.", "Must Have", "3.2.7"],
        ["P-02", "Telebirr Integration", "Support Telebirr mobile money.", "Not Started", "Not Started", "Not Started", "Research Telebirr API, implement adapter.", "Should Have", "3.2.7"],
        ["P-03", "CBE Birr Integration", "Support CBE Birr payments.", "Not Started", "Not Started", "Not Started", "Research CBE Birr API, implement adapter.", "Could Have", "3.2.7"],
    ]),
    ("Payment Workflow", [
        ["P-04", "Advance Deposit Payment", "Couple makes deposit to secure booking.", "Not Started", "Not Started", "Not Started", "Build deposit flow with gateway redirect + webhook.", "Must Have", "3.2.7"],
        ["P-05", "Final Payment", "Complete remaining payment.", "Not Started", "Not Started", "Not Started", "Build final payment trigger, amount calculation.", "Must Have", "3.2.7"],
        ["P-06", "Escrow-Style Workflow", "Hold funds until service delivered.", "Not Started", "Not Started", "Not Started", "Design escrow state machine, hold/release/refund.", "Should Have", "3.2.7"],
        ["P-07", "Payment History", "View all transactions.", "Not Started", "Not Started", "Not Started", "Build payments table, transaction history API and page.", "Must Have", "3.2.7"],
        ["P-08", "Payment-Expense Linkage", "Auto-record payments as budget expenses.", "Not Started", "Not Started", "Not Started", "Link payment confirmation to expense creation.", "Should Have", "3.2.7"],
    ]),
]

REVIEW_DATA = [
    ("Review Submission", [
        ["RV-01", "Submit Rating & Review", "Couple rates vendor with stars and text.", "Not Started", "Not Started", "Not Started", "Create reviews table, API, review form UI.", "Must Have", "3.2.8"],
        ["RV-02", "One Review Per Booking", "Prevent duplicate reviews.", "Not Started", "Not Started", "Not Started", "Add unique constraint on booking_id.", "Must Have", "3.2.8"],
        ["RV-03", "Photo Upload with Review", "Couple uploads photos as feedback.", "Not Started", "Not Started", "Not Started", "Add review_photos table, image upload, gallery.", "Could Have", "US-C-023"],
    ]),
    ("Review Management", [
        ["RV-04", "Vendor Rating Aggregation", "Auto-calculate average rating.", "Not Started", "Not Started", "Not Started", "Build trigger to recalculate avg on new review.", "Must Have", "3.2.8"],
        ["RV-05", "Admin Review Moderation", "Admin approves/removes reviews.", "Not Started", "Not Started", "Not Started", "Add isApproved field, admin queue API, moderation UI.", "Must Have", "3.4.2 UC15"],
        ["RV-06", "Review Display on Vendor Profile", "Show reviews on public vendor page.", "Not Started", "Not Started", "Not Started", "Fetch reviews, display review cards with stars.", "Must Have", "3.2.8"],
        ["RV-07", "Review Feeds AI Recommendations", "Use review data for collaborative filtering.", "Not Started", "Not Started", "Not Started", "Include review scores in recommendation matrix.", "Should Have", "3.2.4"],
        ["RV-08", "Vendor Notification on New Review", "Notify vendor on new review.", "Not Started", "Not Started", "Not Started", "Wire review creation to notification service.", "Should Have", "Scenario 3"],
    ]),
]

ADMIN_DATA = [
    ("User Management", [
        ["AD-01", "View All Users", "Admin sees all registered users.", "In Progress", "In Progress", "Not Started", "Build admin users list page with filters.", "Must Have", "US-A-001"],
        ["AD-02", "Suspend/Ban User", "Admin suspends or bans accounts.", "In Progress", "Done", "Not Started", "Build user action buttons, confirmation dialogs.", "Must Have", "US-A-002"],
        ["AD-03", "Reactivate User", "Admin reactivates suspended accounts.", "In Progress", "Done", "Not Started", "Add reactivate button, confirmation flow.", "Should Have", "US-A-003"],
        ["AD-04", "Impersonate User Session", "Admin views platform as specific user.", "In Progress", "Done", "Not Started", "Build impersonation button, exit-impersonation banner.", "Could Have", "US-A-004"],
    ]),
    ("Vendor Verification (Admin)", [
        ["AD-05", "Vendor Verification Dashboard", "Dedicated vendor review panel.", "Done", "Done", "Done", "", "Must Have", "3.2.3"],
        ["AD-06", "Document Review Interface", "View uploaded vendor docs inline.", "Done", "Done", "Done", "", "Must Have", "3.2.3"],
    ]),
    ("Analytics & Reporting", [
        ["AD-07", "Platform Analytics Dashboard", "User counts, vendor stats, booking volume.", "Not Started", "Not Started", "Not Started", "Build analytics queries, stats API, charts.", "Should Have", "4.6.8"],
        ["AD-08", "Revenue Reports", "Track platform revenue.", "Not Started", "Not Started", "Not Started", "Build revenue aggregation, report generation.", "Could Have", "4.6.8"],
        ["AD-09", "User Activity Reports", "Monitor engagement, trends.", "Not Started", "Not Started", "Not Started", "Build activity logging, trend calculations.", "Could Have", "4.6.8"],
        ["AD-10", "System Health Monitoring", "Uptime, response times, error rates.", "Not Started", "Not Started", "Not Started", "Integrate monitoring, build status page.", "Should Have", "3.3.2"],
    ]),
    ("Content Moderation", [
        ["AD-11", "Review Moderation Queue", "Admin reviews user reviews.", "Not Started", "Not Started", "Not Started", "Build moderation queue, approve/reject actions.", "Must Have", "3.4.2 UC15"],
        ["AD-12", "Content Moderator Role", "Assign limited admin privileges.", "Not Started", "Not Started", "Not Started", "Add role assignment UI, moderator permissions.", "Could Have", "4.11.3"],
    ]),
]

COUPLE_DATA = [
    ("Wedding Project", [
        ["C-01", "Create Wedding Plan", "Initialize wedding with date, location, theme.", "Not Started", "Not Started", "Not Started", "Create wedding_projects table, CRUD API, setup wizard.", "Must Have", "3.4.2 UC4"],
        ["C-02", "Wedding Profile Customization", "Set and update wedding preferences.", "Not Started", "Not Started", "Not Started", "Build preferences page with theme, style, location.", "Must Have", "US-C-003"],
        ["C-03", "Wedding Dashboard", "Central overview of planning progress.", "In Progress", "In Progress", "In Progress", "Connect dashboard stats to real data.", "Must Have", "3.4.1"],
    ]),
    ("Checklist & Timeline", [
        ["C-04", "Wedding Checklist", "Pre-built and custom task checklist.", "Not Started", "Not Started", "Not Started", "Build checklist_items table, CRUD API, checklist page.", "Must Have", "1.5.1"],
        ["C-05", "Task Deadlines & Reminders", "Due dates with reminders.", "Not Started", "Not Started", "Not Started", "Add due_date field, reminder triggers, overdue highlighting.", "Should Have", "1.5.1"],
        ["C-06", "Planning Timeline/Milestones", "Visual timeline of milestones.", "Not Started", "Not Started", "Not Started", "Build milestone model, timeline visualization.", "Could Have", "1.5.1"],
    ]),
    ("Guest Management", [
        ["C-07", "Guest List", "Manage invited guests.", "Not Started", "Not Started", "Not Started", "Build guests table, CRUD API, guest list page.", "Must Have", "1.5.1"],
        ["C-08", "RSVP Tracking", "Track guest responses.", "Not Started", "Not Started", "Not Started", "Add RSVP status, guest-facing RSVP page.", "Should Have", "1.5.1"],
        ["C-09", "Guest Count & Seating", "Track guests and seating.", "Not Started", "Not Started", "Not Started", "Build seating plan model, table assignment UI.", "Could Have", "1.5.1"],
    ]),
    ("Discovery & Favorites", [
        ["C-10", "Vendor Favorites List", "Save vendors to compare.", "Not Started", "Not Started", "Not Started", "Build favorites table, toggle API, favorites page.", "Could Have", "US-C-008"],
        ["C-11", "Vendor Comparison", "Side-by-side comparison.", "Not Started", "Not Started", "Not Started", "Build comparison view with key metrics.", "Could Have", "US-C-005"],
    ]),
]

NONFUNCTIONAL_DATA = [
    ("Performance", [
        ["NF-01", "Sub-2-Second Response Time", "95% of requests respond within 1s.", "In Progress", "In Progress", "In Progress", "Add monitoring, load testing, optimize queries.", "Must Have", "3.3.1"],
        ["NF-02", "5000 Concurrent Users", "System supports 5000+ users.", "Not Started", "Not Started", "Not Started", "Load test, configure scaling, optimize WebSocket.", "Should Have", "3.3.1"],
        ["NF-03", "Recommendation < 800ms", "AI recommendations within 800ms.", "Not Started", "Not Started", "Not Started", "Depends on Redis caching once AI is built.", "Should Have", "3.3.1"],
    ]),
    ("Security", [
        ["NF-04", "HTTPS/TLS Encryption", "All communication encrypted.", "In Progress", "In Progress", "In Progress", "Configure TLS for production.", "Must Have", "3.3.3"],
        ["NF-05", "Data Encryption at Rest", "Sensitive fields encrypted.", "Not Started", "Not Started", "Not Started", "Identify PII, implement encryption layer.", "Should Have", "3.3.3"],
        ["NF-06", "Input Validation (OWASP)", "All inputs validated.", "In Progress", "In Progress", "In Progress", "Add Zod/Joi validation on all endpoints.", "Must Have", "3.3.3"],
        ["NF-07", "SQL Injection Prevention", "Parameterized queries throughout.", "Done", "Done", "Done", "", "Must Have", "3.3.3"],
    ]),
    ("Reliability & Availability", [
        ["NF-08", "99.5% Monthly Uptime", "System available 99.5%.", "Not Started", "Not Started", "Not Started", "Configure health checks, monitoring alerts.", "Should Have", "3.3.2"],
        ["NF-09", "Daily Automated Backups", "PostgreSQL backup RPO <= 24h.", "Not Started", "Not Started", "Not Started", "Configure automated backups, verify PITR.", "Should Have", "3.3.2"],
        ["NF-10", "Disaster Recovery (RTO <= 4h)", "Recover within 4 hours.", "Not Started", "Not Started", "Not Started", "Document recovery, test backup restoration.", "Should Have", "3.3.2"],
    ]),
    ("Scalability & Infrastructure", [
        ["NF-11", "Horizontal Backend Scaling", "Backend supports multiple instances.", "In Progress", "In Progress", "Not Started", "Add Socket.IO Redis adapter, load balancer.", "Should Have", "3.3.6"],
        ["NF-12", "Database Indexing & Caching", "Efficient queries with indexes.", "In Progress", "In Progress", "Not Started", "Add Redis for frequent data, optimize queries.", "Should Have", "3.3.6"],
        ["NF-13", "Cloud Deployment", "Deploy to AWS/GCP/Vercel.", "Not Started", "Not Started", "Not Started", "Configure CI/CD pipeline, deploy.", "Should Have", "5.5"],
    ]),
    ("Mobile & Accessibility", [
        ["NF-14", "Flutter Mobile App", "Cross-platform mobile app.", "Not Started", "Not Started", "Not Started", "Initialize Flutter project, implement core screens.", "Should Have", "1.11"],
        ["NF-15", "Responsive Web UI", "Works on mobile and tablets.", "In Progress", "Done", "In Progress", "Test all pages on mobile, fix responsive issues.", "Must Have", "3.3.4"],
        ["NF-16", "Accessibility Standards", "WCAG compliance.", "Not Started", "Not Started", "Not Started", "Add ARIA labels, keyboard navigation, screen reader.", "Should Have", "3.3.4"],
    ]),
]

# ========== ASSIGNMENT MAPPING ==========

ASSIGNMENTS = {
    "1. Auth & User Mgmt": ("Chera", AUTH_DATA),
    "2. Vendor Management": ("Chera", VENDOR_DATA),
    "3. Real-Time Comms": ("Chera", REALTIME_DATA),
    "4. Budgeting & Expense": ("Abdi", BUDGET_DATA),
    "5. AI Recommendations": ("Tamirat", AI_DATA),
    "6. Booking & Scheduling": ("Abdi", BOOKING_DATA),
    "7. Payment Integration": ("Tamirat", PAYMENT_DATA),
    "8. Review & Feedback": ("Tamirat", REVIEW_DATA),
    "9. Admin & Analytics": ("Chera", ADMIN_DATA),
    "10. Couple Planning": ("Abdi", COUPLE_DATA),
    "11. Non-Functional": ("Chera", NONFUNCTIONAL_DATA),
}


def style_cell(cell, fill=None, font=None, align=None, border=THIN_BORDER):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border


def write_feature_sheet(wb, sheet_name, assignee, sections):
    ws = wb.create_sheet(title=sheet_name)

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col_idx, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        style_cell(cell, HEADER_FILL, HEADER_FONT,
                   Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.freeze_panes = "A2"

    row = 2
    for section_title, items in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
        cell = ws.cell(row=row, column=1, value=section_title)
        style_cell(cell, SUBHEADER_FILL, SUBHEADER_FONT, Alignment(vertical="center"))
        for col in range(2, len(COLUMNS) + 1):
            style_cell(ws.cell(row=row, column=col), fill=SUBHEADER_FILL)
        row += 1

        for item in items:
            task_id, name, desc, status, be, fe, remaining, priority, doc_ref = item
            values = [task_id, name, desc, assignee, status, be, fe, remaining, priority, doc_ref]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = THIN_BORDER

            if row % 2 == 0:
                for col_idx in range(1, len(COLUMNS) + 1):
                    ws.cell(row=row, column=col_idx).fill = ALT_ROW_FILL

            # Status coloring
            for col_idx, val in [(5, status), (6, be), (7, fe)]:
                if val in STATUS_STYLES:
                    c = ws.cell(row=row, column=col_idx)
                    fill, font = STATUS_STYLES[val]
                    c.fill = fill
                    c.font = font
                    c.alignment = Alignment(horizontal="center", vertical="center")

            # Assignee coloring
            assignee_cell = ws.cell(row=row, column=4)
            if assignee in ASSIGNEE_STYLES:
                assignee_cell.fill = ASSIGNEE_STYLES[assignee]
            assignee_cell.font = Font(name="Segoe UI", size=10, bold=True, color="333333")
            assignee_cell.alignment = Alignment(horizontal="center", vertical="center")

            row += 1
        row += 1


def create_overview(wb):
    ws = wb.active
    ws.title = "Overview"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 16

    row = 1
    ws.cell(row=row, column=2, value="TWEDAR - Task Assignment Sheet").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=2, value="Feature Module Assignments for Chera, Abdi & Tamirat").font = SUBTITLE_FONT
    row += 2

    ws.cell(row=row, column=2, value="ASSIGNEE LEGEND").font = BOLD_BODY
    row += 1
    for name, fill in ASSIGNEE_STYLES.items():
        cell = ws.cell(row=row, column=2, value=name)
        cell.fill = fill
        cell.font = BOLD_BODY
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="STATUS LEGEND").font = BOLD_BODY
    row += 1
    for label, (fill, font) in STATUS_STYLES.items():
        cell = ws.cell(row=row, column=2, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        row += 1

    row += 2
    ws.cell(row=row, column=2, value="MODULE ASSIGNMENTS").font = BOLD_BODY
    row += 1

    headers = ["Module", "Assignee", "Total", "Done", "In Progress", "Not Started", "Completion %"]
    for col_idx, h in enumerate(headers, 2):
        cell = ws.cell(row=row, column=col_idx, value=h)
        style_cell(cell, HEADER_FILL, HEADER_FONT, Alignment(horizontal="center"))
    row += 1

    person_totals = {"Chera": [0, 0, 0, 0], "Abdi": [0, 0, 0, 0], "Tamirat": [0, 0, 0, 0]}
    grand = [0, 0, 0, 0]

    for sheet_name, (assignee, sections) in ASSIGNMENTS.items():
        total = sum(len(items) for _, items in sections)
        done = sum(1 for _, items in sections for i in items if i[3] == "Done")
        progress = sum(1 for _, items in sections for i in items if i[3] == "In Progress")
        not_started = sum(1 for _, items in sections for i in items if i[3] == "Not Started")
        pct = round((done / total) * 100) if total > 0 else 0

        person_totals[assignee][0] += total
        person_totals[assignee][1] += done
        person_totals[assignee][2] += progress
        person_totals[assignee][3] += not_started
        grand[0] += total
        grand[1] += done
        grand[2] += progress
        grand[3] += not_started

        values = [sheet_name, assignee, total, done, progress, not_started, f"{pct}%"]
        for col_idx, val in enumerate(values, 2):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left")
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL

        # Assignee color
        ac = ws.cell(row=row, column=3)
        if assignee in ASSIGNEE_STYLES:
            ac.fill = ASSIGNEE_STYLES[assignee]
        ac.font = BOLD_BODY
        ac.alignment = Alignment(horizontal="center")

        # Completion color
        pct_cell = ws.cell(row=row, column=8)
        if pct >= 80:
            pct_cell.fill = DONE_FILL
            pct_cell.font = DONE_FONT
        elif pct >= 30:
            pct_cell.fill = PROGRESS_FILL
            pct_cell.font = PROGRESS_FONT
        else:
            pct_cell.fill = NOT_STARTED_FILL
            pct_cell.font = NOT_STARTED_FONT

        row += 1

    # Grand total
    grand_pct = round((grand[1] / grand[0]) * 100) if grand[0] > 0 else 0
    totals = ["TOTAL", "", grand[0], grand[1], grand[2], grand[3], f"{grand_pct}%"]
    for col_idx, val in enumerate(totals, 2):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="1B2A4A")
        cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left")
        cell.border = THIN_BORDER
        cell.fill = TOTAL_FILL

    row += 2
    ws.cell(row=row, column=2, value="PER-PERSON SUMMARY").font = BOLD_BODY
    row += 1

    p_headers = ["Team Member", "", "Total Tasks", "Done", "In Progress", "Not Started", "Completion %"]
    for col_idx, h in enumerate(p_headers, 2):
        cell = ws.cell(row=row, column=col_idx, value=h)
        style_cell(cell, HEADER_FILL, HEADER_FONT, Alignment(horizontal="center"))
    row += 1

    for name, (total, done, prog, ns) in person_totals.items():
        pct = round((done / total) * 100) if total > 0 else 0
        values = [name, "", total, done, prog, ns, f"{pct}%"]
        for col_idx, val in enumerate(values, 2):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        nc = ws.cell(row=row, column=2)
        if name in ASSIGNEE_STYLES:
            nc.fill = ASSIGNEE_STYLES[name]
        nc.font = BOLD_BODY

        pct_cell = ws.cell(row=row, column=8)
        if pct >= 80:
            pct_cell.fill = DONE_FILL
            pct_cell.font = DONE_FONT
        elif pct >= 30:
            pct_cell.fill = PROGRESS_FILL
            pct_cell.font = PROGRESS_FONT
        else:
            pct_cell.fill = NOT_STARTED_FILL
            pct_cell.font = NOT_STARTED_FONT

        row += 1


def main():
    wb = Workbook()
    create_overview(wb)

    for sheet_name, (assignee, sections) in ASSIGNMENTS.items():
        write_feature_sheet(wb, sheet_name, assignee, sections)

    output_path = "/home/chera/Public/my_stuffs/astu/twedar/Twedar_Task_Assignments.xlsx"
    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Sheets: {wb.sheetnames}")

    for name, (assignee, sections) in ASSIGNMENTS.items():
        total = sum(len(items) for _, items in sections)
        remaining = sum(1 for _, items in sections for i in items if i[3] != "Done")
        print(f"  {name} -> {assignee} ({total} tasks, {remaining} remaining)")


if __name__ == "__main__":
    main()
